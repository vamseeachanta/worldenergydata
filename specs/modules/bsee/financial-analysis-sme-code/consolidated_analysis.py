#!/usr/bin/env python3
"""
Consolidated Financial Analysis Script
Based on SME Roy's V20 implementation
Unified approach for BSEE financial analysis
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple

# Version tracking
VERSION = "Consolidated_V1.0"
BASED_ON = "SME_Roy_V20"

# Standard lease grouping map
LEASE_GROUP_MAP = {
    'Stones': 'Stones',
    'Cascade': 'Cascade Chinook', 
    'Chinook': 'Cascade Chinook',
    'Julia': 'Julia',
    'Anchor': 'Anchor',
    'Jack': 'Jack',
    'St Malo': 'St Malo',
    'Kaskida': 'Kaskida',
    'Tiber': 'Tiber',
    'Shenandoah': 'Shenandoah',
    'North Platte': 'North Platte',
    'Big Foot': 'Big Foot',
}

class FinancialAnalyzer:
    """Main orchestrator for BSEE financial analysis"""
    
    def __init__(self, config_path: Optional[str] = None):
        """Initialize with optional configuration file"""
        self.config = self._load_config(config_path)
        self.lease_groups = LEASE_GROUP_MAP.copy()
        
    def _load_config(self, config_path: Optional[str]) -> dict:
        """Load configuration from YAML or use defaults"""
        default_config = {
            'discount_rate': 0.10,  # 10% annual
            'tax_rate': 0.35,       # 35% combined
            'royalty_rate': 0.1875, # 18.75% standard
            'opex_per_bbl': 10.0,   # $10/bbl
            'gap_threshold_days': 300,
        }
        
        if config_path and Path(config_path).exists():
            # Load from YAML if provided
            import yaml
            with open(config_path, 'r') as f:
                user_config = yaml.safe_load(f)
                default_config.update(user_config)
                
        return default_config
    
    def analyze(self, 
                leases_file: str,
                production_file: str,
                drilling_completion_file: str,
                oil_price_file: str,
                assumptions_file: Optional[str] = None) -> Dict:
        """
        Run complete financial analysis
        
        Args:
            leases_file: Path to leases configuration Excel
            production_file: Path to production data Excel  
            drilling_completion_file: Path to D&C data Excel
            oil_price_file: Path to WTI price data Excel
            assumptions_file: Optional path to lease assumptions Excel
            
        Returns:
            Dictionary with analysis results
        """
        
        # Load all input data
        leases_df = self._load_leases(leases_file)
        production_df = self._load_production(production_file)
        dc_data = self._load_drilling_completion(drilling_completion_file)
        oil_prices = self._load_oil_prices(oil_price_file)
        
        if assumptions_file:
            assumptions = self._load_assumptions(assumptions_file)
        else:
            assumptions = self._generate_default_assumptions(leases_df)
        
        # Process each lease group
        results = {}
        for lease_name in leases_df['LEASE_NAME'].unique():
            if pd.isna(lease_name):
                continue
                
            # Get lease group
            group = self.lease_groups.get(lease_name, lease_name)
            
            # Calculate financials
            lease_results = self._calculate_lease_financials(
                lease_name=lease_name,
                group_name=group,
                production=production_df,
                dc_data=dc_data,
                oil_prices=oil_prices,
                assumptions=assumptions
            )
            
            results[lease_name] = lease_results
            
        # Generate summary metrics
        summary = self._generate_summary(results)
        
        return {
            'lease_results': results,
            'summary': summary,
            'config': self.config,
            'version': VERSION
        }
    
    def _load_leases(self, filepath: str) -> pd.DataFrame:
        """Load and process lease configuration"""
        df = pd.read_excel(filepath)
        # Standardize column names
        df.columns = [col.strip().upper() for col in df.columns]
        return df
    
    def _load_production(self, filepath: str) -> pd.DataFrame:
        """Load production data, handle matrix or timeseries format"""
        xls = pd.ExcelFile(filepath)
        all_production = {}
        
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            
            # Check if matrix format (wide) or timeseries (long)
            if self._is_matrix_format(df):
                df = self._convert_matrix_to_timeseries(df)
            
            all_production[sheet_name] = df
            
        return all_production
    
    def _is_matrix_format(self, df: pd.DataFrame) -> bool:
        """Check if DataFrame is in matrix format"""
        # Matrix format has well names and year-month columns
        cols = df.columns.astype(str)
        has_well = any('WELL' in col.upper() for col in cols)
        has_months = any('-' in col for col in cols)
        return has_well and has_months
    
    def _convert_matrix_to_timeseries(self, df: pd.DataFrame) -> pd.DataFrame:
        """Convert matrix format to timeseries"""
        # Implementation based on V20 matrix_to_timeseries
        well_col = [c for c in df.columns if 'WELL' in c.upper()][0]
        month_cols = [c for c in df.columns if '-' in str(c)]
        
        # Melt to long format
        melted = df.melt(
            id_vars=[well_col],
            value_vars=month_cols,
            var_name='Month',
            value_name='Production'
        )
        
        melted['Month'] = pd.to_datetime(melted['Month'])
        melted['Production'] = pd.to_numeric(melted['Production'], errors='coerce').fillna(0)
        
        return melted
    
    def _load_drilling_completion(self, filepath: str) -> Dict:
        """Load drilling and completion data"""
        dc_data = {}
        xls = pd.ExcelFile(filepath)
        
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            dc_data[sheet] = df
            
        return dc_data
    
    def _load_oil_prices(self, filepath: str) -> pd.DataFrame:
        """Load WTI oil price data"""
        df = pd.read_excel(filepath)
        df['Date'] = pd.to_datetime(df.iloc[:, 0])
        df['Price'] = pd.to_numeric(df.iloc[:, 1], errors='coerce')
        df = df.set_index('Date')['Price']
        return df
    
    def _load_assumptions(self, filepath: str) -> pd.DataFrame:
        """Load lease-specific assumptions"""
        return pd.read_excel(filepath)
    
    def _generate_default_assumptions(self, leases_df: pd.DataFrame) -> pd.DataFrame:
        """Generate default assumptions if not provided"""
        assumptions = []
        for lease in leases_df['LEASE_NAME'].unique():
            if pd.isna(lease):
                continue
            assumptions.append({
                'LEASE_NAME': lease,
                'DISCOUNT_RATE': self.config['discount_rate'],
                'TAX_RATE': self.config['tax_rate'],
                'ROYALTY_RATE': self.config['royalty_rate'],
                'OPEX_PER_BBL': self.config['opex_per_bbl']
            })
        return pd.DataFrame(assumptions)
    
    def _calculate_lease_financials(self, 
                                   lease_name: str,
                                   group_name: str,
                                   production: Dict,
                                   dc_data: Dict,
                                   oil_prices: pd.Series,
                                   assumptions: pd.DataFrame) -> Dict:
        """Calculate financial metrics for a single lease"""
        
        # Get lease-specific data
        lease_prod = self._get_lease_production(lease_name, production)
        lease_dc = self._get_lease_dc(lease_name, dc_data)
        lease_assumptions = assumptions[assumptions['LEASE_NAME'] == lease_name].iloc[0]
        
        # Create monthly timeline
        start_date = lease_prod.index.min() if not lease_prod.empty else pd.Timestamp('2014-01-01')
        end_date = pd.Timestamp('2050-12-31')
        months = pd.date_range(start=start_date, end=end_date, freq='MS')
        
        # Initialize results DataFrame
        results = pd.DataFrame(index=months)
        results['Month'] = results.index
        results['Lease'] = lease_name
        results['Group'] = group_name
        
        # Add production
        results['Production_BBL'] = lease_prod.reindex(results.index, fill_value=0)
        
        # Add oil prices
        results['Oil_Price'] = oil_prices.reindex(results.index, method='ffill')
        
        # Calculate revenue
        royalty = lease_assumptions.get('ROYALTY_RATE', self.config['royalty_rate'])
        results['Gross_Revenue'] = results['Production_BBL'] * results['Oil_Price']
        results['Net_Revenue'] = results['Gross_Revenue'] * (1 - royalty)
        
        # Add OPEX
        opex_rate = lease_assumptions.get('OPEX_PER_BBL', self.config['opex_per_bbl'])
        results['OPEX'] = results['Production_BBL'] * opex_rate
        
        # Add CAPEX (drilling and completion)
        results['Drilling_CAPEX'] = self._allocate_drilling_costs(lease_dc, results.index)
        results['Completion_CAPEX'] = self._allocate_completion_costs(lease_dc, results.index)
        results['Total_CAPEX'] = results['Drilling_CAPEX'] + results['Completion_CAPEX']
        
        # Calculate EBITDA
        results['EBITDA'] = results['Net_Revenue'] - results['OPEX'] - results['Total_CAPEX']
        
        # Apply taxes
        tax_rate = lease_assumptions.get('TAX_RATE', self.config['tax_rate'])
        results['Tax'] = results['EBITDA'].apply(lambda x: max(0, x * tax_rate))
        
        # Net cash flow
        results['Net_Cash_Flow'] = results['EBITDA'] - results['Tax']
        
        # Calculate NPV
        discount_rate = lease_assumptions.get('DISCOUNT_RATE', self.config['discount_rate'])
        npv = self._calculate_npv(results['Net_Cash_Flow'], discount_rate)
        
        return {
            'monthly_data': results,
            'npv': npv,
            'total_production': results['Production_BBL'].sum(),
            'total_revenue': results['Gross_Revenue'].sum(),
            'total_capex': results['Total_CAPEX'].sum(),
            'total_opex': results['OPEX'].sum(),
            'total_tax': results['Tax'].sum(),
        }
    
    def _get_lease_production(self, lease_name: str, production: Dict) -> pd.Series:
        """Extract production data for specific lease"""
        # Look for lease in production sheets
        if lease_name in production:
            df = production[lease_name]
            if 'Month' in df.columns:
                return df.set_index('Month')['Production']
        return pd.Series()
    
    def _get_lease_dc(self, lease_name: str, dc_data: Dict) -> pd.DataFrame:
        """Extract drilling/completion data for lease"""
        # Implementation would extract relevant D&C data
        # This is simplified for the consolidated script
        return pd.DataFrame()
    
    def _allocate_drilling_costs(self, dc_data: pd.DataFrame, months: pd.DatetimeIndex) -> pd.Series:
        """Allocate drilling costs over drilling period"""
        # Simplified allocation logic
        costs = pd.Series(index=months, data=0.0)
        # Would implement actual cost allocation based on dc_data
        return costs
    
    def _allocate_completion_costs(self, dc_data: pd.DataFrame, months: pd.DatetimeIndex) -> pd.Series:
        """Allocate completion costs at completion date"""
        # Simplified allocation logic
        costs = pd.Series(index=months, data=0.0)
        # Would implement actual cost allocation based on dc_data
        return costs
    
    def _calculate_npv(self, cash_flows: pd.Series, discount_rate: float) -> float:
        """Calculate Net Present Value"""
        months = np.arange(len(cash_flows))
        discount_factors = (1 + discount_rate) ** (months / 12)
        return (cash_flows / discount_factors).sum()
    
    def _generate_summary(self, results: Dict) -> pd.DataFrame:
        """Generate summary metrics for all leases"""
        summary_data = []
        
        for lease_name, lease_results in results.items():
            summary_data.append({
                'Lease': lease_name,
                'NPV': lease_results['npv'],
                'Total_Production_BBL': lease_results['total_production'],
                'Total_Revenue': lease_results['total_revenue'],
                'Total_CAPEX': lease_results['total_capex'],
                'Total_OPEX': lease_results['total_opex'],
                'Total_Tax': lease_results['total_tax'],
            })
            
        return pd.DataFrame(summary_data)
    
    def export_to_excel(self, results: Dict, output_file: str):
        """Export results to formatted Excel file"""
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write README
            readme_data = [
                ['Financial Analysis Output'],
                ['Version:', VERSION],
                ['Based on:', BASED_ON],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                [''],
                ['This workbook contains:'],
                ['- Executive Summary: Key metrics by lease'],
                ['- Monthly data sheets for each lease group'],
                ['- Detailed cash flow analysis'],
            ]
            pd.DataFrame(readme_data).to_excel(
                writer, sheet_name='README', index=False, header=False
            )
            
            # Write Executive Summary
            results['summary'].to_excel(
                writer, sheet_name='Executive_Summary', index=False
            )
            
            # Write lease group sheets
            for lease_name, lease_results in results['lease_results'].items():
                group = self.lease_groups.get(lease_name, lease_name)
                sheet_name = group[:31]  # Excel sheet name limit
                
                monthly = lease_results['monthly_data']
                monthly.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply formatting
        wb = load_workbook(output_file)
        for sheet in wb.worksheets:
            if sheet.title != 'README':
                # Format column widths
                sheet.column_dimensions['A'].width = 12  # Date column
                for col in range(2, 10):
                    sheet.column_dimensions[get_column_letter(col)].width = 15
                
                # Format numbers
                for row in range(2, sheet.max_row + 1):
                    # Date formatting
                    if sheet.cell(row=row, column=1).value:
                        sheet.cell(row=row, column=1).number_format = 'mm/dd/yyyy'
                    
                    # Number formatting for values
                    for col in range(3, 10):
                        if sheet.cell(row=row, column=col).value:
                            sheet.cell(row=row, column=col).number_format = '#,##0'
        
        wb.save(output_file)
        print(f"✅ Results exported to {output_file}")


def main():
    """Main execution function"""
    analyzer = FinancialAnalyzer()
    
    # Example usage (would be parameterized in production)
    results = analyzer.analyze(
        leases_file='leases.xlsx',
        production_file='multi_year_lease_matrix_with_charts.xlsx',
        drilling_completion_file='drilling_and_completion_days_by_api.xlsx',
        oil_price_file='wti_full_monthly.xlsx',
        assumptions_file='leases_assumptions.xlsx'
    )
    
    # Export results
    analyzer.export_to_excel(results, 'financial_analysis_output.xlsx')
    
    # Print summary
    print(f"\n📊 Analysis Complete")
    print(f"Total NPV: ${results['summary']['NPV'].sum():,.0f}")
    print(f"Leases Analyzed: {len(results['lease_results'])}")


if __name__ == '__main__':
    main()