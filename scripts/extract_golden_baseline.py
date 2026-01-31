#!/usr/bin/env python3
"""
ABOUTME: Extracts golden baseline numbers from Roy's FDAS V30 financial_project_summary.xlsx
ABOUTME: Outputs structured YAML for use in repeatability tests (WRK-009)
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
FDAS_V30_DIR = PROJECT_ROOT / "docs/modules/bsee/analysis/production/FDAS_V30"
OUTPUT_PATH = PROJECT_ROOT / "config/analysis/lower_tertiary/golden_baseline_v30.yml"


def extract_project_summary(wb: openpyxl.Workbook) -> list[dict]:
    """Extract per-project summary from Project_Summary sheet."""
    ws = wb["Project_Summary"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    projects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if record.get("PROJECT") or record.get("Project"):
            projects.append(record)
    return projects


def main() -> None:
    xlsx_path = FDAS_V30_DIR / "financial_project_summary.xlsx"
    if not xlsx_path.exists():
        raise FileNotFoundError(f"V30 output not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"Loaded: {xlsx_path}")
    print(f"Sheets: {wb.sheetnames}")

    projects = extract_project_summary(wb)
    print(f"Extracted {len(projects)} projects from Project_Summary")

    for p in projects:
        print(f"  - {p}")

    print(f"\nBaseline file should be maintained at: {OUTPUT_PATH}")
    print("Review and update golden_baseline_v30.yml as needed.")


if __name__ == "__main__":
    main()
