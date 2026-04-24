"""
Report builder for generating comprehensive reports matching go-by format
Creates structured reports with 14-row format and proper aggregations
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .hierarchical_aggregator import HierarchicalAggregator

logger = logging.getLogger(__name__)


class GoByReportBuilder:
    """
    Report builder that generates reports matching go-by Excel format
    Implements the 14-row structure from Jack/Julia/St Malo/Stones field reports
    """

    # Standard 14-row format from go-by reports
    GOBY_ROW_STRUCTURE = [
        "Field Name",
        "Lease Number",
        "Well Name",
        "API Number",
        "Water Depth (ft)",
        "Total Depth (ft)",
        "Spud Date",
        "Status",
        "Oil Production (bbls)",
        "Gas Production (mcf)",
        "Water Production (bbls)",
        "Gross Revenue ($)",
        "Operating Cost ($)",
        "Net Income ($)",
    ]

    def __init__(self, aggregator: Optional[HierarchicalAggregator] = None):
        """
        Initialize report builder

        Args:
            aggregator: Hierarchical aggregator for data processing
        """
        self.aggregator = aggregator or HierarchicalAggregator()
        self.report_data = None

    def build_report(
        self,
        report_type: str,
        entity_identifier: str,
        cfg: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Build comprehensive report for specified entity

        Args:
            report_type: Type of report ('block', 'field', 'lease')
            entity_identifier: Identifier for the entity (block number, field name, or lease number)
            cfg: Configuration dictionary

        Returns:
            Structured report data matching go-by format
        """
        logger.info(f"Building {report_type} report for {entity_identifier}")

        # Set up filters based on report type
        block_number = None
        field_name = None
        lease_number = None

        if report_type == "block":
            block_number = entity_identifier
        elif report_type == "field":
            field_name = entity_identifier
        elif report_type == "lease":
            lease_number = entity_identifier
        else:
            raise ValueError(f"Invalid report type: {report_type}")

        # Generate hierarchical data using aggregator
        self.report_data = self.aggregator.generate_report_data(
            block_number=block_number,
            field_name=field_name,
            lease_number=lease_number,
            cfg=cfg,
        )

        # Build structured report sections
        report = {
            "metadata": self._build_metadata(report_type, entity_identifier),
            "summary": self._build_summary_section(),
            "detail_rows": self._build_detail_rows(report_type),
            "aggregations": self._build_aggregation_section(report_type),
            "charts_data": self._prepare_chart_data(),
        }

        return report

    def _build_metadata(
        self, report_type: str, entity_identifier: str
    ) -> Dict[str, Any]:
        """
        Build report metadata section

        Args:
            report_type: Type of report
            entity_identifier: Entity identifier

        Returns:
            Metadata dictionary
        """
        return {
            "report_type": report_type,
            "entity": entity_identifier,
            "generated_date": datetime.now().strftime("%Y-%m-%d"),
            "generated_time": datetime.now().strftime("%H:%M:%S"),
            "data_source": "BSEE",
            "report_format": "Go-By Standard 14-Row",
            "version": "1.0",
        }

    def _build_summary_section(self) -> Dict[str, Any]:
        """
        Build executive summary section

        Returns:
            Summary dictionary
        """
        if not self.report_data:
            return {}

        summary = self.report_data.get("summary", {})

        return {
            "total_entities": summary.get("entity_count", 0),
            "total_oil_production": f"{summary.get('total_oil_bbls', 0):,.0f} bbls",
            "total_gas_production": f"{summary.get('total_gas_mcf', 0):,.0f} mcf",
            "total_water_production": f"{summary.get('total_water_bbls', 0):,.0f} bbls",
            "total_boe": f"{summary.get('total_boe', 0):,.0f} BOE",
            "gross_revenue": f"${summary.get('total_gross_revenue', 0):,.2f}",
            "total_costs": f"${summary.get('total_costs', 0):,.2f}",
            "net_income": f"${summary.get('total_net_income', 0):,.2f}",
            "profit_margin": f"{summary.get('profit_margin', 0):.1f}%",
        }

    def _build_detail_rows(self, report_type: str) -> List[List[Any]]:
        """
        Build detail rows in 14-row go-by format

        Args:
            report_type: Type of report

        Returns:
            List of detail rows
        """
        if not self.report_data:
            return []

        detail_rows = []
        data = self.report_data.get("data", {})

        for entity_id, entity_data in data.items():
            if report_type == "block":
                # For block report, show field-level summaries
                for field in entity_data.get("fields", []):
                    for lease in field.get("leases", []):
                        for well in lease.get("wells", []):
                            row = self._create_well_row(field, lease, well)
                            detail_rows.append(row)

            elif report_type == "field":
                # For field report, show lease-level details
                for lease in entity_data.get("leases", []):
                    for well in lease.get("wells", []):
                        row = self._create_well_row(entity_data, lease, well)
                        detail_rows.append(row)

            elif report_type == "lease":
                # For lease report, show well-level details
                for well in entity_data.get("wells", []):
                    row = self._create_well_row(None, entity_data, well)
                    detail_rows.append(row)

        return detail_rows

    def _create_well_row(
        self, field: Optional[Dict], lease: Dict, well: Dict
    ) -> List[Any]:
        """
        Create a single well row in go-by format

        Args:
            field: Field data (optional)
            lease: Lease data
            well: Well data

        Returns:
            Row data list matching 14-row structure
        """
        return [
            field.get("field_name", "N/A") if field else "N/A",  # Field Name
            lease.get("lease_number", lease.get("lease_id", "N/A")),  # Lease Number
            well.get("well_name", "N/A"),  # Well Name
            well.get("api_number", "N/A"),  # API Number
            well.get("water_depth_ft", 0),  # Water Depth
            well.get("total_depth_ft", 0),  # Total Depth
            (
                well.get("spud_date", "N/A") if well.get("spud_date") else "N/A"
            ),  # Spud Date
            well.get("status", "Unknown"),  # Status
            round(well.get("oil_production_bbls", 0), 0),  # Oil Production
            round(well.get("gas_production_mcf", 0), 0),  # Gas Production
            round(well.get("water_production_bbls", 0), 0),  # Water Production
            round(well.get("gross_revenue", 0), 2),  # Gross Revenue
            round(well.get("operating_cost", 0), 2),  # Operating Cost
            round(well.get("net_income", 0), 2),  # Net Income
        ]

    def _build_aggregation_section(self, report_type: str) -> Dict[str, Any]:
        """
        Build aggregation section with rollup totals

        Args:
            report_type: Type of report

        Returns:
            Aggregation dictionary
        """
        if not self.report_data:
            return {}

        data = self.report_data.get("data", {})
        aggregations = {}

        if report_type == "block":
            # Aggregate by field
            field_totals = {}
            for block_id, block_data in data.items():
                for field in block_data.get("fields", []):
                    field_name = field.get("field_name")
                    field_totals[field_name] = {
                        "oil_bbls": field.get("oil_production_bbls", 0),
                        "gas_mcf": field.get("gas_production_mcf", 0),
                        "revenue": field.get("gross_revenue", 0),
                        "wells": field.get("total_wells", 0),
                        "leases": field.get("total_leases", 0),
                    }
            aggregations["by_field"] = field_totals

        elif report_type == "field":
            # Aggregate by lease
            lease_totals = {}
            for field_id, field_data in data.items():
                for lease in field_data.get("leases", []):
                    lease_num = lease.get("lease_number")
                    lease_totals[lease_num] = {
                        "oil_bbls": lease.get("oil_production_bbls", 0),
                        "gas_mcf": lease.get("gas_production_mcf", 0),
                        "revenue": lease.get("gross_revenue", 0),
                        "wells": lease.get("total_wells", 0),
                    }
            aggregations["by_lease"] = lease_totals

        elif report_type == "lease":
            # Aggregate by well status
            status_totals = {"active": 0, "inactive": 0, "abandoned": 0}
            for lease_id, lease_data in data.items():
                for well in lease_data.get("wells", []):
                    status = well.get("status", "unknown").lower()
                    if status in status_totals:
                        status_totals[status] += 1
                    else:
                        status_totals["inactive"] += 1
            aggregations["by_status"] = status_totals

        return aggregations

    def _prepare_chart_data(self) -> Dict[str, Any]:
        """
        Prepare data for visualization charts

        Returns:
            Chart data dictionary
        """
        if not self.report_data:
            return {}

        summary = self.report_data.get("summary", {})

        return {
            "production_pie": {
                "labels": ["Oil (BOE)", "Gas (BOE)", "NGL (BOE)"],
                "values": [
                    summary.get("total_oil_bbls", 0),
                    summary.get("total_gas_mcf", 0) / 6,  # Convert to BOE
                    0,  # NGL if available
                ],
            },
            "revenue_breakdown": {
                "labels": ["Operating Cost", "Royalties", "Taxes", "Net Income"],
                "values": [
                    summary.get("total_costs", 0) * 0.5,  # Approximate breakdown
                    summary.get("total_costs", 0) * 0.3,
                    summary.get("total_costs", 0) * 0.2,
                    summary.get("total_net_income", 0),
                ],
            },
        }

    def export_to_excel(
        self, report: Dict[str, Any], output_path: Path, filename: Optional[str] = None
    ) -> Path:
        """
        Export report to Excel format matching go-by structure

        Args:
            report: Report data
            output_path: Output directory path
            filename: Optional filename (auto-generated if not provided)

        Returns:
            Path to generated Excel file
        """
        # Generate filename if not provided
        if not filename:
            metadata = report.get("metadata", {})
            report_type = metadata.get("report_type", "report")
            entity = metadata.get("entity", "unknown")
            date_str = datetime.now().strftime("%Y%m%d")
            filename = f"{report_type}_{entity}_{date_str}.xlsx"

        # Ensure output directory exists
        output_path = Path(output_path)
        output_path.mkdir(parents=True, exist_ok=True)
        filepath = output_path / filename

        # Create Excel workbook
        from openpyxl import Workbook

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary, report)

        # Detail sheet
        ws_detail = wb.create_sheet("Details")
        self._write_detail_sheet(ws_detail, report)

        # Aggregations sheet
        ws_agg = wb.create_sheet("Aggregations")
        self._write_aggregation_sheet(ws_agg, report)

        # Save workbook
        wb.save(filepath)
        logger.info(f"Report exported to {filepath}")

        return filepath

    def _write_summary_sheet(self, worksheet, report: Dict[str, Any]):
        """
        Write summary data to Excel worksheet

        Args:
            worksheet: Openpyxl worksheet
            report: Report data
        """
        # Title
        worksheet["A1"] = "COMPREHENSIVE REPORT SUMMARY"
        worksheet["A1"].font = Font(bold=True, size=14)

        # Metadata
        metadata = report.get("metadata", {})
        worksheet["A3"] = "Report Information"
        worksheet["A3"].font = Font(bold=True)
        worksheet["A4"] = f"Type: {metadata.get('report_type', 'N/A').upper()}"
        worksheet["A5"] = f"Entity: {metadata.get('entity', 'N/A')}"
        worksheet["A6"] = f"Generated: {metadata.get('generated_date', 'N/A')}"

        # Summary metrics
        summary = report.get("summary", {})
        worksheet["A8"] = "Key Metrics"
        worksheet["A8"].font = Font(bold=True)

        row = 9
        for key, value in summary.items():
            worksheet[f"A{row}"] = key.replace("_", " ").title()
            worksheet[f"B{row}"] = str(value)
            row += 1

        # Format columns
        worksheet.column_dimensions["A"].width = 25
        worksheet.column_dimensions["B"].width = 20

    def _write_detail_sheet(self, worksheet, report: Dict[str, Any]):
        """
        Write detail rows to Excel worksheet in go-by format

        Args:
            worksheet: Openpyxl worksheet
            report: Report data
        """
        # Write headers (14-row structure)
        for col, header in enumerate(self.GOBY_ROW_STRUCTURE, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        # Write detail rows
        detail_rows = report.get("detail_rows", [])
        for row_idx, row_data in enumerate(detail_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Add totals row
        if detail_rows:
            total_row = len(detail_rows) + 2
            worksheet.cell(row=total_row, column=1, value="TOTAL")
            worksheet.cell(row=total_row, column=1).font = Font(bold=True)

            # Sum numeric columns (9-14)
            for col in range(9, 15):
                col_letter = get_column_letter(col)
                formula = f"=SUM({col_letter}2:{col_letter}{total_row-1})"
                worksheet.cell(row=total_row, column=col, value=formula)
                worksheet.cell(row=total_row, column=col).font = Font(bold=True)

        # Auto-adjust column widths
        for col in range(1, 15):
            worksheet.column_dimensions[get_column_letter(col)].width = 15

    def _write_aggregation_sheet(self, worksheet, report: Dict[str, Any]):
        """
        Write aggregation data to Excel worksheet

        Args:
            worksheet: Openpyxl worksheet
            report: Report data
        """
        worksheet["A1"] = "AGGREGATIONS"
        worksheet["A1"].font = Font(bold=True, size=14)

        aggregations = report.get("aggregations", {})
        row = 3

        for agg_type, agg_data in aggregations.items():
            # Section header
            worksheet[f"A{row}"] = agg_type.replace("_", " ").title()
            worksheet[f"A{row}"].font = Font(bold=True)
            row += 1

            # Aggregation data
            if isinstance(agg_data, dict):
                for key, value in agg_data.items():
                    worksheet[f"B{row}"] = str(key)
                    if isinstance(value, dict):
                        col = 3
                        for metric_key, metric_value in value.items():
                            worksheet.cell(row=row, column=col, value=metric_value)
                            col += 1
                    else:
                        worksheet[f"C{row}"] = value
                    row += 1
            row += 1  # Add spacing between sections

    def generate_step_by_step_report(
        self,
        report_type: str,
        entity_identifier: str,
        output_path: Path,
        cfg: Optional[Dict[str, Any]] = None,
    ) -> Path:
        """
        Generate complete report following step-by-step process

        Args:
            report_type: Type of report ('block', 'field', 'lease')
            entity_identifier: Entity identifier
            output_path: Output directory
            cfg: Configuration dictionary

        Returns:
            Path to generated report file
        """
        logger.info("Starting step-by-step report generation")

        # Step 1: Build report data structure
        logger.info("Step 1: Building report data structure")
        report = self.build_report(report_type, entity_identifier, cfg)

        # Step 2: Validate data completeness
        logger.info("Step 2: Validating data completeness")
        self._validate_report_data(report)

        # Step 3: Apply go-by formatting rules
        logger.info("Step 3: Applying go-by formatting rules")
        formatted_report = self._apply_goby_formatting(report)

        # Step 4: Export to Excel
        logger.info("Step 4: Exporting to Excel")
        excel_path = self.export_to_excel(formatted_report, output_path)

        # Step 5: Log completion
        logger.info(f"Report generation complete: {excel_path}")

        return excel_path

    def _validate_report_data(self, report: Dict[str, Any]):
        """
        Validate report data for completeness

        Args:
            report: Report data to validate
        """
        required_sections = ["metadata", "summary", "detail_rows", "aggregations"]
        for section in required_sections:
            if section not in report:
                logger.warning(f"Missing report section: {section}")

        # Validate detail rows match go-by structure
        detail_rows = report.get("detail_rows", [])
        if detail_rows:
            for row in detail_rows:
                if len(row) != len(self.GOBY_ROW_STRUCTURE):
                    logger.warning(
                        f"Row length mismatch: expected {len(self.GOBY_ROW_STRUCTURE)}, got {len(row)}"  # noqa: E501
                    )

    def _apply_goby_formatting(self, report: Dict[str, Any]) -> Dict[str, Any]:
        """
        Apply go-by specific formatting rules

        Args:
            report: Report data

        Returns:
            Formatted report
        """
        # Format numbers according to go-by standards
        detail_rows = report.get("detail_rows", [])
        for row in detail_rows:
            # Format production volumes (no decimals)
            for i in [8, 9, 10]:  # Oil, Gas, Water columns
                if isinstance(row[i], (int, float)):
                    row[i] = int(row[i])

            # Format currency values (2 decimals)
            for i in [11, 12, 13]:  # Revenue, Cost, Net Income columns
                if isinstance(row[i], (int, float)):
                    row[i] = round(row[i], 2)

        return report
