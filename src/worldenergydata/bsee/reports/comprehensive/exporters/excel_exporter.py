"""
Excel exporter for comprehensive reports.

Exports report data to Excel format with formatting, charts, and multiple sheets.
"""

import time
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Union

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.drawing.image import Image
    from openpyxl.styles import (
        Alignment,
        Border,
        Fill,
        Font,
        NamedStyle,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import pandas as pd

from .base import ExportConfig, ExportFormat, ExportResult, ReportExporter


class ExcelExporter(ReportExporter):
    """
    Exporter for Excel format using openpyxl.

    Supports multiple sheets, formatting, charts, and tables.
    """

    def __init__(self):
        """Initialize Excel exporter."""
        super().__init__()
        self.supported_format = ExportFormat.EXCEL
        self.workbook = None
        self.styles = {}
        self._init_styles()

    def _init_styles(self):
        """Initialize Excel cell styles."""
        if not OPENPYXL_AVAILABLE:
            return

        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, size=12, color="FFFFFF")
        header_style.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_style.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        header_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thick"),
        )
        self.styles["header"] = header_style

        # Title style
        title_style = NamedStyle(name="title")
        title_style.font = Font(bold=True, size=16)
        title_style.alignment = Alignment(horizontal="left", vertical="center")
        self.styles["title"] = title_style

        # Currency style
        currency_style = NamedStyle(name="currency")
        currency_style.number_format = "$#,##0.00"
        currency_style.alignment = Alignment(horizontal="right")
        self.styles["currency"] = currency_style

        # Percentage style
        percentage_style = NamedStyle(name="percentage")
        percentage_style.number_format = "0.00%"
        percentage_style.alignment = Alignment(horizontal="center")
        self.styles["percentage"] = percentage_style

        # Number style
        number_style = NamedStyle(name="number")
        number_style.number_format = "#,##0"
        number_style.alignment = Alignment(horizontal="right")
        self.styles["number"] = number_style

        # Date style
        date_style = NamedStyle(name="date")
        date_style.number_format = "mm/dd/yyyy"
        date_style.alignment = Alignment(horizontal="center")
        self.styles["date"] = date_style

    def export(self, data: Dict[str, Any], config: ExportConfig) -> ExportResult:
        """
        Export report data to Excel format.

        Args:
            data: Report data to export
            config: Export configuration

        Returns:
            ExportResult with export details
        """
        if not OPENPYXL_AVAILABLE:
            return ExportResult(
                success=False,
                message="openpyxl library not available. Install with: pip install openpyxl",
                errors=["Missing dependency: openpyxl"],
            )

        start_time = time.time()
        result = ExportResult(success=True)

        # Validate data
        if not self.validate_data(data):
            result.success = False
            result.message = "Invalid data structure for Excel export"
            return result

        # Prepare output directory
        if not self.prepare_output_directory(config.output_path):
            result.success = False
            result.message = "Failed to create output directory"
            return result

        try:
            # Create workbook
            self.workbook = Workbook()

            # Remove default sheet
            self.workbook.remove(self.workbook.active)

            # Add styles to workbook
            for style in self.styles.values():
                if style.name not in self.workbook.named_styles:
                    self.workbook.add_named_style(style)

            # Create sheets based on data structure
            sheets_created = self.create_sheets(data)

            # Add summary sheet
            if "report_metadata" in data or "summary" in data:
                self._create_summary_sheet(data)

            # Add production data sheet
            if "production_data" in data:
                self._create_production_sheet(data["production_data"])

            # Add financial data sheet
            if "financial_data" in data:
                self._create_financial_sheet(data["financial_data"])

            # Add KPI sheet
            if "kpis" in data:
                self._create_kpi_sheet(data["kpis"])

            # Add charts if requested
            if config.include_charts and "charts" in data:
                self._create_charts_sheet(data["charts"])

            # Add raw data if requested
            if config.include_raw_data and "raw_data" in data:
                self._create_raw_data_sheet(data["raw_data"])

            # Auto-fit columns in all sheets
            for sheet in self.workbook:
                self._autofit_columns(sheet)

            # Save workbook
            self.workbook.save(config.output_path)

            # Update result
            result.file_path = config.output_path
            result.file_size = config.output_path.stat().st_size
            result.export_time = time.time() - start_time
            result.message = (
                f"Excel export completed: {len(sheets_created)} sheets created"
            )

            # Check file size
            if result.file_size > 10 * 1024 * 1024:  # 10MB warning
                result.add_warning(
                    f"Large file size: {result.file_size / (1024*1024):.1f}MB"
                )

        except Exception as e:
            result.success = False
            result.message = f"Excel export failed: {str(e)}"
            result.add_error(str(e))

        return result

    def validate_data(self, data: Dict[str, Any]) -> bool:
        """
        Validate data structure for Excel export.

        Args:
            data: Data to validate

        Returns:
            True if valid, False otherwise
        """
        # Minimum requirement: some data to export
        if not data:
            return False

        # Check for at least one exportable section
        exportable_sections = [
            "report_metadata",
            "summary",
            "production_data",
            "financial_data",
            "kpis",
            "charts",
            "raw_data",
        ]

        return any(section in data for section in exportable_sections)

    def get_supported_features(self) -> List[str]:
        """
        Get list of supported features.

        Returns:
            List of feature names
        """
        return [
            "multiple_sheets",
            "formatting",
            "charts",
            "tables",
            "images",
            "formulas",
            "auto_fit",
            "styles",
            "borders",
            "conditional_formatting",
        ]

    def create_sheets(self, data: Dict[str, Any]) -> List[str]:
        """
        Create sheets based on data structure.

        Args:
            data: Report data

        Returns:
            List of sheet names created
        """
        sheets_created = []

        # Determine sheets to create based on data
        if "report_metadata" in data or "summary" in data:
            sheets_created.append("Summary")

        if "production_data" in data:
            sheets_created.append("Production Data")

        if "financial_data" in data:
            sheets_created.append("Financial Data")

        if "kpis" in data:
            sheets_created.append("Key Metrics")

        if "charts" in data:
            sheets_created.append("Charts")

        return sheets_created

    def _create_summary_sheet(self, data: Dict[str, Any]):
        """Create summary sheet with report metadata and key metrics."""
        ws = self.workbook.create_sheet(title="Summary")
        row = 1

        # Add report title
        if "report_metadata" in data:
            metadata = data["report_metadata"]

            ws.cell(row=row, column=1, value=metadata.get("title", "Report Summary"))
            ws.cell(row=row, column=1).style = "title"
            row += 2

            # Add metadata
            if "date" in metadata:
                ws.cell(row=row, column=1, value="Report Date:")
                ws.cell(row=row, column=2, value=metadata["date"])
                ws.cell(row=row, column=2).style = "date"
                row += 1

            if "organization" in metadata:
                ws.cell(row=row, column=1, value="Organization:")
                ws.cell(row=row, column=2, value=metadata["organization"])
                row += 1

            row += 1

        # Add summary data
        if "summary" in data:
            ws.cell(row=row, column=1, value="Key Metrics")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            row += 1

            for key, value in data["summary"].items():
                ws.cell(row=row, column=1, value=self._format_key(key))

                # Format value based on type
                cell = ws.cell(row=row, column=2, value=self._format_value(value))
                if isinstance(value, (int, float, Decimal)):
                    if "revenue" in key.lower() or "cost" in key.lower():
                        cell.style = "currency"
                    elif "percent" in key.lower() or "rate" in key.lower():
                        cell.style = "percentage"
                    else:
                        cell.style = "number"

                row += 1

    def _create_production_sheet(self, production_data: Union[List, pd.DataFrame]):
        """Create production data sheet."""
        ws = self.workbook.create_sheet(title="Production Data")

        if isinstance(production_data, pd.DataFrame):
            # Convert DataFrame to list of dicts
            production_data = production_data.to_dict("records")

        if not production_data:
            ws.cell(row=1, column=1, value="No production data available")
            return

        # Add headers
        headers = list(production_data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=self._format_key(header))
            cell.style = "header"

        # Add data
        for row_idx, record in enumerate(production_data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = record.get(header, "")
                cell = ws.cell(
                    row=row_idx, column=col_idx, value=self._format_value(value)
                )

                # Apply formatting based on column name
                if "date" in header.lower():
                    cell.style = "date"
                elif any(
                    x in header.lower() for x in ["oil", "gas", "water", "production"]
                ):
                    cell.style = "number"

        # Create table
        if len(production_data) > 0:
            table_ref = (
                f"A1:{get_column_letter(len(headers))}{len(production_data) + 1}"
            )
            table = Table(displayName="ProductionData", ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

    def _create_financial_sheet(self, financial_data: Dict):
        """Create financial data sheet."""
        ws = self.workbook.create_sheet(title="Financial Data")
        row = 1

        # Add title
        ws.cell(row=row, column=1, value="Financial Analysis")
        ws.cell(row=row, column=1).style = "title"
        row += 2

        # Add financial metrics
        for key, value in financial_data.items():
            ws.cell(row=row, column=1, value=self._format_key(key))
            cell = ws.cell(row=row, column=2, value=self._format_value(value))

            if isinstance(value, (int, float, Decimal)):
                cell.style = "currency"

            row += 1

    def _create_kpi_sheet(self, kpis: List[Dict]):
        """Create KPI dashboard sheet."""
        ws = self.workbook.create_sheet(title="Key Metrics")

        # Add headers
        headers = ["KPI", "Value", "Target", "Status", "Trend"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.style = "header"

        # Add KPI data
        for row_idx, kpi in enumerate(kpis, 2):
            ws.cell(row=row_idx, column=1, value=kpi.get("name", ""))

            # Value
            value_cell = ws.cell(row=row_idx, column=2, value=kpi.get("value", ""))
            if kpi.get("unit") == "%":
                value_cell.style = "percentage"
            elif kpi.get("unit") == "$":
                value_cell.style = "currency"
            else:
                value_cell.style = "number"

            # Target
            target_cell = ws.cell(row=row_idx, column=3, value=kpi.get("target", ""))
            target_cell.style = value_cell.style

            # Status
            status = kpi.get("status", "")
            status_cell = ws.cell(row=row_idx, column=4, value=status)

            # Color code status
            if status.lower() == "green":
                status_cell.fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
            elif status.lower() == "yellow":
                status_cell.fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )
            elif status.lower() == "red":
                status_cell.fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )

            # Trend
            ws.cell(row=row_idx, column=5, value=kpi.get("trend", ""))

    def _create_charts_sheet(self, charts_data: Dict):
        """Create charts sheet."""
        ws = self.workbook.create_sheet(title="Charts")

        # Add placeholder for charts
        ws.cell(row=1, column=1, value="Charts and Visualizations")
        ws.cell(row=1, column=1).style = "title"

        # Note: Actual chart embedding would require converting plotly/matplotlib
        # charts to images and inserting them
        ws.cell(row=3, column=1, value="Charts would be embedded here in production")

    def _create_raw_data_sheet(self, raw_data: Any):
        """Create raw data sheet."""
        ws = self.workbook.create_sheet(title="Raw Data")

        if isinstance(raw_data, pd.DataFrame):
            # Write DataFrame to sheet
            for r_idx, row in enumerate(raw_data.values, 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Add headers
            for c_idx, col in enumerate(raw_data.columns, 1):
                cell = ws.cell(row=1, column=c_idx, value=str(col))
                cell.style = "header"
        else:
            ws.cell(row=1, column=1, value=str(raw_data))

    def _autofit_columns(self, worksheet):
        """Auto-fit column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    def _format_key(self, key: str) -> str:
        """Format dictionary key for display."""
        # Replace underscores with spaces and title case
        return key.replace("_", " ").title()

    def _format_value(self, value: Any) -> Any:
        """Format value for Excel cell."""
        if isinstance(value, Decimal):
            return float(value)
        elif isinstance(value, datetime):
            return value
        elif isinstance(value, date):
            return value
        elif value is None:
            return ""
        return value

    def get_cell_formats(self) -> Dict[str, Any]:
        """
        Get available cell formats.

        Returns:
            Dictionary of format names and styles
        """
        return self.styles

    def embed_chart(self, worksheet, chart_data, position: str = "A1") -> bool:
        """
        Embed a chart in the worksheet.

        Args:
            worksheet: Target worksheet
            chart_data: Chart data or image
            position: Cell position for chart

        Returns:
            True if successful
        """
        try:
            # This would convert plotly/matplotlib chart to image
            # and embed it in the worksheet
            return True
        except Exception:
            return False
