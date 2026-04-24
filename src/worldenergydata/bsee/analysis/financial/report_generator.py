"""
SME Financial Analysis Report Generator

This module generates Excel reports matching the V20 format from SME Roy's implementation,
including executive summary, project summary, and detailed cash flow sheets.
"""

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (  # noqa: F401
        Alignment,
        Border,
        Fill,
        Font,
        NamedStyle,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo  # noqa: F401

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    load_workbook = None

logger = logging.getLogger(__name__)


def format_currency(value: Union[int, float]) -> str:
    """Format value as currency string"""
    if pd.isna(value):
        return ""
    if value < 0:
        return f"-${int(round(abs(value))):,}"
    return f"${int(round(value)):,}"


def format_number(value: Union[int, float]) -> str:
    """Format value as number string with commas"""
    if pd.isna(value):
        return ""
    return f"{int(round(value)):,}"


def format_percentage(value: Union[int, float]) -> str:
    """Format value as percentage string"""
    if pd.isna(value):
        return ""
    return f"{value*100:.2f}%"


class ExcelFormatter:
    """
    Excel formatting utilities for V20 report format
    """

    def __init__(self):
        """Initialize formatter with styles"""
        self.styles = {}
        if OPENPYXL_AVAILABLE:
            self._init_styles()

    def _init_styles(self):
        """Initialize Excel cell styles matching V20 format"""
        # Header style
        header_style = NamedStyle(name="sme_header")
        header_style.font = Font(bold=True, size=11)
        header_style.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        header_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.styles["header"] = header_style

        # Title style
        title_style = NamedStyle(name="sme_title")
        title_style.font = Font(bold=True, size=14)
        title_style.alignment = Alignment(horizontal="left", vertical="center")
        self.styles["title"] = title_style

        # Currency style
        currency_style = NamedStyle(name="sme_currency")
        currency_style.number_format = "#,##0"
        currency_style.alignment = Alignment(horizontal="right")
        self.styles["currency"] = currency_style

        # Percentage style
        percentage_style = NamedStyle(name="sme_percentage")
        percentage_style.number_format = "0.00%"
        percentage_style.alignment = Alignment(horizontal="center")
        self.styles["percentage"] = percentage_style

        # Number style
        number_style = NamedStyle(name="sme_number")
        number_style.number_format = "#,##0"
        number_style.alignment = Alignment(horizontal="right")
        self.styles["number"] = number_style

        # Date style
        date_style = NamedStyle(name="sme_date")
        date_style.number_format = "mmm-yy"
        date_style.alignment = Alignment(horizontal="center")
        self.styles["date"] = date_style

    def apply_header_style(self, worksheet):
        """Apply header style to first row of worksheet"""
        for cell in worksheet[1]:
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )
            cell.font = Font(bold=True)

    def set_column_widths(self, worksheet, width: float = 15.0):
        """Set all column widths in worksheet"""
        for col in range(1, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = width

    def format_numbers(
        self,
        worksheet,
        percent_headers: List[str] = None,
        skip_headers: List[str] = None,
    ):
        """Apply number formatting to worksheet cells"""
        percent_headers = set(percent_headers or [])
        skip_headers = set(skip_headers or [])

        # Get header mapping
        header_map = {}
        for col in range(1, worksheet.max_column + 1):
            val = worksheet.cell(row=1, column=col).value
            if val:
                header_map[str(val).strip()] = col

        # Apply formatting
        for header, col in header_map.items():
            if header in skip_headers:
                continue

            if header in percent_headers:
                num_format = "0.00%"
            elif header == "FO Month":
                num_format = "mmm-yy"
            else:
                num_format = "#,##0"

            # Apply to data cells
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                val = cell.value

                if header == "FO Month":
                    cell.number_format = num_format
                elif isinstance(val, (int, float)) and val is not None:
                    cell.number_format = num_format


class ReportGenerator:
    """
    Generate Excel reports for SME financial analysis
    """

    def __init__(self):
        """Initialize report generator"""
        self.formatter = ExcelFormatter()
        self.workbook = None

    def generate_report(
        self,
        development_data: Dict[str, Dict[str, Any]],
        output_path: str,
        version: str = "V20",
        assumptions: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Generate comprehensive Excel report

        Args:
            development_data: Dictionary of development name -> data
                Each development should have:
                - cash_flow: DataFrame with monthly cash flows
                - metrics: Dictionary of financial metrics
                - first_oil: First oil date
                - development_type: 'subsea' or 'dry_tree'
                - producer_count: Number of producing wells
            output_path: Path to save Excel file
            version: Report version
            assumptions: Optional assumptions data

        Returns:
            Path to generated file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        # Create workbook
        self.workbook = Workbook()

        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

        # Create README sheet (first)
        self.create_readme_sheet(
            version=version,
            generated_date=datetime.now(),
            input_files=assumptions.get("input_files", []) if assumptions else [],
        )

        # Create Executive Summary
        self.create_executive_summary(development_data, assumptions)

        # Create Project Summary
        self.create_project_summary(development_data)

        # Create QC sheets if D&C data available
        if any("dnc_data" in dev for dev in development_data.values()):
            self.create_qc_sheets(development_data)

        # Create individual development sheets
        for dev_name, dev_data in development_data.items():
            if "cash_flow" in dev_data:
                self.create_development_sheet(dev_name, dev_data["cash_flow"])

        # Save workbook
        self._save_workbook(output_path)

        # Apply V20 formatting
        self._apply_v20_formatting(output_path)

        logger.info(f"Report generated: {output_path}")
        return output_path

    def create_readme_sheet(
        self, version: str, generated_date: datetime, input_files: List[str]
    ):
        """Create README sheet with report metadata"""
        ws = self.workbook.create_sheet("README", 0)

        # Title
        ws["A1"] = f"README — Financial Analysis {version} Output Workbook"
        ws["A1"].font = Font(bold=True, size=14)

        # Metadata
        ws["A2"] = "Generated On"
        ws["B2"] = generated_date.strftime("%B %d, %Y %H:%M")

        ws["A3"] = "File Name"
        ws["B3"] = "Financial Analysis Report"

        # Overview
        ws["A5"] = "Overview"
        ws["B5"] = (
            f"This workbook contains per-development financial results produced by the {version} engine. "  # noqa: E501
            "It aggregates monthly oil production, drilling & completion (D&C) day allocations, facilities CAPEX, and OPEX "  # noqa: E501
            "to construct after-tax cash flows per development. Economics (NPV and MIRR) are computed from those monthly cash flows."  # noqa: E501
        )

        # Inputs
        ws["A7"] = "Inputs (files)"
        inputs_text = "- leases.xlsx (DEV_NAME, LEASE_NAME/NUM, DEV_TYPE/DEV_SYSTEM)\n"
        inputs_text += "- leases_assumptions.xlsx (rates, costs, taxes)\n"
        inputs_text += "- production data (monthly production by lease/well)\n"
        inputs_text += (
            "- drilling_and_completion data (D&C per-well totals/monthlies)\n"
        )
        inputs_text += (
            "- WTI prices (optional monthly WTI; otherwise flat WTI base is used)"
        )
        ws["B7"] = inputs_text

        # Calculations
        ws["A10"] = "Calculations (summary)"
        ws["B10"] = (
            "• Revenue: monthly gross oil barrels × WTI, minus royalty & severance.\n"
            "• OPEX: variable per-barrel + fixed monthly.\n"
            "• CAPEX: D&C days × dayrate (MODU or DRY per system/FO timing), plus facilities allocations.\n"  # noqa: E501
            "• Taxes: corporate tax savings applied to CAPEX in-month.\n"
            "• Net Cash Flow: Revenue_Net − OPEX − CAPEX + Tax_Savings.\n"
            "• NPV10 & MIRR: computed from monthly cash flows."
        )

        # Outputs
        ws["A15"] = "Outputs (tabs)"
        ws["B15"] = (
            f"- Executive Summary ({version}): high-level per-development metrics with NPV/MIRR.\n"
            f"- Project Summary ({version}): detailed CAPEX/OPEX breakdown, wells & equipment counts.\n"  # noqa: E501
            "- QC sheets: reconciliation tables of input vs allocated D&C days.\n"
            "- <DEV_NAME> tabs: month-by-month cash flow build and cumulative cash flow."
        )

        # Format cells
        for cell in [ws["B5"], ws["B7"], ws["B10"], ws["B15"]]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 120

    def create_executive_summary(
        self,
        development_data: Dict[str, Dict[str, Any]],
        assumptions: Optional[Dict[str, Any]] = None,
    ):
        """Create Executive Summary sheet"""
        ws = self.workbook.create_sheet("Executive Summary (V20)")

        # Prepare data
        exec_data = self._format_executive_summary_data(development_data)

        # Create DataFrame
        df = pd.DataFrame(exec_data)

        # Write to worksheet
        self._write_dataframe_to_sheet(ws, df)

        # Apply formatting
        self.formatter.apply_header_style(ws)
        self.formatter.set_column_widths(ws, 15.0)

        # Add rate columns if assumptions provided
        if assumptions:
            self._add_rate_columns(ws, assumptions)

    def create_project_summary(self, development_data: Dict[str, Dict[str, Any]]):
        """Create Project Summary sheet"""
        ws = self.workbook.create_sheet("Project Summary (V20)")

        # Prepare data
        proj_data = self._format_project_summary_data(development_data)

        # Create DataFrame
        df = pd.DataFrame(proj_data)

        # Write to worksheet
        self._write_dataframe_to_sheet(ws, df)

        # Apply formatting
        self.formatter.apply_header_style(ws)
        self.formatter.set_column_widths(ws, 14.0)

    def create_development_sheet(self, dev_name: str, cash_flow_df: pd.DataFrame):
        """Create individual development cash flow sheet"""
        # Truncate name if needed (Excel limit is 31 chars)
        sheet_name = self._truncate_sheet_name(dev_name)
        ws = self.workbook.create_sheet(sheet_name)

        # Reset index to include YearMonth as column
        df = cash_flow_df.reset_index()

        # Write to worksheet
        self._write_dataframe_to_sheet(ws, df)

        # Apply formatting
        self.formatter.apply_header_style(ws)
        self.formatter.set_column_widths(ws, 12.0)

    def create_qc_sheets(self, development_data: Dict[str, Dict[str, Any]]):
        """Create QC sheets for D&C allocation"""
        # QC - D&C Allocation sheet
        ws = self.workbook.create_sheet("QC — D&C Allocation")

        qc_rows = []
        for dev_name, dev_data in development_data.items():
            if "dnc_data" in dev_data:
                for well, well_data in dev_data["dnc_data"].items():
                    qc_rows.append(
                        {
                            "DEV_NAME": dev_name,
                            "WELL": well,
                            "SPUD": well_data.get("spud_date"),
                            "TD": well_data.get("td_date"),
                            "FO_MONTH": well_data.get("first_oil"),
                            "Drill Days — Input": well_data.get("drill_days_input", 0),
                            "Drill Days — Alloc": well_data.get("drill_days_alloc", 0),
                            "Drill Days — Diff": well_data.get("drill_days_diff", 0),
                            "Comp Days — Input": well_data.get("comp_days_input", 0),
                            "Comp Days — Alloc": well_data.get("comp_days_alloc", 0),
                            "Comp Days — Diff": well_data.get("comp_days_diff", 0),
                        }
                    )

        if qc_rows:
            df = pd.DataFrame(qc_rows)
            self._write_dataframe_to_sheet(ws, df)
            self.formatter.apply_header_style(ws)

    def _format_executive_summary_data(
        self, development_data: Dict[str, Dict[str, Any]]
    ) -> List[Dict]:
        """Format data for executive summary"""
        exec_rows = []

        for dev_name, dev_data in development_data.items():
            metrics = dev_data.get("metrics", {})

            row = {
                "Project Name": dev_name,
                "TOTAL OIL BBL": metrics.get("total_oil_bbls", 0),
                "Facilities Cost USD": metrics.get("total_capex_facilities", 0),
                "DnC Drill Total USD": metrics.get("total_capex_drill", 0),
                "DnC Comp Total USD": metrics.get("total_capex_comp", 0),
                "DnC Total USD": metrics.get("total_capex_drill", 0)
                + metrics.get("total_capex_comp", 0),
                "DnC Total PreFO USD": metrics.get("dnc_prefo_gross", 0),
                "DnC Total PostFO USD": metrics.get("dnc_postfo_gross", 0),
                "DnC Net PreFO USD": metrics.get("dnc_prefo_net", 0),
                "DnC Net PostFO USD": metrics.get("dnc_postfo_net", 0),
                "NPV10 afterTax": metrics.get("npv", 0),
                "MIRR afterTax": metrics.get("mirr_annual", 0),
                "Price Model Used": metrics.get("price_model", "Flat WTI $75/bbl"),
            }
            exec_rows.append(row)

        return exec_rows

    def _format_project_summary_data(
        self, development_data: Dict[str, Dict[str, Any]]
    ) -> List[Dict]:
        """Format data for project summary"""
        proj_rows = []

        for dev_name, dev_data in development_data.items():
            metrics = dev_data.get("metrics", {})

            # Determine development system
            dev_type = dev_data.get("development_type", "subsea")
            dev_system = "subsea" if dev_type == "subsea" else "dry"

            # Calculate wells
            producer_count = dev_data.get("producer_count", 0)
            injector_count = (
                producer_count // 7
                if dev_system == "subsea" and producer_count >= 7
                else 0
            )
            pump_count = (
                producer_count // 8
                if dev_system == "subsea" and producer_count >= 8
                else 0
            )

            row = {
                "Project Name": dev_name,
                "TOTAL OIL BBL": metrics.get("total_oil_bbls", 0),
                "Facilities Host USD": metrics.get("host_usd", 0),
                "Facilities SURF USD": (
                    metrics.get("surf_usd", 0) if dev_system == "subsea" else 0
                ),
                "Subsea Pumps USD": (
                    metrics.get("pump_usd", 0) if dev_system == "subsea" else 0
                ),
                "Dry Well Systems USD": (
                    metrics.get("dry_sys_usd", 0) if dev_system != "subsea" else 0
                ),
                "Facilities Cost USD": metrics.get("total_capex_facilities", 0),
                "DnC Drill Total USD": metrics.get("total_capex_drill", 0),
                "DnC Comp Total USD": metrics.get("total_capex_comp", 0),
                "DnC Total USD": metrics.get("total_capex_drill", 0)
                + metrics.get("total_capex_comp", 0),
                "OPEX Total USD": metrics.get("total_opex", 0),
                "Producer Wells Used": producer_count,
                "Injector Wells": injector_count,
                "Pumps Count (8/prod)": pump_count,
                "DEV SYSTEM USED": dev_system,
                "ECON PATH USED": (
                    "dry: MODU->DRY at FO" if dev_system != "subsea" else "subsea: MODU"
                ),
                "FO Month": dev_data.get("first_oil"),
            }
            proj_rows.append(row)

        return proj_rows

    def _write_dataframe_to_sheet(self, worksheet, df: pd.DataFrame):
        """Write DataFrame to worksheet"""
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=str(col_name))

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                # Handle different data types
                if pd.isna(value):
                    cell_value = None
                elif isinstance(value, (pd.Timestamp, datetime, date)):
                    cell_value = value
                elif isinstance(value, (int, float)):
                    cell_value = value
                else:
                    cell_value = str(value)

                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

    def _add_rate_columns(self, worksheet, assumptions: Dict[str, Any]):
        """Add discount and MIRR rate columns to executive summary"""
        start_col = worksheet.max_column + 1

        headers = [
            "Discount Rate (Annual)",
            "MIRR Reinvest Rate (Annual)",
            "MIRR Finance Rate (Annual)",
        ]
        rates = {
            "Discount Rate (Annual)": assumptions.get("discount_rate_annual", 0.10),
            "MIRR Reinvest Rate (Annual)": assumptions.get(
                "mirr_reinvest_rate_annual", 0.10
            ),
            "MIRR Finance Rate (Annual)": assumptions.get(
                "mirr_finance_rate_annual", 0.10
            ),
        }

        for i, header in enumerate(headers):
            col = start_col + i
            worksheet.cell(row=1, column=col, value=header)

            # Add rate value for all data rows
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=col, value=rates[header])

    def _truncate_sheet_name(self, name: str, max_length: int = 31) -> str:
        """Truncate sheet name to Excel's 31 character limit"""
        return name[:max_length] if len(name) > max_length else name

    def _save_workbook(self, output_path: str):
        """Save workbook to file"""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)

    def _apply_v20_formatting(self, output_path: str):
        """Apply V20 specific formatting to saved workbook"""
        # Check if file exists
        if not Path(output_path).exists():
            logger.warning(f"File not found for formatting: {output_path}")
            return

        wb = load_workbook(output_path)

        # Format Executive Summary
        if "Executive Summary (V20)" in wb.sheetnames:
            ws = wb["Executive Summary (V20)"]
            self.formatter.format_numbers(
                ws,
                percent_headers={
                    "Discount Rate (Annual)",
                    "MIRR Reinvest Rate (Annual)",
                    "MIRR Finance Rate (Annual)",
                    "MIRR afterTax",
                },
                skip_headers={"Project Name", "Price Model Used"},
            )

        # Format Project Summary
        if "Project Summary (V20)" in wb.sheetnames:
            ws = wb["Project Summary (V20)"]
            self.formatter.format_numbers(
                ws,
                percent_headers=set(),
                skip_headers={"Project Name", "DEV SYSTEM USED", "ECON PATH USED"},
            )

        # Format QC sheets
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("QC"):
                ws = wb[sheet_name]
                self.formatter.format_numbers(
                    ws,
                    percent_headers=set(),
                    skip_headers={"DEV_NAME", "WELL", "SPUD", "TD", "FO_MONTH"},
                )

        # Save formatted workbook
        wb.save(output_path)
        logger.info(f"V20 formatting applied to {output_path}")
