"""
FDAS Excel Report Generator

Creates formatted Excel workbooks with financial summaries, cashflow
details, and development economics compatible with FDAS output format.

Author: WorldEnergyData Team
Date: 2025-10-03
Source: Ported from FDAS generate_financial_summary.py Excel output
"""

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


class ReportGenerationError(Exception):
    """Raised when report generation fails"""

    pass


class ExcelReportGenerator:
    """
    Generates formatted Excel reports for FDAS financial analysis.

    Creates workbooks with multiple sheets including project summary,
    cashflow details, and development-specific information.
    """

    # Excel formatting constants
    HEADER_FILL = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=14)
    CURRENCY_FORMAT = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    NUMBER_FORMAT = "#,##0"
    PERCENT_FORMAT = "0.00%"
    DATE_FORMAT = "yyyy-mm-dd"

    def __init__(self, output_path: str):
        """
        Initialize report generator.

        Args:
            output_path: Path for output Excel file
        """
        self.output_path = Path(output_path)
        self.workbook = Workbook()
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]

    def create_project_summary_sheet(
        self,
        project_name: str,
        financial_metrics: Dict[str, float],
        development_info: Dict[str, Any],
    ) -> None:
        """
        Create project summary sheet with key metrics.

        Args:
            project_name: Project/development name
            financial_metrics: Dict with NPV, MIRR, IRR, etc.
            development_info: Dict with development details

        Examples:
            >>> generator = ExcelReportGenerator('output.xlsx')
            >>> generator.create_project_summary_sheet(
            ...     'Anchor',
            ...     {'npv': 1500.0, 'mirr_annual': 0.18},
            ...     {'dev_system': 'subsea20', 'water_depth': 7500}
            ... )
        """
        ws = self.workbook.create_sheet("Project Summary")

        # Title
        ws["A1"] = "FDAS Financial Summary"
        ws["A1"].font = self.TITLE_FONT
        ws["A2"] = f"Project: {project_name}"
        ws["A3"] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'

        # Development Information
        row = 5
        ws[f"A{row}"] = "Development Information"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        info_items = [
            ("Development System", development_info.get("dev_system", "N/A")),
            ("Water Depth (ft)", development_info.get("water_depth", "N/A")),
            ("First Oil Date", development_info.get("first_oil_date", "N/A")),
            ("Well Count", development_info.get("well_count", "N/A")),
        ]

        for label, value in info_items:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        # Financial Metrics
        row += 2
        ws[f"A{row}"] = "Financial Metrics"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        # Header row
        ws[f"A{row}"] = "Metric"
        ws[f"B{row}"] = "Value"
        ws[f"A{row}"].fill = self.HEADER_FILL
        ws[f"B{row}"].fill = self.HEADER_FILL
        ws[f"A{row}"].font = self.HEADER_FONT
        ws[f"B{row}"].font = self.HEADER_FONT
        row += 1

        # Metrics
        metrics_items = [
            ("NPV (10% discount) $MM", financial_metrics.get("npv", 0.0)),
            ("MIRR (Annual)", financial_metrics.get("mirr_annual", 0.0)),
            ("IRR (Annual)", financial_metrics.get("irr_annual", 0.0)),
            ("Payback Period (years)", financial_metrics.get("payback_years", 0.0)),
        ]

        for label, value in metrics_items:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value

            # Format based on metric type
            if "$MM" in label:
                ws[f"B{row}"].number_format = self.CURRENCY_FORMAT
            elif "MIRR" in label or "IRR" in label:
                ws[f"B{row}"].number_format = self.PERCENT_FORMAT
            else:
                ws[f"B{row}"].number_format = self.NUMBER_FORMAT

            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def create_cashflow_detail_sheet(
        self, cashflow_df: pd.DataFrame, sheet_name: str = "Monthly Cashflow"
    ) -> None:
        """
        Create detailed monthly cashflow sheet.

        Args:
            cashflow_df: DataFrame with monthly cashflow data
            sheet_name: Name for the sheet

        Examples:
            >>> generator = ExcelReportGenerator('output.xlsx')
            >>> generator.create_cashflow_detail_sheet(cashflow_df)
        """
        ws = self.workbook.create_sheet(sheet_name)

        # Title
        ws["A1"] = "Monthly Cashflow Detail"
        ws["A1"].font = self.TITLE_FONT

        # Write DataFrame starting at row 3
        start_row = 3

        # Header row
        headers = list(cashflow_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(cashflow_df.values, start=start_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format based on column name
                col_name = headers[col_idx - 1]
                if (
                    "usd" in col_name.lower()
                    or "capex" in col_name.lower()
                    or "opex" in col_name.lower()
                ):
                    cell.number_format = self.CURRENCY_FORMAT
                elif "bbl" in col_name.lower() or "production" in col_name.lower():
                    cell.number_format = self.NUMBER_FORMAT

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                max(len(str(header)) + 2, 15)
            )

    def create_capex_summary_sheet(
        self, capex_breakdown: Dict[str, float], sheet_name: str = "CAPEX Summary"
    ) -> None:
        """
        Create CAPEX breakdown summary sheet.

        Args:
            capex_breakdown: Dict with CAPEX categories and amounts
            sheet_name: Name for the sheet

        Examples:
            >>> generator = ExcelReportGenerator('output.xlsx')
            >>> generator.create_capex_summary_sheet({
            ...     'Host Platform': 450.0,
            ...     'Drilling & Completion': 320.0,
            ...     'Subsea Facilities': 180.0
            ... })
        """
        ws = self.workbook.create_sheet(sheet_name)

        # Title
        ws["A1"] = "CAPEX Breakdown"
        ws["A1"].font = self.TITLE_FONT

        # Header row
        ws["A3"] = "Category"
        ws["B3"] = "Amount ($MM)"
        ws["A3"].fill = self.HEADER_FILL
        ws["B3"].fill = self.HEADER_FILL
        ws["A3"].font = self.HEADER_FONT
        ws["B3"].font = self.HEADER_FONT

        # Data
        row = 4
        total = 0.0
        for category, amount in capex_breakdown.items():
            ws[f"A{row}"] = category
            ws[f"B{row}"] = amount
            ws[f"B{row}"].number_format = self.CURRENCY_FORMAT
            total += amount
            row += 1

        # Total row
        ws[f"A{row}"] = "Total CAPEX"
        ws[f"B{row}"] = total
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].number_format = self.CURRENCY_FORMAT

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def create_production_summary_sheet(
        self, production_df: pd.DataFrame, sheet_name: str = "Production Summary"
    ) -> None:
        """
        Create production summary sheet.

        Args:
            production_df: DataFrame with production data
            sheet_name: Name for the sheet
        """
        ws = self.workbook.create_sheet(sheet_name)

        # Title
        ws["A1"] = "Production Summary"
        ws["A1"].font = self.TITLE_FONT

        # Write DataFrame
        start_row = 3

        # Header
        headers = list(production_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        # Data
        for row_idx, row_data in enumerate(production_df.values, start=start_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format numbers
                col_name = headers[col_idx - 1]
                if "bbl" in col_name.lower() or "volume" in col_name.lower():
                    cell.number_format = self.NUMBER_FORMAT

        # Auto-adjust columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                15
            )

    def save(self) -> None:
        """
        Save workbook to file.

        Raises:
            ReportGenerationError: If save fails
        """
        try:
            self.workbook.save(self.output_path)
        except Exception as e:
            raise ReportGenerationError(f"Failed to save workbook: {e}")


def format_financial_summary(financial_metrics: Dict[str, float]) -> pd.DataFrame:
    """
    Format financial metrics as DataFrame for reporting.

    Args:
        financial_metrics: Dict with financial metrics

    Returns:
        Formatted DataFrame

    Examples:
        >>> metrics = {'npv': 1500.0, 'mirr_annual': 0.18, 'irr_annual': 0.22}
        >>> df = format_financial_summary(metrics)
    """
    data = []

    metric_display = {
        "npv": ("NPV (10% discount) $MM", lambda x: f"${x:,.2f}"),
        "mirr_annual": ("MIRR (Annual)", lambda x: f"{x:.2%}"),
        "irr_annual": ("IRR (Annual)", lambda x: f"{x:.2%}"),
        "payback_years": ("Payback Period (years)", lambda x: f"{x:.1f}"),
        "total_capex": ("Total CAPEX $MM", lambda x: f"${x:,.2f}"),
        "total_opex": ("Total OPEX $MM", lambda x: f"${x:,.2f}"),
        "cumulative_production": ("Cumulative Oil (MMBO)", lambda x: f"{x:,.1f}"),
    }

    for key, (label, formatter) in metric_display.items():
        if key in financial_metrics:
            data.append(
                {
                    "Metric": label,
                    "Value": financial_metrics[key],
                    "Formatted": formatter(financial_metrics[key]),
                }
            )

    return pd.DataFrame(data)


def create_project_summary_sheet(
    workbook: Workbook,
    project_name: str,
    metrics: Dict[str, float],
    info: Dict[str, Any],
) -> None:
    """
    Convenience function to create project summary sheet.

    Args:
        workbook: openpyxl Workbook
        project_name: Project name
        metrics: Financial metrics
        info: Development information

    Examples:
        >>> wb = Workbook()
        >>> create_project_summary_sheet(wb, 'Anchor', metrics, info)
    """
    generator = ExcelReportGenerator.__new__(ExcelReportGenerator)
    generator.workbook = workbook
    generator.create_project_summary_sheet(project_name, metrics, info)


class FDASReportBuilder:
    """
    High-level builder for complete FDAS reports.

    Orchestrates creation of all report sheets with proper formatting.
    """

    def __init__(self, output_path: str):
        """
        Initialize report builder.

        Args:
            output_path: Path for output Excel file
        """
        self.generator = ExcelReportGenerator(output_path)

    def build_complete_report(
        self,
        project_name: str,
        financial_metrics: Dict[str, float],
        development_info: Dict[str, Any],
        cashflow_df: pd.DataFrame,
        production_df: Optional[pd.DataFrame] = None,
        capex_breakdown: Optional[Dict[str, float]] = None,
    ) -> None:
        """
        Build complete FDAS report with all sheets.

        Args:
            project_name: Project name
            financial_metrics: Financial metrics dict
            development_info: Development information
            cashflow_df: Monthly cashflow DataFrame
            production_df: Optional production summary
            capex_breakdown: Optional CAPEX breakdown

        Examples:
            >>> builder = FDASReportBuilder('anchor_analysis.xlsx')
            >>> builder.build_complete_report(
            ...     'Anchor',
            ...     metrics,
            ...     info,
            ...     cashflow_df,
            ...     production_df,
            ...     capex_breakdown
            ... )
            >>> builder.save()
        """
        # Project summary
        self.generator.create_project_summary_sheet(
            project_name, financial_metrics, development_info
        )

        # Cashflow detail
        self.generator.create_cashflow_detail_sheet(cashflow_df)

        # Production summary
        if production_df is not None:
            self.generator.create_production_summary_sheet(production_df)

        # CAPEX breakdown
        if capex_breakdown is not None:
            self.generator.create_capex_summary_sheet(capex_breakdown)

    def save(self) -> None:
        """Save the report."""
        self.generator.save()
