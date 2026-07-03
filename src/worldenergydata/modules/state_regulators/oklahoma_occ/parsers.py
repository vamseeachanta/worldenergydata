"""Parsers for Oklahoma OCC completion workbooks (#740)."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

CORE_COLUMNS = {
    "API_Number",
    "Completion_No",
    "Test_Date",
    "Shut_In_Pressure",
    "Flow_Tubing_Pressure",
}

CONTEXT_COLUMNS = [
    "Well_Name",
    "Well_Number",
    "Operator_Name",
    "Operator_Number",
    "County",
    "Formation_Name",
    "Formation_Code",
]

NUMERIC_COLUMNS = [
    "Shut_In_Pressure",
    "Flow_Tubing_Pressure",
    "Gas_MCF_Per_Day",
    "Oil_BBL_Per_Day",
    "Water_BBL_Per_Day",
    "Gas_Oil_Ratio",
    "Oil_Gravity",
    "Measured_Total_Depth",
    "True_Vertical_Depth",
    "Total_Depth",
    "Formation_Depth",
    "Perforated_Top_Depth",
    "Perforated_Bottom_Depth",
]

READ_COLUMNS = sorted(CORE_COLUMNS | set(CONTEXT_COLUMNS) | set(NUMERIC_COLUMNS))
TEXT_COLUMNS = sorted(
    set(READ_COLUMNS)
    - set(NUMERIC_COLUMNS)
    - {"API_Number", "Completion_No", "Test_Date"}
)

OUTPUT_COLUMNS = [
    "state",
    "well_key",
    "api_number",
    "api14",
    "completion_no",
    "well_name",
    "well_number",
    "operator",
    "operator_number",
    "county",
    "field",
    "formation_name",
    "formation_code",
    "test_date",
    "test_year",
    "test_type",
    "pressure_psig_reported",
    "pressure_psia",
    "pressure_kind",
    "flow_tubing_pressure_psig",
    "gas_mcf_per_day",
    "oil_bbl_per_day",
    "water_bbl_per_day",
    "reference_depth_ft",
    "reference_depth_source",
    "gradient_psi_ft",
    "gradient_method",
    "is_earliest_observation",
    "screen_observation_priority",
    "source_file",
]


def read_completion_workbook(path: str | Path) -> pd.DataFrame:
    """Read an OCC completion workbook into a typed DataFrame."""
    frame = _read_selected_columns(Path(path), READ_COLUMNS)
    _validate_columns(frame, CORE_COLUMNS, Path(path).name)
    for column in NUMERIC_COLUMNS:
        if column in frame:
            frame[column] = pd.to_numeric(frame[column], errors="coerce")
    for column in TEXT_COLUMNS:
        if column in frame:
            frame[column] = frame[column].map(_clean_text).astype("string")
    frame["API_Number"] = frame["API_Number"].map(_clean_identifier).astype("string")
    frame["Completion_No"] = (
        frame["Completion_No"].map(_clean_completion_no).astype("string")
    )
    frame["Test_Date"] = pd.to_datetime(frame["Test_Date"], errors="coerce")
    return frame


def _read_selected_columns(path: Path, wanted_columns: list[str]) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return pd.DataFrame(columns=wanted_columns)
        column_indexes = {
            name: index for index, name in enumerate(header) if name in wanted_columns
        }
        records = []
        for row in rows:
            records.append(
                {
                    column: row[index] if index < len(row) else None
                    for column, index in column_indexes.items()
                }
            )
    finally:
        workbook.close()
    return pd.DataFrame.from_records(records, columns=list(column_indexes))


def build_pressure_observations(
    completions: pd.DataFrame, settings: dict
) -> pd.DataFrame:
    """Build curated pressure observations from structured OCC completions."""
    frame = completions.copy()
    _validate_columns(frame, CORE_COLUMNS, "OCC completions")
    for column in CONTEXT_COLUMNS + NUMERIC_COLUMNS:
        if column not in frame:
            frame[column] = pd.NA
    for column in NUMERIC_COLUMNS:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    frame["API_Number"] = frame["API_Number"].map(_clean_identifier)
    frame["Completion_No"] = frame["Completion_No"].map(_clean_completion_no)
    test_date = pd.to_datetime(frame["Test_Date"], errors="coerce")
    pressure = _select_pressure(frame)
    depth, depth_source = _select_reference_depth(frame, settings["depth_priority"])
    test_year = test_date.dt.year
    min_year, max_year = _year_window(settings)
    in_window = test_year.between(min_year, max_year)
    usable = (
        pressure["pressure_psig"].gt(0) & depth.gt(0) & test_date.notna() & in_window
    )
    filtered = frame[usable].copy()
    pressure = pressure[usable].reset_index(drop=True)
    depth = depth[usable].reset_index(drop=True)
    depth_source = depth_source[usable].reset_index(drop=True)
    test_date = test_date[usable].reset_index(drop=True)

    observations = pd.DataFrame(
        {
            "state": "OK",
            "well_key": filtered["API_Number"].to_numpy(),
            "api_number": filtered["API_Number"].to_numpy(),
            "api14": [
                _api14(api, completion)
                for api, completion in zip(
                    filtered["API_Number"], filtered["Completion_No"], strict=False
                )
            ],
            "completion_no": filtered["Completion_No"].to_numpy(),
            "well_name": filtered["Well_Name"].to_numpy(),
            "well_number": filtered["Well_Number"].to_numpy(),
            "operator": filtered["Operator_Name"].to_numpy(),
            "operator_number": filtered["Operator_Number"].to_numpy(),
            "county": filtered["County"].to_numpy(),
            "field": filtered["Formation_Name"].to_numpy(),
            "formation_name": filtered["Formation_Name"].to_numpy(),
            "formation_code": filtered["Formation_Code"].to_numpy(),
            "test_date": test_date,
            "test_year": test_date.dt.year.astype("Int64"),
            "test_type": settings["test_type"],
            "pressure_psig_reported": pressure["pressure_psig"].to_numpy(),
            "pressure_psia": (
                pressure["pressure_psig"] + settings["atmospheric_psi"]
            ).to_numpy(),
            "pressure_kind": pressure["pressure_kind"].to_numpy(),
            "flow_tubing_pressure_psig": filtered["Flow_Tubing_Pressure"].to_numpy(),
            "gas_mcf_per_day": filtered["Gas_MCF_Per_Day"].to_numpy(),
            "oil_bbl_per_day": filtered["Oil_BBL_Per_Day"].to_numpy(),
            "water_bbl_per_day": filtered["Water_BBL_Per_Day"].to_numpy(),
            "reference_depth_ft": depth.to_numpy(),
            "reference_depth_source": depth_source.to_numpy(),
            "gradient_psi_ft": (
                (pressure["pressure_psig"] + settings["atmospheric_psi"]) / depth
            ).to_numpy(),
            "gradient_method": settings["gradient_method"],
            "source_file": "completions-wells-formations-base.xlsx",
        }
    )
    observations["is_earliest_observation"] = _earliest_flags(observations)
    observations["screen_observation_priority"] = (
        ~observations["is_earliest_observation"].astype(bool)
    ).astype(int)
    observations["is_earliest_observation"] = observations[
        "is_earliest_observation"
    ].astype(object)
    return observations[OUTPUT_COLUMNS]


def build_quality_stats(
    completions: pd.DataFrame, observations: pd.DataFrame, settings: dict | None = None
) -> dict:
    pressure = _select_pressure(completions)
    depth, _ = _select_reference_depth(
        completions,
        [
            "True_Vertical_Depth",
            "Formation_Depth",
            "Measured_Total_Depth",
            "Total_Depth",
        ],
    )
    test_date = pd.to_datetime(completions["Test_Date"], errors="coerce")
    test_year = test_date.dt.year
    min_year, max_year = _year_window(settings or {})
    in_window = test_year.between(min_year, max_year)
    has_pressure = pressure["pressure_psig"].gt(0)
    has_depth = depth.gt(0)
    return {
        "input_rows": int(len(completions)),
        "curated_count": int(len(observations)),
        "filtered_missing_pressure_count": int((~has_pressure).sum()),
        "filtered_missing_depth_count": int((has_pressure & ~has_depth).sum()),
        "filtered_missing_test_date_count": int(
            (has_pressure & has_depth & test_date.isna()).sum()
        ),
        "filtered_out_of_window_test_year_count": int(
            (has_pressure & has_depth & test_date.notna() & ~in_window).sum()
        ),
        "test_year_window": [int(min_year), int(max_year)],
        "wells_with_pressure_observation": int(observations["well_key"].nunique()),
        "completion_observation_count": int(
            observations[["well_key", "completion_no"]].drop_duplicates().shape[0]
        ),
        "pressure_kind_counts": _counts(observations["pressure_kind"]),
        "reference_depth_source_counts": _counts(
            observations["reference_depth_source"]
        ),
        "test_year_range": _year_range(observations),
    }


def _select_pressure(frame: pd.DataFrame) -> pd.DataFrame:
    shut_in = pd.to_numeric(frame["Shut_In_Pressure"], errors="coerce")
    flowing = pd.to_numeric(frame["Flow_Tubing_Pressure"], errors="coerce")
    uses_shut_in = shut_in.gt(0)
    pressure = shut_in.where(uses_shut_in, flowing)
    kind = pd.Series(
        np.where(uses_shut_in, "WHP_shut_in", "WHP_flowing_tubing"),
        index=frame.index,
    )
    return pd.DataFrame({"pressure_psig": pressure, "pressure_kind": kind})


def _select_reference_depth(
    frame: pd.DataFrame, depth_priority: list[str]
) -> tuple[pd.Series, pd.Series]:
    depth = pd.Series(np.nan, index=frame.index, dtype="float64")
    source = pd.Series(pd.NA, index=frame.index, dtype="object")
    for column in depth_priority:
        if column in frame:
            values = pd.to_numeric(frame[column], errors="coerce")
        else:
            values = pd.Series(np.nan, index=frame.index, dtype="float64")
        use = depth.isna() & values.gt(0)
        depth.loc[use] = values.loc[use]
        source.loc[use] = column
    return depth, source


def _earliest_flags(observations: pd.DataFrame) -> pd.Series:
    order = observations.sort_values(["well_key", "test_date", "completion_no"])
    first_index = order.groupby("well_key", sort=False).head(1).index
    flags = pd.Series(False, index=observations.index, dtype=object)
    flags.loc[first_index] = True
    return flags


def _clean_identifier(value: object) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.zfill(10) if text else None


def _clean_completion_no(value: object) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.zfill(2) if text else None


def _clean_text(value: object) -> str | None:
    if pd.isna(value):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") else text


def _api14(api_number: str | None, completion_no: str | None) -> str | None:
    if not api_number or not completion_no:
        return None
    return f"{api_number}00{completion_no}"


def _validate_columns(
    frame: pd.DataFrame, required: set[str], source_name: str
) -> None:
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(
            f"{source_name} missing required OCC completion columns: "
            + ", ".join(missing)
        )


def _counts(series: pd.Series) -> dict:
    return {str(key): int(value) for key, value in series.value_counts().items()}


def _year_range(observations: pd.DataFrame) -> list[int] | None:
    if observations.empty:
        return None
    return [
        int(observations["test_year"].min()),
        int(observations["test_year"].max()),
    ]


def _year_window(settings: dict) -> tuple[int, int]:
    min_year = int(settings.get("min_test_year", 1900))
    if "max_test_year" in settings:
        max_year = int(settings["max_test_year"])
    else:
        max_year = datetime.now(timezone.utc).year + int(
            settings.get("max_future_years", 0)
        )
    return min_year, max_year
