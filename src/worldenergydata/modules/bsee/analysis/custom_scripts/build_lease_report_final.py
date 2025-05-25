import pandas as pd
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime
from loguru import logger

class BseeCustomAnalysis:

    def router(self, cfg):
        self.build_report(cfg)
        return cfg

    def read_file(self, filepath):
        df = pd.read_pickle(filepath)
        for column in df.select_dtypes(include="object"):
            df[column] = df[column].str.strip().str.strip('"')
        return df

    def build_report(self, cfg):
        lease = cfg['data']['groups']['lease']
        lease = lease.strip().upper()
        if not lease.startswith("G"):
            lease = "G" + lease

        # 1) main → SN_WAR & API
        war_filepath = cfg['data']['groups']['war']
        war_df = self.read_file(war_filepath)
        lease_war_df = war_df[war_df["BOTM_LEASE_NUM"] == lease]
        if lease_war_df.empty:
            logger.error(f"No records for lease {lease}")
            return

        war_sn_list = lease_war_df["SN_WAR"].dropna().tolist()
        api_list = lease_war_df["API_WELL_NUMBER"].dropna().tolist()

        # 2) boreholes by API
        borehole_filepath = cfg['data']['groups']['boreholes']
        borehole_df = self.read_file(borehole_filepath)
        api12_filtered_borehole_df = borehole_df[borehole_df["API_WELL_NUMBER"].isin(api_list)].copy()
        api12_filtered_borehole_df = api12_filtered_borehole_df[[
            "API_WELL_NUMBER",
            "WELL_SPUD_DATE",
            "TOTAL_DEPTH_DATE",
            "BH_TOTAL_MD",
            "WELL_BORE_TVD"
        ]]

        # 3) prop for mud weight
        war_prop_filepath = cfg['data']['groups']['war_prop']
        mud_weight_df = self.read_file(war_prop_filepath)
        sn_war_mud_weight_df = mud_weight_df[mud_weight_df["SN_WAR"].isin(war_sn_list)]
        sn_war_mud_weight_df = sn_war_mud_weight_df[["SN_WAR", "DRILL_FLUID_WGT"]].rename(
            columns={"DRILL_FLUID_WGT": "Max Mud Weight (ppg)"}
        )

        # Merge and select max mud weight per API
        merged_df = lease_war_df[["SN_WAR", "API_WELL_NUMBER"]].drop_duplicates()
        merged_df = merged_df.merge(api12_filtered_borehole_df, on="API_WELL_NUMBER", how="left")
        merged_df = merged_df.merge(sn_war_mud_weight_df, on="SN_WAR", how="left")
        merged_df["Max Mud Weight (ppg)"] = pd.to_numeric(merged_df["Max Mud Weight (ppg)"], errors="coerce")
        max_weight_indices = merged_df.groupby("API_WELL_NUMBER")["Max Mud Weight (ppg)"].idxmax()
        merged_df = merged_df.loc[max_weight_indices].reset_index(drop=True)

        # Convert dates to datetime
        merged_df["WELL_SPUD_DATE"] = pd.to_datetime(merged_df["WELL_SPUD_DATE"], errors='coerce')
        merged_df["TOTAL_DEPTH_DATE"] = pd.to_datetime(merged_df["TOTAL_DEPTH_DATE"], errors='coerce')

        # build Excel
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = lease

        headers = [
            "SN_WAR",
            "API_WELL_NUMBER",
            "Spud Date",
            "Total Depth Date",
            "Drilling Days",
            "TMD",
            "TVD",
            "Max Mud Weight (ppg)",
            "Bottom Hole Pressure"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(1, col_idx, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal="center")

        for row_idx, row in enumerate(merged_df.itertuples(index=False), start=2):
            sn_war, api_num, spud_date, td_date, tmd, tvd, mud_weight = row

            worksheet.cell(row_idx, 1, sn_war)
            worksheet.cell(row_idx, 2, api_num)

            worksheet.cell(row_idx, 3, spud_date if pd.notna(spud_date) else "")
            worksheet.cell(row_idx, 4, td_date if pd.notna(td_date) else "")

            if pd.notna(spud_date) and pd.notna(td_date):
                worksheet.cell(row_idx, 5, (td_date - spud_date).days)
            else:
                worksheet.cell(row_idx, 5, "")

            worksheet.cell(row_idx, 6, pd.to_numeric(tmd, errors="coerce"))
            worksheet.cell(row_idx, 7, pd.to_numeric(tvd, errors="coerce"))
            worksheet.cell(row_idx, 8, pd.to_numeric(mud_weight, errors="coerce"))
            worksheet.cell(row_idx, 9, f"=H{row_idx}*0.052*G{row_idx}")

        # Apply formatting
        column_formatting = {
            1: ("@", 12),
            2: ("0", 14),
            3: ("mm/dd/yyyy", 12),
            4: ("mm/dd/yyyy", 12),
            5: ("0", 10),
            6: ("#,##0", 12),
            7: ("#,##0", 12),
            8: ("0.0", 15),
            9: ("#,##0", 12)
        }
        for col_idx, (format_code, width) in column_formatting.items():
            column_letter = worksheet.cell(1, col_idx).column_letter
            worksheet.column_dimensions[column_letter].width = width
            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(wrap_text=True)
                if cell.row > 1 and format_code != "@":
                    cell.number_format = format_code

        result_path = cfg['Analysis']['result_folder']
        label = cfg['meta']['label']
        output_filename = f"julia_summary_{lease}.xlsx"
        workbook.save(os.path.join(result_path, output_filename))
        logger.info(f"✅ Wrote fully formatted summary to {output_filename}")
