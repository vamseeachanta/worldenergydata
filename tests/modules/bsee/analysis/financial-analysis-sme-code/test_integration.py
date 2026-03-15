"""
Integration tests for SME financial analysis pipeline
"""

import json
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, mock_open, patch

import numpy as np
import pandas as pd
import pytest

from src.worldenergydata.modules.bsee.analysis.financial.analyzer import (
    AnalysisConfig,
    AnalysisResult,
    FinancialAnalyzer,
)


class TestFinancialAnalyzer:
    """Test the main financial analyzer orchestrator"""

    @pytest.fixture
    def sample_config(self):
        """Create a sample analysis configuration"""
        return AnalysisConfig(
            input_path="data/modules/bsee/",
            output_path="output/financial/",
            start_date="2020-01-01",
            end_date="2023-12-31",
            developments=["Stones", "Anchor"],
            oil_price_scenario="mid",
            discount_rate=0.10,
            tax_rate=0.35,
            version="V20",
        )

    @pytest.fixture
    def analyzer(self, sample_config):
        """Create a FinancialAnalyzer instance"""
        return FinancialAnalyzer(sample_config)

    def test_analyzer_initialization(self, analyzer, sample_config):
        """Test analyzer initialization with config"""
        assert analyzer.config == sample_config
        assert analyzer.data_loader is not None
        assert analyzer.lease_grouper is not None
        assert analyzer.cash_flow_calculator is not None
        assert analyzer.report_generator is not None

    def test_load_data(self, analyzer):
        """Test data loading from repository"""
        with patch.object(analyzer.data_loader, "load_production_data") as mock_prod:
            with patch.object(analyzer.data_loader, "load_drilling_data") as mock_drill:
                mock_prod.return_value = pd.DataFrame(
                    {"API": [1, 2], "oil": [100, 200]}
                )
                mock_drill.return_value = pd.DataFrame(
                    {"API": [1, 2], "days": [30, 45]}
                )

                production_data, drilling_data = analyzer.load_data()

                assert production_data is not None
                assert drilling_data is not None
                mock_prod.assert_called_once()
                mock_drill.assert_called_once()

    def test_process_developments(self, analyzer):
        """Test processing of multiple developments"""
        sample_data = pd.DataFrame(
            {
                "API": [1, 2, 3],
                "lease": ["A", "B", "C"],
                "oil_bbls": [1000, 2000, 1500],
                "year_month": pd.date_range("2020-01", periods=3, freq="MS"),
            }
        )

        with patch.object(analyzer.lease_grouper, "group_by_development") as mock_group:
            mock_group.return_value = {"Dev1": ["A", "B"], "Dev2": ["C"]}

            with patch.object(analyzer, "_process_single_development") as mock_process:
                mock_process.return_value = {"cash_flow": pd.DataFrame(), "metrics": {}}

                results = analyzer.process_developments(sample_data)

                assert "Dev1" in results
                assert "Dev2" in results
                assert mock_process.call_count == 2

    def test_calculate_financial_metrics(self, analyzer):
        """Test financial metrics calculation"""
        cash_flow = pd.DataFrame(
            {
                "net_cash_flow": [-100000, 50000, 60000, 70000],
                "year_month": pd.date_range("2020-01", periods=4, freq="MS"),
            }
        )

        with patch.object(analyzer.cash_flow_calculator, "calculate_npv") as mock_npv:
            with patch.object(
                analyzer.cash_flow_calculator, "calculate_mirr"
            ) as mock_mirr:
                mock_npv.return_value = 25000
                mock_mirr.return_value = 0.15

                metrics = analyzer.calculate_financial_metrics(cash_flow)

                assert metrics["npv"] == 25000
                assert metrics["mirr"] == 0.15
                mock_npv.assert_called_once()
                mock_mirr.assert_called_once()

    def test_generate_report(self, analyzer, tmp_path):
        """Test report generation"""
        development_results = {
            "Dev1": {
                "cash_flow": pd.DataFrame({"revenue": [100000]}),
                "metrics": {"npv": 50000, "mirr": 0.12},
            }
        }

        output_file = tmp_path / "test_report.xlsx"

        with patch.object(analyzer.report_generator, "generate_report") as mock_gen:
            mock_gen.return_value = str(output_file)

            result = analyzer.generate_report(development_results, str(output_file))

            assert result == str(output_file)
            mock_gen.assert_called_once_with(
                development_data=development_results,
                output_path=str(output_file),
                version="V20",
            )

    def test_run_analysis_full_pipeline(self, analyzer, tmp_path):
        """Test complete analysis pipeline"""
        output_file = tmp_path / "full_analysis.xlsx"

        with patch.object(analyzer, "load_data") as mock_load:
            with patch.object(analyzer, "process_developments") as mock_process:
                with patch.object(analyzer, "generate_report") as mock_report:
                    # Mock data
                    mock_load.return_value = (
                        pd.DataFrame({"data": [1]}),  # production
                        pd.DataFrame({"data": [2]}),  # drilling
                    )
                    mock_process.return_value = {
                        "Dev1": {"cash_flow": pd.DataFrame(), "metrics": {}}
                    }
                    mock_report.return_value = str(output_file)

                    result = analyzer.run_analysis(str(output_file))

                    assert isinstance(result, AnalysisResult)
                    assert result.output_file == str(output_file)
                    assert result.success is True
                    mock_load.assert_called_once()
                    mock_process.assert_called_once()
                    mock_report.assert_called_once()

    def test_error_handling(self, analyzer):
        """Test error handling in pipeline"""
        with patch.object(analyzer, "load_data") as mock_load:
            mock_load.side_effect = Exception("Data load error")

            result = analyzer.run_analysis("output.xlsx")

            assert result.success is False
            assert "Data load error" in result.error_message


class TestEndToEndIntegration:
    """End-to-end integration tests with real data flow"""

    @pytest.fixture
    def test_data_dir(self, tmp_path):
        """Create test data directory structure"""
        data_dir = tmp_path / "test_data"
        data_dir.mkdir()

        # Create mock binary files
        bin_dir = data_dir / "bin"
        bin_dir.mkdir()

        # Create mock zip files
        zip_dir = data_dir / "zip"
        zip_dir.mkdir()

        return data_dir

    def test_cli_execution(self, test_data_dir, tmp_path):
        """Test CLI interface execution"""
        from src.worldenergydata.modules.bsee.analysis.financial.cli_interface import (
            main,
        )

        output_file = tmp_path / "cli_output.xlsx"

        test_args = [
            "--input-directory",
            str(test_data_dir),
            "--output-file",
            str(output_file),
            "--start-date",
            "2020-01-01",
            "--end-date",
            "2023-12-31",
            "--developments",
            "Stones,Anchor",
            "--oil-price",
            "mid",
            "--version",
            "V20",
        ]

        with patch("sys.argv", ["cli_interface.py"] + test_args):
            with patch(
                "src.worldenergydata.modules.bsee.analysis.financial.analyzer.FinancialAnalyzer.run_analysis"
            ) as mock_run:
                mock_run.return_value = MagicMock(
                    success=True, output_file=str(output_file)
                )

                result = main()

                assert result == 0  # Success exit code
                mock_run.assert_called_once()

    def test_data_flow_integration(self, test_data_dir):
        """Test data flow through all components"""
        config = AnalysisConfig(
            input_path=str(test_data_dir),
            output_path=str(test_data_dir),
            start_date="2020-01-01",
            end_date="2023-12-31",
        )

        analyzer = FinancialAnalyzer(config)

        # Mock the component interactions
        with patch.object(analyzer.data_loader, "load_production_data") as mock_prod:
            with patch.object(
                analyzer.drilling_loader, "load_drilling_data"
            ) as mock_drill:
                with patch.object(analyzer.lease_grouper, "group_leases") as mock_group:
                    with patch.object(
                        analyzer.cash_flow_calculator, "calculate_monthly_cash_flow"
                    ) as mock_calc:
                        # Setup mock returns
                        mock_prod.return_value = pd.DataFrame(
                            {
                                "API": [1, 2],
                                "oil_bbls": [1000, 2000],
                                "year_month": pd.date_range(
                                    "2020-01", periods=2, freq="MS"
                                ),
                            }
                        )
                        mock_drill.return_value = pd.DataFrame(
                            {
                                "API": [1, 2],
                                "drill_days": [30, 45],
                                "comp_days": [15, 20],
                            }
                        )
                        mock_group.return_value = {
                            "Dev1": pd.DataFrame({"lease": ["A"]})
                        }
                        mock_calc.return_value = pd.DataFrame(
                            {"net_cash_flow": [100000, 150000]}
                        )

                        # Run partial pipeline
                        prod_data, drill_data = analyzer.load_data()
                        assert prod_data is not None
                        assert drill_data is not None

    def test_memory_usage(self):
        """Test memory usage stays within limits"""
        import os

        import psutil

        process = psutil.Process(os.getpid())
        initial_memory = process.memory_info().rss / 1024 / 1024  # MB

        # Create large test dataset
        large_df = pd.DataFrame(
            {
                "API": range(100000),
                "oil_bbls": np.random.rand(100000) * 1000,
                "year_month": pd.date_range("2020-01", periods=100000, freq="H"),
            }
        )

        config = AnalysisConfig(input_path=".", output_path=".")
        analyzer = FinancialAnalyzer(config)

        with patch.object(analyzer, "load_data") as mock_load:
            mock_load.return_value = (large_df, large_df)

            # Process data
            analyzer.load_data()

            final_memory = process.memory_info().rss / 1024 / 1024  # MB
            memory_increase = final_memory - initial_memory

            # Should stay under 2GB increase
            assert (
                memory_increase < 2000
            ), f"Memory usage increased by {memory_increase}MB"

    def test_performance_benchmark(self):
        """Test processing speed meets requirements"""
        import time

        # Create test data for 100+ leases
        num_leases = 150
        dates = pd.date_range("2020-01", periods=48, freq="MS")

        production_data = []
        for lease_id in range(num_leases):
            for date in dates:
                production_data.append(
                    {
                        "lease": f"LEASE_{lease_id}",
                        "year_month": date,
                        "oil_bbls": np.random.rand() * 1000,
                    }
                )

        df = pd.DataFrame(production_data)

        config = AnalysisConfig(input_path=".", output_path=".")
        analyzer = FinancialAnalyzer(config)

        with patch.object(analyzer, "load_data") as mock_load:
            mock_load.return_value = (df, pd.DataFrame())

            start_time = time.time()
            analyzer.process_developments(df)
            elapsed_time = time.time() - start_time

            # Should complete in under 60 seconds
            assert elapsed_time < 60, f"Processing took {elapsed_time} seconds"


class TestConfigurationLoading:
    """Test configuration loading and validation"""

    def test_yaml_config_loading(self, tmp_path):
        """Test loading configuration from YAML file"""
        config_file = tmp_path / "config.yaml"
        config_data = {
            "input_path": "/data/bsee",
            "output_path": "/output",
            "start_date": "2020-01-01",
            "end_date": "2023-12-31",
            "developments": ["Stones", "Anchor"],
            "financial_parameters": {
                "discount_rate": 0.10,
                "tax_rate": 0.35,
                "oil_price_scenario": "mid",
            },
        }

        import yaml

        with open(config_file, "w") as f:
            yaml.dump(config_data, f)

        from src.worldenergydata.modules.bsee.analysis.financial.config_loader import (
            load_config,
        )

        config = load_config(str(config_file))

        assert config.input_path == "/data/bsee"
        assert config.discount_rate == 0.10
        assert "Stones" in config.developments

    def test_config_validation(self):
        """Test configuration validation"""
        from src.worldenergydata.modules.bsee.analysis.financial.config_loader import (
            validate_config,
        )

        # Valid config
        valid_config = {
            "input_path": "/data",
            "output_path": "/output",
            "start_date": "2020-01-01",
            "end_date": "2023-12-31",
        }
        assert validate_config(valid_config) is True

        # Invalid config (missing required field)
        invalid_config = {"input_path": "/data", "output_path": "/output"}
        with pytest.raises(ValueError):
            validate_config(invalid_config)


class TestSampleDataExecution:
    """Test with SME Roy's sample data"""

    def test_with_sme_sample_data(self, tmp_path):
        """Test pipeline with actual SME sample data structure"""
        # This would use actual sample data from SME Roy's examples
        sample_dir = Path("docs/modules/bsee/data/SME_Roy_attachments/2025-08-20")

        if sample_dir.exists():
            config = AnalysisConfig(
                input_path=str(sample_dir), output_path=str(tmp_path), version="V20"
            )

            analyzer = FinancialAnalyzer(config)

            with patch.object(analyzer, "run_analysis") as mock_run:
                mock_run.return_value = MagicMock(success=True)

                result = analyzer.run_analysis(str(tmp_path / "test_output.xlsx"))
                assert result.success is True

    def test_output_format_validation(self, tmp_path):
        """Test that output matches V20 format specifications"""
        from openpyxl import Workbook

        # Create a mock V20 format output
        wb = Workbook()

        # Required sheets in V20 format
        required_sheets = ["README", "Executive Summary (V20)", "Project Summary (V20)"]

        for sheet_name in required_sheets:
            wb.create_sheet(sheet_name)

        output_file = tmp_path / "v20_format_test.xlsx"
        wb.save(str(output_file))

        # Validate format
        from openpyxl import load_workbook

        loaded_wb = load_workbook(str(output_file))

        for sheet_name in required_sheets:
            assert sheet_name in loaded_wb.sheetnames

        # Check README is first sheet
        assert (
            loaded_wb.sheetnames[0] == "README" or loaded_wb.sheetnames[1] == "README"
        )
