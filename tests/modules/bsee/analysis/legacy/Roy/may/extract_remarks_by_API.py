import csv
import os
import pickle
import re
from collections import defaultdict
from datetime import datetime

import pandas as pd
from loguru import logger
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ExtractRemarksbyAPI:
    """
    Extract all API_WELL_NUMBER remarks from war data.
    Gets the mapping df from the war file, which contains API_WELL_NUMBER and SN_WAR.
    Reads the war remarks file, filters by SN_WAR, and maps to API_WELL_NUMBER.
    The output Excel file contains the API_WELL_NUMBER, REMARK_DATE, and REMARK_TEXT.
    """

    def router(self, cfg):
        self.extract_remarks_by_api(cfg)
        return cfg

    def extract_remarks_by_api(self, cfg):
        mapping_df = self.get_mapping_file(cfg)
        remark_file = cfg["data"]["groups"]["war_remarks"]

        try:
            mapping_df["SN_WAR"] = mapping_df["SN_WAR"].apply(
                lambda x: (
                    str(abs(int(str(x).strip())))
                    if str(x).strip().lstrip("-").isdigit()
                    else ""
                )
            )
            mapping_df = mapping_df[mapping_df["SN_WAR"] != ""]
        except Exception as e:
            logger.error(f"❌ Failed to read API-SN_WAR mapping: {e}")
            return

        snwar_to_api = mapping_df.set_index("SN_WAR")["API_WELL_NUMBER"].to_dict()
        remarks_by_api = defaultdict(list)

        def date_to_sortable_int(date_str):
            """Convert date string in various formats to sortable integer (YYYYMMDD)"""
            try:
                # Try to extract month, day, year with flexible regex
                match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", date_str)
                if match:
                    month, day, year = match.groups()
                    month = int(month)
                    day = int(day)
                    year = int(year)

                    # Handle 2-digit years (assuming 20xx for years < 50, 19xx otherwise)
                    if year < 50:
                        year += 2000
                    elif year < 100:
                        year += 1900

                    return year * 10000 + month * 100 + day
            except:
                pass
            return 0  # Default value for invalid dates

        try:
            # Load data from pickle file
            with open(remark_file, "rb") as f:
                remarks_df = pickle.load(f)  # This is a pandas DataFrame

            # Iterate through DataFrame rows properly
            for _, row in remarks_df.iterrows():
                raw_snwar = (
                    str(row["SN_WAR"]).strip() if "SN_WAR" in remarks_df.columns else ""
                )
                try:
                    snwar = str(abs(int(raw_snwar)))
                except:
                    continue

                if snwar in snwar_to_api:
                    api = snwar_to_api[snwar]
                    text = (
                        str(row["TEXT_REMARK"]).strip()
                        if "TEXT_REMARK" in remarks_df.columns
                        else ""
                    )
                    if not text:
                        continue

                    for line in text.splitlines():
                        match = re.match(
                            r"^[A-Za-z]{2} (\d{1,2}/\d{1,2}/\d{2,4}) - (.+)", line
                        )
                        if match:
                            date_str, remark = match.groups()
                            remarks_by_api[api].append((date_str, remark.strip()[:500]))

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "API Remarks"

            # Headers
            ws.append(["API_WELL_NUMBER", "REMARK_DATE", "REMARK_TEXT"])

            # Populate data by API
            for api, entries in remarks_by_api.items():
                # Sort by numerical date value
                sorted_entries = sorted(
                    entries, key=lambda x: date_to_sortable_int(x[0])
                )

                # Calculate duration in days (if we have at least 2 valid dates)
                duration = ""
                if len(sorted_entries) >= 2:
                    try:
                        date1 = sorted_entries[0][0]
                        date2 = sorted_entries[-1][0]
                        days = (
                            date_to_sortable_int(date2) - date_to_sortable_int(date1)
                        ) // 10000  # Approximate days
                        duration = f"{days} days" if days >= 0 else ""
                    except:
                        pass

                for date_str, remark in sorted_entries:
                    ws.append([api, date_str, remark])

                # # Insert duration row
                # ws.append(["duration", duration, ""])
                # # Insert 2 additional blank rows
                # ws.append(["", "", ""])

            # Adjust column widths
            ws.column_dimensions[get_column_letter(1)].width = 16
            ws.column_dimensions[get_column_letter(2)].width = 10
            ws.column_dimensions[get_column_letter(3)].width = 100

            result_folder = cfg["Analysis"]["result_folder"]
            filename = "api_remarks_output.xlsx"

            wb.save(os.path.join(result_folder, filename))
            logger.info(
                f"✅ Extracted remarks to {filename} with duration rows and spacing"
            )

        except Exception as e:
            logger.error(f"❌ Failed to process remark file: {e}")

    def get_mapping_file(self, cfg):

        war_file = cfg["data"]["groups"]["war"]
        war_df = pd.read_pickle(war_file)

        df = war_df[["API_WELL_NUMBER", "SN_WAR"]]

        df = df.dropna().drop_duplicates()
        df["API_WELL_NUMBER"] = df["API_WELL_NUMBER"].astype(str).str.strip()

        df["SN_WAR"] = df["SN_WAR"].apply(
            lambda x: str(abs(int(float(x)))) if pd.notna(x) else ""
        )

        return df
