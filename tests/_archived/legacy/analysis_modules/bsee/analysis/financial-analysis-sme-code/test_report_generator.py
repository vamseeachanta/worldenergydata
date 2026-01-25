"""
Tests for SME financial analysis report generator
"""

import pytest
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch, mock_open
import tempfile
import shutil

from src.worldenergydata.modules.bsee.analysis.financial.report_generator import (
    ReportGenerator,
    ExcelFormatter,
    format_number,
    format_currency,
    format_percentage
)


class TestExcelFormatter:
    """Test the ExcelFormatter class"""
    
    def test_initialization(self):
        """Test formatter initialization"""
        formatter = ExcelFormatter()
        assert formatter.styles is not None
        assert 'header' in formatter.styles
        assert 'title' in formatter.styles
        assert 'currency' in formatter.styles
        assert 'percentage' in formatter.styles
    
    def test_format_currency(self):
        """Test currency formatting"""
        assert format_currency(1000000) == "$1,000,000"
        assert format_currency(1234.56) == "$1,235"
        assert format_currency(0) == "$0"
        assert format_currency(-1000) == "-$1,000"
    
    def test_format_percentage(self):
        """Test percentage formatting"""
        assert format_percentage(0.1234) == "12.34%"
        assert format_percentage(1.0) == "100.00%"
        assert format_percentage(0) == "0.00%"
        assert format_percentage(-0.05) == "-5.00%"
    
    def test_format_number(self):
        """Test number formatting"""
        assert format_number(1000000) == "1,000,000"
        assert format_number(1234.56) == "1,235"
        assert format_number(0) == "0"
        assert format_number(-1000) == "-1,000"
    
    @patch('openpyxl.Workbook')
    def test_apply_header_style(self, mock_workbook):
        """Test applying header style to worksheet"""
        formatter = ExcelFormatter()
        mock_ws = MagicMock()
        mock_ws.__getitem__.return_value = [MagicMock() for _ in range(5)]
        
        formatter.apply_header_style(mock_ws)
        
        # Check that style was applied to header row
        assert mock_ws.__getitem__.called


class TestReportGenerator:
    """Test the ReportGenerator class"""
    
    @pytest.fixture
    def sample_cash_flow_data(self):
        """Create sample cash flow data"""
        dates = pd.date_range('2020-01-01', periods=12, freq='MS')
        return pd.DataFrame({
            'YearMonth': dates,
            'WELL_1': [1000] * 12,
            'WELL_2': [800] * 12,
            'Gross_Oil_bbls': [1800] * 12,
            'WTI_Price': [75.0] * 12,
            'Revenue_Gross': [135000] * 12,
            'Revenue_Net': [100000] * 12,
            'OPEX': [20000] * 12,
            'CAPEX_Drill': [0] * 3 + [500000] + [0] * 8,
            'CAPEX_Comp': [0] * 4 + [300000] + [0] * 7,
            'CAPEX_Facilities': [100000] * 3 + [0] * 9,
            'CAPEX': [100000] * 3 + [500000] + [300000] + [0] * 7,
            'Net_Cash_Flow': [-20000] * 3 + [-420000] + [-220000] + [80000] * 7,
            'Cum_Cash_Flow': pd.Series([-20000] * 3 + [-420000] + [-220000] + [80000] * 7).cumsum()
        }).set_index('YearMonth')
    
    @pytest.fixture
    def sample_metrics(self):
        """Create sample financial metrics"""
        return {
            'total_oil_bbls': 1000000,
            'total_revenue': 75000000,
            'total_opex': 12000000,
            'total_capex': 50000000,
            'npv': 8500000,
            'mirr_annual': 0.1523
        }
    
    @pytest.fixture
    def sample_development_data(self):
        """Create sample development data"""
        return {
            'Stones': {
                'cash_flow': pd.DataFrame(),  # Would contain actual cash flow
                'metrics': {
                    'total_oil_bbls': 5000000,
                    'npv': 45000000,
                    'mirr_annual': 0.18
                },
                'first_oil': pd.Timestamp('2020-04-01'),
                'development_type': 'subsea',
                'producer_count': 4
            },
            'Anchor': {
                'cash_flow': pd.DataFrame(),
                'metrics': {
                    'total_oil_bbls': 3000000,
                    'npv': 25000000,
                    'mirr_annual': 0.15
                },
                'first_oil': pd.Timestamp('2020-06-01'),
                'development_type': 'dry_tree',
                'producer_count': 3
            }
        }
    
    @pytest.fixture
    def report_generator(self):
        """Create a ReportGenerator instance"""
        return ReportGenerator()
    
    def test_initialization(self, report_generator):
        """Test report generator initialization"""
        assert report_generator.formatter is not None
        assert report_generator.workbook is None
    
    def test_create_readme_sheet(self, report_generator):
        """Test README sheet creation"""
        with patch('openpyxl.Workbook') as mock_wb:
            mock_workbook = MagicMock()
            mock_ws = MagicMock()
            mock_workbook.create_sheet.return_value = mock_ws
            report_generator.workbook = mock_workbook
            
            report_generator.create_readme_sheet(
                version='V20',
                generated_date=datetime.now(),
                input_files=['file1.xlsx', 'file2.xlsx']
            )
            
            # Check that sheet was created
            mock_workbook.create_sheet.assert_called_once()
            assert mock_ws.__setitem__.called
    
    def test_create_executive_summary(self, report_generator, sample_development_data):
        """Test executive summary sheet creation"""
        with patch('openpyxl.Workbook') as mock_wb:
            mock_workbook = MagicMock()
            mock_ws = MagicMock()
            mock_workbook.create_sheet.return_value = mock_ws
            report_generator.workbook = mock_workbook
            
            report_generator.create_executive_summary(sample_development_data)
            
            # Check that sheet was created
            mock_workbook.create_sheet.assert_called_with("Executive Summary (V20)")
    
    def test_create_project_summary(self, report_generator, sample_development_data):
        """Test project summary sheet creation"""
        with patch('openpyxl.Workbook') as mock_wb:
            mock_workbook = MagicMock()
            mock_ws = MagicMock()
            mock_workbook.create_sheet.return_value = mock_ws
            report_generator.workbook = mock_workbook
            
            report_generator.create_project_summary(sample_development_data)
            
            # Check that sheet was created
            mock_workbook.create_sheet.assert_called_with("Project Summary (V20)")
    
    def test_create_development_sheet(self, report_generator, sample_cash_flow_data):
        """Test individual development sheet creation"""
        with patch('openpyxl.Workbook') as mock_wb:
            mock_workbook = MagicMock()
            mock_ws = MagicMock()
            mock_workbook.create_sheet.return_value = mock_ws
            report_generator.workbook = mock_workbook
            
            report_generator.create_development_sheet('TestDev', sample_cash_flow_data)
            
            # Check that sheet was created with truncated name if needed
            mock_workbook.create_sheet.assert_called()
    
    def test_generate_report(self, report_generator, sample_development_data, tmp_path):
        """Test complete report generation"""
        output_file = tmp_path / "test_report.xlsx"
        
        with patch.object(report_generator, '_save_workbook') as mock_save:
            with patch.object(report_generator, '_apply_v20_formatting'):
                result = report_generator.generate_report(
                    development_data=sample_development_data,
                    output_path=str(output_file),
                    version='V20'
                )
                
                assert result == str(output_file)
                mock_save.assert_called_once()
    
    def test_format_executive_summary_data(self, report_generator, sample_development_data):
        """Test formatting of executive summary data"""
        exec_data = report_generator._format_executive_summary_data(sample_development_data)
        
        assert len(exec_data) == 2  # Two developments
        assert 'Project Name' in exec_data[0]
        assert 'TOTAL OIL BBL' in exec_data[0]
        assert 'NPV10 afterTax' in exec_data[0]
        assert 'MIRR afterTax' in exec_data[0]
    
    def test_format_project_summary_data(self, report_generator, sample_development_data):
        """Test formatting of project summary data"""
        proj_data = report_generator._format_project_summary_data(sample_development_data)
        
        assert len(proj_data) == 2  # Two developments
        assert 'Project Name' in proj_data[0]
        assert 'DEV SYSTEM USED' in proj_data[0]
        assert 'Producer Wells Used' in proj_data[0]
        assert 'FO Month' in proj_data[0]
    
    def test_apply_v20_formatting(self, report_generator, tmp_path):
        """Test V20 specific formatting"""
        # Create a valid Excel file
        test_file = tmp_path / "test.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        wb.create_sheet("README")
        wb.create_sheet("Executive Summary (V20)")
        wb.create_sheet("Project Summary (V20)")
        wb.save(str(test_file))
        
        with patch('src.worldenergydata.modules.bsee.analysis.financial.report_generator.load_workbook') as mock_load:
            mock_wb = MagicMock()
            mock_load.return_value = mock_wb
            
            # Create mock sheets
            mock_wb.sheetnames = ['README', 'Executive Summary (V20)', 'Project Summary (V20)']
            mock_sheets = {name: MagicMock() for name in mock_wb.sheetnames}
            mock_wb.__getitem__ = lambda self, x: mock_sheets[x]
            
            # Add mock max_column and max_row
            for sheet in mock_sheets.values():
                sheet.max_column = 10
                sheet.max_row = 5
                sheet.cell.return_value.value = "test"
            
            report_generator._apply_v20_formatting(str(test_file))
            
            mock_load.assert_called_once_with(str(test_file))
            mock_wb.save.assert_called_once_with(str(test_file))
    
    def test_sheet_name_truncation(self, report_generator):
        """Test that long sheet names are truncated to 31 characters"""
        long_name = "This is a very long development name that exceeds limit"
        truncated = report_generator._truncate_sheet_name(long_name)
        
        assert len(truncated) <= 31
        assert truncated == long_name[:31]
    
    def test_number_formatting_in_sheets(self, report_generator):
        """Test that numbers are formatted correctly in sheets"""
        with patch('openpyxl.Workbook') as mock_wb:
            mock_workbook = MagicMock()
            mock_ws = MagicMock()
            mock_workbook.create_sheet.return_value = mock_ws
            report_generator.workbook = mock_workbook
            
            # Test data with various number types
            test_data = pd.DataFrame({
                'Currency': [1000000.50],
                'Percentage': [0.1523],
                'Integer': [1000],
                'Date': [pd.Timestamp('2020-01-01')]
            })
            
            report_generator._write_dataframe_to_sheet(mock_ws, test_data)
            
            # Verify cells were written
            assert mock_ws.cell.called


class TestIntegration:
    """Integration tests for the report generator"""
    
    def test_full_report_generation_workflow(self, tmp_path):
        """Test complete workflow from cash flow to Excel report"""
        # Create test data
        dates = pd.date_range('2020-01-01', periods=24, freq='MS')
        cash_flow_df = pd.DataFrame({
            'WELL_1': [1000] * 24,
            'WELL_2': [800] * 24,
            'Gross_Oil_bbls': [1800] * 24,
            'Revenue_Net': [100000] * 24,
            'OPEX': [20000] * 24,
            'CAPEX': [500000] + [0] * 23,
            'Net_Cash_Flow': [-420000] + [80000] * 23
        }, index=dates)
        cash_flow_df.index.name = 'YearMonth'
        
        development_data = {
            'TestDev': {
                'cash_flow': cash_flow_df,
                'metrics': {
                    'total_oil_bbls': cash_flow_df['Gross_Oil_bbls'].sum(),
                    'total_revenue': cash_flow_df['Revenue_Net'].sum(),
                    'total_opex': cash_flow_df['OPEX'].sum(),
                    'total_capex': cash_flow_df['CAPEX'].sum(),
                    'npv': 1000000,
                    'mirr_annual': 0.15
                },
                'first_oil': pd.Timestamp('2020-01-01'),
                'development_type': 'subsea',
                'producer_count': 2
            }
        }
        
        # Generate report
        generator = ReportGenerator()
        output_file = tmp_path / "integration_test.xlsx"
        
        with patch.object(generator, '_save_workbook'):
            with patch.object(generator, '_apply_v20_formatting'):
                with patch('openpyxl.Workbook'):
                    result = generator.generate_report(
                        development_data=development_data,
                        output_path=str(output_file),
                        version='V20'
                    )
                    
                    assert result == str(output_file)
    
    def test_multi_development_report(self, tmp_path):
        """Test report generation with multiple developments"""
        developments = {}
        
        # Create multiple development data
        for i, dev_name in enumerate(['Dev1', 'Dev2', 'Dev3']):
            dates = pd.date_range('2020-01-01', periods=12, freq='MS')
            cash_flow = pd.DataFrame({
                'Gross_Oil_bbls': [1000 * (i + 1)] * 12,
                'Revenue_Net': [50000 * (i + 1)] * 12,
                'OPEX': [10000] * 12,
                'CAPEX': [100000] + [0] * 11,
                'Net_Cash_Flow': [-60000 + 40000 * i] + [40000 * (i + 1)] * 11
            }, index=dates)
            
            developments[dev_name] = {
                'cash_flow': cash_flow,
                'metrics': {
                    'total_oil_bbls': cash_flow['Gross_Oil_bbls'].sum(),
                    'npv': 100000 * (i + 1),
                    'mirr_annual': 0.10 + 0.02 * i
                },
                'development_type': 'subsea' if i % 2 == 0 else 'dry_tree'
            }
        
        generator = ReportGenerator()
        output_file = tmp_path / "multi_dev_test.xlsx"
        
        with patch.object(generator, '_save_workbook'):
            with patch.object(generator, '_apply_v20_formatting'):
                with patch('openpyxl.Workbook'):
                    result = generator.generate_report(
                        development_data=developments,
                        output_path=str(output_file)
                    )
                    
                    assert result == str(output_file)