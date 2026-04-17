"""Tests for Excel exporter validate_data, get_supported_features, _init_styles."""

import pytest

from worldenergydata.bsee.reports.comprehensive.exporters.excel_exporter import (
    ExcelExporter,
)

# ---------------------------------------------------------------------------
# _init_styles
# ---------------------------------------------------------------------------


class TestInitStyles:
    def test_styles_created(self):
        exporter = ExcelExporter()
        assert "header" in exporter.styles
        assert "title" in exporter.styles
        assert "currency" in exporter.styles
        assert "percentage" in exporter.styles
        assert "number" in exporter.styles
        assert "date" in exporter.styles

    def test_header_style_bold(self):
        exporter = ExcelExporter()
        style = exporter.styles["header"]
        assert style.font.bold is True

    def test_currency_format(self):
        exporter = ExcelExporter()
        style = exporter.styles["currency"]
        assert "$" in style.number_format


# ---------------------------------------------------------------------------
# validate_data
# ---------------------------------------------------------------------------


class TestValidateData:
    def test_empty_data(self):
        exporter = ExcelExporter()
        assert exporter.validate_data({}) is False

    def test_none_data(self):
        exporter = ExcelExporter()
        assert exporter.validate_data(None) is False

    def test_with_summary(self):
        exporter = ExcelExporter()
        assert exporter.validate_data({"summary": {}}) is True

    def test_with_production_data(self):
        exporter = ExcelExporter()
        assert exporter.validate_data({"production_data": []}) is True

    def test_with_financial_data(self):
        exporter = ExcelExporter()
        assert exporter.validate_data({"financial_data": []}) is True

    def test_with_kpis(self):
        exporter = ExcelExporter()
        assert exporter.validate_data({"kpis": {}}) is True

    def test_with_charts(self):
        exporter = ExcelExporter()
        assert exporter.validate_data({"charts": {}}) is True

    def test_with_raw_data(self):
        exporter = ExcelExporter()
        assert exporter.validate_data({"raw_data": []}) is True

    def test_with_report_metadata(self):
        exporter = ExcelExporter()
        assert exporter.validate_data({"report_metadata": {}}) is True

    def test_unrecognized_section(self):
        exporter = ExcelExporter()
        assert exporter.validate_data({"other_key": "value"}) is False


# ---------------------------------------------------------------------------
# get_supported_features
# ---------------------------------------------------------------------------


class TestGetSupportedFeatures:
    def test_returns_list(self):
        exporter = ExcelExporter()
        features = exporter.get_supported_features()
        assert isinstance(features, list)

    def test_contains_key_features(self):
        exporter = ExcelExporter()
        features = exporter.get_supported_features()
        assert "multiple_sheets" in features
        assert "formatting" in features
        assert "charts" in features
        assert "styles" in features

    def test_feature_count(self):
        exporter = ExcelExporter()
        assert len(exporter.get_supported_features()) == 10


# ---------------------------------------------------------------------------
# create_sheets
# ---------------------------------------------------------------------------


class TestCreateSheets:
    def test_summary_sheet(self):
        exporter = ExcelExporter()
        from openpyxl import Workbook

        exporter.workbook = Workbook()
        exporter.workbook.remove(exporter.workbook.active)
        sheets = exporter.create_sheets({"summary": {"total": 100}})
        assert "Summary" in sheets

    def test_production_sheet(self):
        exporter = ExcelExporter()
        from openpyxl import Workbook

        exporter.workbook = Workbook()
        exporter.workbook.remove(exporter.workbook.active)
        sheets = exporter.create_sheets({"production_data": [{"oil": 1000}]})
        assert "Production Data" in sheets

    def test_empty_data(self):
        exporter = ExcelExporter()
        from openpyxl import Workbook

        exporter.workbook = Workbook()
        exporter.workbook.remove(exporter.workbook.active)
        sheets = exporter.create_sheets({})
        assert sheets == []


# ---------------------------------------------------------------------------
# supported_format
# ---------------------------------------------------------------------------


class TestSupportedFormat:
    def test_is_excel(self):
        exporter = ExcelExporter()
        assert exporter.supported_format.value == "excel"
