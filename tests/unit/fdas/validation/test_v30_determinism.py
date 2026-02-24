"""
Phase A: V30 Determinism Tests

Verify that Roy's FDAS V30 golden baseline output (financial_project_summary.xlsx)
matches the captured reference values in golden_baseline_v30.yml.

This confirms the baseline is stable and correctly captured before
we attempt to reproduce it from our pipeline (Phase B).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
import yaml

PROJECT_ROOT = Path(__file__).resolve().parents[4]
FDAS_V30_DIR = PROJECT_ROOT / "docs/modules/bsee/analysis/production/FDAS_V30"
GOLDEN_BASELINE_PATH = (
    PROJECT_ROOT / "config/analysis/lower_tertiary/golden_baseline_v30.yml"
)


@pytest.fixture(scope="module")
def golden_baseline():
    """Load golden baseline reference values."""
    with GOLDEN_BASELINE_PATH.open("r") as f:
        return yaml.safe_load(f)


@pytest.fixture(scope="module")
def v30_output():
    """Load Roy's V30 financial_project_summary.xlsx Project_Summary sheet."""
    xlsx_path = FDAS_V30_DIR / "financial_project_summary.xlsx"
    if not xlsx_path.exists():
        pytest.skip(f"V30 output not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Project_Summary"]

    # Extract headers from row 1
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)

    # Extract all project rows
    projects = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        # Find the project name column (first non-None string column)
        project_name = None
        for val in row:
            if isinstance(val, str) and val.strip():
                project_name = val.strip()
                break
        if project_name:
            projects[project_name] = record

    return projects


@pytest.fixture(scope="module")
def v30_monthly_sheets():
    """Load per-project monthly cashflow sheets from V30 output."""
    xlsx_path = FDAS_V30_DIR / "financial_project_summary.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    sheets = {}
    for name in wb.sheetnames:
        if name == "Project_Summary":
            continue
        ws = wb[name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        sheets[name] = rows

    return sheets


class TestV30BaselineCaptured:
    """Verify golden baseline YAML was correctly populated."""

    def test_golden_baseline_has_all_producing_projects(self, golden_baseline):
        """All 7 producing projects present in baseline."""
        producing = [
            "jack_st_malo",
            "stones",
            "julia",
            "big_foot",
            "cascade_chinook",
            "anchor",
            "shenandoah",
        ]
        for proj_id in producing:
            assert proj_id in golden_baseline["projects"], f"Missing: {proj_id}"

    def test_golden_baseline_has_exploration_projects(self, golden_baseline):
        """All 3 exploration projects present."""
        exploration = ["north_platte", "kaskida", "tiber"]
        for proj_id in exploration:
            assert proj_id in golden_baseline["projects"], f"Missing: {proj_id}"

    def test_golden_baseline_has_tolerances(self, golden_baseline):
        """Tolerances defined for key metrics."""
        assert "tolerances" in golden_baseline
        assert "npv_relative" in golden_baseline["tolerances"]
        assert "mirr_absolute" in golden_baseline["tolerances"]

    def test_golden_baseline_metadata(self, golden_baseline):
        """Metadata present and valid."""
        assert "metadata" in golden_baseline
        assert "source_file" in golden_baseline["metadata"]


class TestV30OutputMatchesBaseline:
    """Compare V30 Excel output against captured golden baseline."""

    # Map from golden baseline project IDs to V30 sheet/summary names
    PROJECT_NAME_MAP = {
        "jack_st_malo": "Jack St Malo",
        "stones": "Stones",
        "julia": "Julia",
        "big_foot": "Big Foot",
        "cascade_chinook": "Cascade Chinook",
        "anchor": "Anchor",
        "shenandoah": "Shenandoah",
    }

    def _get_monthly_production(self, sheets, project_name):
        """Sum monthly oil production from per-project sheet."""
        if project_name not in sheets:
            return None
        rows = sheets[project_name]
        total = 0
        for row in rows:
            vol = row.get("MONTHLY_OIL_VOLUME") or row.get("Monthly_Oil_Volume") or 0
            if vol and not isinstance(vol, str):
                total += float(vol)
        return total

    def _get_monthly_revenue(self, sheets, project_name):
        """Sum monthly revenue from per-project sheet."""
        if project_name not in sheets:
            return None
        rows = sheets[project_name]
        total = 0
        for row in rows:
            rev = row.get("Revenue_USD") or row.get("REVENUE_USD") or 0
            if rev and not isinstance(rev, str):
                total += float(rev)
        return total

    @pytest.mark.parametrize("project_id,project_name", list(PROJECT_NAME_MAP.items()))
    def test_production_matches(
        self, golden_baseline, v30_monthly_sheets, project_id, project_name
    ):
        """Monthly sheet production totals match golden baseline."""
        baseline = golden_baseline["projects"][project_id]
        expected_oil = baseline["total_oil_bbl"]

        actual_oil = self._get_monthly_production(v30_monthly_sheets, project_name)
        if actual_oil is None:
            pytest.skip(f"Sheet '{project_name}' not found in V30 output")

        tol = golden_baseline["tolerances"]["production_relative"]
        assert actual_oil == pytest.approx(
            expected_oil, rel=tol
        ), f"{project_name}: oil {actual_oil:,.0f} vs baseline {expected_oil:,.0f}"

    @pytest.mark.parametrize("project_id,project_name", list(PROJECT_NAME_MAP.items()))
    def test_revenue_matches(
        self, golden_baseline, v30_monthly_sheets, project_id, project_name
    ):
        """Monthly sheet revenue totals match golden baseline."""
        baseline = golden_baseline["projects"][project_id]
        expected_rev = baseline["revenue_usd"]

        actual_rev = self._get_monthly_revenue(v30_monthly_sheets, project_name)
        if actual_rev is None:
            pytest.skip(f"Sheet '{project_name}' not found in V30 output")

        tol = golden_baseline["tolerances"]["revenue_relative"]
        assert actual_rev == pytest.approx(
            expected_rev, rel=tol
        ), f"{project_name}: revenue {actual_rev:,.0f} vs baseline {expected_rev:,.0f}"


class TestV30MonthlyDataIntegrity:
    """Verify V30 monthly data quality."""

    def test_all_project_sheets_present(self, v30_monthly_sheets):
        """All 10 per-project sheets exist."""
        expected = {
            "Stones",
            "Cascade Chinook",
            "Julia",
            "Anchor",
            "Jack St Malo",
            "Kaskida",
            "Tiber",
            "Shenandoah",
            "North Platte",
            "Big Foot",
        }
        actual = set(v30_monthly_sheets.keys())
        assert expected == actual, f"Missing sheets: {expected - actual}"

    def test_monthly_sheets_have_data(self, v30_monthly_sheets):
        """Each sheet has rows of data."""
        for name, rows in v30_monthly_sheets.items():
            assert len(rows) > 0, f"Sheet '{name}' is empty"

    def test_wti_prices_consistent(self, v30_monthly_sheets):
        """WTI price series is the same across all project sheets."""
        price_series = {}
        for name, rows in v30_monthly_sheets.items():
            prices = {}
            for row in rows:
                month = row.get("MONTH")
                wti = row.get("WTI_USD") or row.get("WTI_$")
                if month and wti and not isinstance(wti, str):
                    prices[str(month)] = float(wti)
            if prices:
                price_series[name] = prices

        if len(price_series) < 2:
            pytest.skip("Not enough sheets with WTI data to compare")

        # Compare all sheets against the first
        reference_name, reference_prices = next(iter(price_series.items()))
        for name, prices in price_series.items():
            if name == reference_name:
                continue
            common_months = set(reference_prices.keys()) & set(prices.keys())
            for month in common_months:
                assert prices[month] == pytest.approx(
                    reference_prices[month], rel=1e-6
                ), (
                    f"WTI mismatch at {month}: "
                    f"{name}={prices[month]} vs "
                    f"{reference_name}={reference_prices[month]}"
                )
