"""Curated Big Foot requirement and award-map contract tests (issue #1039)."""

from __future__ import annotations

import csv
import re
from pathlib import Path

import pytest
from pydantic import ValidationError

ROOT = Path(__file__).resolve().parents[3]
CURATED = ROOT / "data/modules/cost/curated"
MONEY_NARRATIVE_PATTERN = re.compile(
    r"(?:\$\s*\d|\bUSD\s*[\d$]|\b\d+(?:\.\d+)?\s*(?:MM|million|bn|billion)\b)",
    re.IGNORECASE,
)


def _rows(name: str) -> list[dict[str, str]]:
    with (CURATED / name).open(encoding="utf-8", newline="") as stream:
        return list(csv.DictReader(stream))


def _resolve_locator(source_table: str, locator: str) -> dict[str, str]:
    predicates = dict(part.split("=", 1) for part in locator.split("|"))
    matches = [
        row
        for row in _rows(source_table)
        if all(row.get(column) == value for column, value in predicates.items())
    ]
    assert len(matches) == 1, (source_table, locator, len(matches))
    return matches[0]


def _resolve_embedded_locator(locator: str) -> tuple[str, dict[str, str]]:
    source_table, separator, predicates = locator.partition(":")
    assert separator == ":"
    return source_table, _resolve_locator(source_table, predicates)


def test_big_foot_requirements_cover_dry_tree_tlp_architecture() -> None:
    from worldenergydata.cost.timeseries.cost_map_schema import (
        Evidence,
        RequiredAsset,
        WorkPackageRequirement,
    )

    requirements = _rows("project_asset_requirements.csv")
    assert {row["WORK_PACKAGE"] for row in requirements} == {
        "host/TLP",
        "dry trees",
        "wells",
        "drilling/completion",
        "marine riser/tensioner",
        "export",
        "installation/hookup",
        "controls",
    }
    for row in requirements:
        Evidence(
            derivation=row["EVIDENCE_DERIVATION"],
            source_provenance=row["SOURCE_PROVENANCE"],
            source_url=row["SOURCE_URL"],
            source_locator=row["SOURCE_LOCATOR"],
            confidence=row["CONFIDENCE"],
        )
        requirement = WorkPackageRequirement(
            requirement_id=row["REQUIREMENT_ID"],
            project_id=row["PROJECT_ID"],
            work_package=row["WORK_PACKAGE"],
            required_assets=(
                RequiredAsset(asset_type=row["ASSET_TYPE"], quantity=row["QUANTITY"]),
            ),
        )
        assert requirement.project_id == "prj-000001"
        assert row["QUANTITY_UNIT"]
        assert row["EVIDENCE_NOTE"]


def test_unknown_quantity_is_valid_but_unproven_number_is_not() -> None:
    from worldenergydata.cost.timeseries.cost_map_schema import RequiredAsset

    rows = _rows("project_asset_requirements.csv")
    by_package = {row["WORK_PACKAGE"]: row for row in rows}
    assert {row["QUANTITY"] for row in rows} == {"unknown"}
    assert by_package["wells"]["QUANTITY"] == "unknown"
    assert by_package["dry trees"]["QUANTITY"] == "unknown"
    assert RequiredAsset(
        asset_type="development well", quantity="unknown"
    ).quantity == ("unknown")
    with pytest.raises(ValidationError):
        RequiredAsset(asset_type="development well", quantity="38")


def test_identity_registries_use_stable_monotonic_opaque_ids() -> None:
    from worldenergydata.cost.timeseries.cost_map_schema import IdentityRegistryEntry

    expected = {
        "cost_project_identity.csv": ("project", "prj-"),
        "cost_award_identity.csv": ("award", "awd-"),
        "cost_requirement_identity.csv": ("requirement", "req-"),
        "cost_event_identity.csv": ("event", "evt-"),
    }
    seen: set[str] = set()
    for name, (kind, prefix) in expected.items():
        rows = _rows(name)
        ids = [row["OPAQUE_ID"] for row in rows]
        assert ids == [f"{prefix}{number:06d}" for number in range(1, len(ids) + 1)]
        for row in rows:
            assert re.fullmatch(rf"{prefix}\d{{6}}", row["OPAQUE_ID"])
            assert row["OPAQUE_ID"] not in seen
            seen.add(row["OPAQUE_ID"])
            IdentityRegistryEntry(
                opaque_id=row["OPAQUE_ID"],
                entity_kind=row["ENTITY_KIND"],
                display_label=row["DISPLAY_LABEL"],
                state=row["STATE"],
                active=row["ACTIVE"] == "true",
                aliases=tuple(filter(None, row["ALIASES"].split("|"))),
                validation_group_id=row["VALIDATION_GROUP_ID"],
            )
            assert row["ENTITY_KIND"] == kind
            assert row["STATE"] == "active"
            assert row["ACTIVE"] == "true"
            assert row["TOMBSTONED_BY_ID"] == ""
            assert row["MIGRATION_NOTE"] == ""


def test_identity_registry_foreign_keys_resolve() -> None:
    projects = {row["OPAQUE_ID"] for row in _rows("cost_project_identity.csv")}
    awards = {row["OPAQUE_ID"] for row in _rows("cost_award_identity.csv")}
    requirements = {row["OPAQUE_ID"] for row in _rows("cost_requirement_identity.csv")}
    requirement_rows = _rows("project_asset_requirements.csv")
    assert {row["REQUIREMENT_ID"] for row in requirement_rows} == requirements
    assert {row["PROJECT_ID"] for row in requirement_rows} <= projects
    for name in (
        "cost_project_identity.csv",
        "cost_award_identity.csv",
        "cost_requirement_identity.csv",
        "cost_event_identity.csv",
    ):
        for row in _rows(name):
            assert row["SOURCE_TABLE"]
            assert row["SOURCE_RECORD_LOCATOR"]
            _resolve_locator(row["SOURCE_TABLE"], row["SOURCE_RECORD_LOCATOR"])
    for row in _rows("award_asset_links.csv"):
        assert row["AWARD_ID"] in awards
        assert row["COMMERCIAL_AMOUNT_ID"] in awards
        assert row["PROJECT_ID"] in projects
        assert set(row["REQUIREMENT_IDS"].split("|")) <= requirements


def test_every_big_foot_award_has_one_resolution_status() -> None:
    from worldenergydata.cost.timeseries.cost_map_schema import (
        AwardRequirementLink,
        CostMapStatus,
        Evidence,
    )

    links = _rows("award_asset_links.csv")
    award_ids = [row["OPAQUE_ID"] for row in _rows("cost_award_identity.csv")]
    assert [row["AWARD_ID"] for row in links] == award_ids
    assert sorted(row["AWARD_ID"] for row in links) == sorted(award_ids)
    assert len({row["AWARD_ID"] for row in links}) == len(links)
    assert not ({"VALUE_LOW_MM", "VALUE_HIGH_MM", "AMOUNT", "CURRENCY"} & set(links[0]))
    for row in links:
        AwardRequirementLink(
            award_id=row["AWARD_ID"],
            requirement_ids=tuple(row["REQUIREMENT_IDS"].split("|")),
            commercial_amount_id=row["COMMERCIAL_AMOUNT_ID"],
            bundle_group_id=row["BUNDLE_GROUP_ID"] or None,
        )
        CostMapStatus(
            link_resolution=row["LINK_RESOLUTION"],
            scope_coverage=row["SCOPE_COVERAGE"],
            bundle_group_id=row["BUNDLE_GROUP_ID"] or None,
            counting_disposition=row["COUNTING_DISPOSITION"],
            counting_reason=row["COUNTING_REASON"] or None,
        )
        Evidence(
            derivation=row["EVIDENCE_DERIVATION"],
            source_provenance=row["SOURCE_PROVENANCE"],
            source_url=row["SOURCE_URL"],
            source_locator=row["SOURCE_LOCATOR"],
            confidence=row["CONFIDENCE"],
        )


def test_award_link_narratives_are_money_free() -> None:
    for value in ("$45", "USD 45", "45MM", "45 million", "4bn"):
        assert MONEY_NARRATIVE_PATTERN.search(value)
    assert MONEY_NARRATIVE_PATTERN.search("component floor") is None
    for row in _rows("award_asset_links.csv"):
        assert MONEY_NARRATIVE_PATTERN.search(row["NOTES"]) is None


def test_requirement_and_link_evidence_matches_live_source_rows() -> None:
    provenance_columns = {
        "sanctioned_projects.csv": "SOURCE_PRIORITY",
        "contract_awards.csv": "PROVENANCE",
    }
    for fixture, narrative_field in (
        ("project_asset_requirements.csv", "EVIDENCE_NOTE"),
        ("award_asset_links.csv", "NOTES"),
    ):
        for row in _rows(fixture):
            source_table, source = _resolve_embedded_locator(row["SOURCE_LOCATOR"])
            assert row["SOURCE_URL"] == source["SOURCE_URL"]
            assert row["SOURCE_PROVENANCE"] == source[provenance_columns[source_table]]
            assert row[narrative_field]

    live_awards = [
        row for row in _rows("contract_awards.csv") if row["PROJECT"] == "Big Foot"
    ]
    identity_sources = {
        row["OPAQUE_ID"]: _resolve_locator(
            row["SOURCE_TABLE"], row["SOURCE_RECORD_LOCATOR"]
        )
        for row in _rows("cost_award_identity.csv")
    }
    identity_keys = {
        award_id: (row["AWARD_YEAR"], row["CONTRACTOR"])
        for award_id, row in identity_sources.items()
    }
    assert identity_keys == {
        "awd-000001": ("2011", "GE Oil & Gas"),
        "awd-000002": ("2009", "Enbridge"),
    }
    assert set(identity_keys.values()) == {
        (row["AWARD_YEAR"], row["CONTRACTOR"]) for row in live_awards
    }
    assert len(identity_keys) == len(live_awards) == 2
    for link in _rows("award_asset_links.csv"):
        _, link_source = _resolve_embedded_locator(link["SOURCE_LOCATOR"])
        identity_source = identity_sources[link["AWARD_ID"]]
        assert (link_source["AWARD_YEAR"], link_source["CONTRACTOR"]) == (
            identity_source["AWARD_YEAR"],
            identity_source["CONTRACTOR"],
        )


def test_big_foot_curated_registry_has_exactly_two_awards_and_zero_not_public() -> None:
    source_rows = [
        row for row in _rows("contract_awards.csv") if row["PROJECT"] == "Big Foot"
    ]
    assert [
        (
            row["AWARD_YEAR"],
            row["CONTRACTOR"],
            row["VALUE_LOW_MM"],
            row["VALUE_HIGH_MM"],
            row["VALUE_BASIS"],
        )
        for row in source_rows
    ] == [
        ("2011", "GE Oil & Gas", "45", "45", "point"),
        ("2009", "Enbridge", "200", "200", "midstream"),
    ]
    assert all(row["VALUE_BASIS"] != "not_public" for row in source_rows)
    registry = _rows("cost_award_identity.csv")
    assert len(registry) == len(source_rows) == 2


def test_thirty_eight_wellbores_does_not_fill_blank_sanctioned_well_count() -> None:
    from openpyxl import load_workbook

    workbook = ROOT / (
        "docs/modules/bsee/analysis/production/FDAS_V30/"
        "drilling_and_completion_days.xlsx"
    )
    sheet = load_workbook(workbook, read_only=True, data_only=True)["Sheet1"]
    rows = list(sheet.iter_rows(values_only=True))
    header = rows[0]
    records = [dict(zip(header, row)) for row in rows[1:]]
    big_foot_wellbores = [
        row
        for row in records
        if row["LEASE_NAME"] == "Big Foot" and row["API_WELL_NUMBER"]
    ]
    assert len(big_foot_wellbores) == 38
    sanctioned = next(
        row for row in _rows("sanctioned_projects.csv") if row["PROJECT"] == "Big Foot"
    )
    assert sanctioned["WELL_COUNT"] == ""
    wells = next(
        row
        for row in _rows("project_asset_requirements.csv")
        if row["WORK_PACKAGE"] == "wells"
    )
    assert wells["QUANTITY"] == "unknown"


def test_midstream_export_is_excluded_from_project_capex() -> None:
    source = next(
        row
        for row in _rows("contract_awards.csv")
        if row["PROJECT"] == "Big Foot" and row["CONTRACTOR"] == "Enbridge"
    )
    assert (
        source["AWARD_YEAR"],
        source["VALUE_LOW_MM"],
        source["VALUE_HIGH_MM"],
        source["VALUE_BASIS"],
    ) == ("2009", "200", "200", "midstream")
    link = next(
        row for row in _rows("award_asset_links.csv") if row["AWARD_ID"] == "awd-000002"
    )
    assert link["REQUIREMENT_IDS"] == "req-000006"
    assert link["COUNTING_DISPOSITION"] == "excluded"
    assert link["COUNTING_REASON"] == "non_capex"


def test_riser_tensioner_point_award_is_a_component_floor() -> None:
    source = next(
        row
        for row in _rows("contract_awards.csv")
        if row["PROJECT"] == "Big Foot" and row["CONTRACTOR"] == "GE Oil & Gas"
    )
    assert (
        source["AWARD_YEAR"],
        source["VALUE_LOW_MM"],
        source["VALUE_HIGH_MM"],
        source["VALUE_BASIS"],
    ) == ("2011", "45", "45", "point")
    link = next(
        row for row in _rows("award_asset_links.csv") if row["AWARD_ID"] == "awd-000001"
    )
    assert link["REQUIREMENT_IDS"] == "req-000005"
    assert link["COUNTING_DISPOSITION"] == "included"
    assert link["COUNTING_REASON"] == ""
    assert "component floor" in link["NOTES"]


def test_missing_host_award_remains_an_explicit_coverage_gap() -> None:
    requirements = _rows("project_asset_requirements.csv")
    host = next(row for row in requirements if row["WORK_PACKAGE"] == "host/TLP")
    assert host["REQUIREMENT_ID"] == "req-000001"
    assert host["QUANTITY"] == "unknown"
    assert "no public host award" in host["EVIDENCE_NOTE"].lower()
    assert "coverage gap" in host["EVIDENCE_NOTE"].lower()
    linked_requirement_ids = {
        requirement_id
        for row in _rows("award_asset_links.csv")
        for requirement_id in row["REQUIREMENT_IDS"].split("|")
    }
    assert host["REQUIREMENT_ID"] not in linked_requirement_ids
