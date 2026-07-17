"""Read-only FDAS assumption crosswalk contract for Big Foot."""

from __future__ import annotations

import csv
import hashlib
from decimal import Decimal
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
CURATED = ROOT / "data/modules/cost/curated"
FDAS = ROOT / "docs/modules/bsee/analysis/production/FDAS_V30"
CROSSWALK = CURATED / "fdas_project_cost_crosswalk.csv"
ABSOLUTE_PATH_PREFIXES = ("/home/", "/mnt/")  # abs-path-allowed
EXPECTED_COLUMNS = (
    "PROJECT_ID",
    "WORKBOOK_FILE",
    "WORKBOOK_SHA256",
    "WORKBOOK_SHEET",
    "WORKBOOK_CELL",
    "WORKBOOK_ROLE",
    "WORKBOOK_CATEGORY",
    "WORKBOOK_VALUE",
    "VALUE_UNIT",
    "CURRENCY",
    "PRICE_BASIS",
    "BASIS_YEAR",
    "OWNERSHIP_BASIS",
    "SCOPE_BASIS",
    "CAPEX_BASIS",
    "REQUIREMENT_IDS",
    "MAPPING_STATUS",
    "COUNTING_DISPOSITION",
    "EVIDENCE_DERIVATION",
    "SOURCE_PROVENANCE",
    "ASSUMPTION_VINTAGE",
    "COMPARISON_ELIGIBILITY",
    "NOTES",
)
FROZEN_HASHES = {
    "lease_assumptions.xlsx": (
        "a1193f669db49ac33b87481733fb13af409844fed890e763b4e8726e329a1407"
    ),
    "financial_project_summary.xlsx": (
        "00f200def283d307293bb93033f070718722618b9a8ace2bbbe11bfbffeddf04"
    ),
    "drilling_and_completion_days.xlsx": (
        "3ecfa1128b33edf73db3a793f8839c98c50bc27184487a8af579c5ef22795e7f"
    ),
}
ALLOWLISTED_COST_CELLS = {
    "financial_project_summary.xlsx": {
        "E3",
        "F3",
        "G3",
        "H3",
        "I3",
        "J3",
        "K3",
        "L3",
        "M3",
        "P3",
        "Q3",
    },
    "lease_assumptions.xlsx": {
        "D9",
        "D10",
        "D12",
        "D13",
        "D15",
        "D16",
        "D17",
        "D36",
        "D37",
        "D38",
    },
}
EXPECTED_UNITS_AND_CURRENCIES = {
    **{
        ("financial_project_summary.xlsx", cell): ("USD", "USD")
        for cell in ALLOWLISTED_COST_CELLS["financial_project_summary.xlsx"]
    },
    ("lease_assumptions.xlsx", "D9"): ("USD MM/day", "USD"),
    ("lease_assumptions.xlsx", "D10"): ("USD MM", "USD"),
    ("lease_assumptions.xlsx", "D12"): ("USD MM/well", "USD"),
    ("lease_assumptions.xlsx", "D13"): ("USD MM/producer", "USD"),
    ("lease_assumptions.xlsx", "D15"): ("USD MM/facility", "USD"),
    ("lease_assumptions.xlsx", "D16"): ("USD/bbl", "USD"),
    ("lease_assumptions.xlsx", "D17"): ("USD MM/year", "USD"),
    ("lease_assumptions.xlsx", "D36"): ("USD MM/pump", "USD"),
    ("lease_assumptions.xlsx", "D37"): ("USD MM/pump", "USD"),
    ("lease_assumptions.xlsx", "D38"): ("USD MM/pump", "USD"),
    ("drilling_and_completion_days.xlsx", "Sheet1!A17:L54"): (
        "wellbore rows; drilling days; completion days",
        "",
    ),
}


def _rows() -> list[dict[str, str]]:
    with CROSSWALK.open(encoding="utf-8", newline="") as stream:
        return list(csv.DictReader(stream))


def _digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_frozen_big_foot_workbook_fingerprints() -> None:
    rows = _rows()

    assert {row["WORKBOOK_FILE"] for row in rows} == set(FROZEN_HASHES)
    assert all(
        row["WORKBOOK_SHA256"] == FROZEN_HASHES[row["WORKBOOK_FILE"]] for row in rows
    )
    assert {name: _digest(FDAS / name) for name in FROZEN_HASHES} == FROZEN_HASHES


def test_big_foot_crosswalk_covers_every_allowlisted_cost_cell() -> None:
    from openpyxl import load_workbook

    with (CURATED / "cost_requirement_identity.csv").open(
        encoding="utf-8", newline=""
    ) as stream:
        known_requirements = {row["OPAQUE_ID"] for row in csv.DictReader(stream)}
    observed = {
        name: {
            row["WORKBOOK_CELL"]
            for row in _rows()
            if row["WORKBOOK_FILE"] == name and row["WORKBOOK_CELL"]
        }
        for name in ALLOWLISTED_COST_CELLS
    }

    assert observed == ALLOWLISTED_COST_CELLS
    for row in _rows():
        assert row["PROJECT_ID"] == "prj-000001"
        assert (
            set(filter(None, row["REQUIREMENT_IDS"].split("|"))) <= known_requirements
        )
        if row["WORKBOOK_ROLE"] not in {"coverage_gap", "fingerprint_reference"}:
            assert row["WORKBOOK_CELL"] in ALLOWLISTED_COST_CELLS[row["WORKBOOK_FILE"]]
            assert row["REQUIREMENT_IDS"] or row["MAPPING_STATUS"] in {
                "unmapped",
                "excluded",
            }
    for name, cells in ALLOWLISTED_COST_CELLS.items():
        workbook = load_workbook(FDAS / name, read_only=True, data_only=True)
        sheet_name = (
            "Project_Summary" if name.startswith("financial") else "assumptions"
        )
        sheet = workbook[sheet_name]
        crosswalk_values = {
            row["WORKBOOK_CELL"]: Decimal(row["WORKBOOK_VALUE"])
            for row in _rows()
            if row["WORKBOOK_FILE"] == name and row["WORKBOOK_CELL"] in cells
        }
        assert crosswalk_values == {
            cell: Decimal(str(sheet[cell].value)) for cell in cells
        }
        workbook.close()


def test_drilling_range_citation_resolves_big_foot_rows_and_days() -> None:
    from openpyxl import load_workbook

    references = [
        row for row in _rows() if row["WORKBOOK_ROLE"] == "fingerprint_reference"
    ]
    assert len(references) == 1
    assert references[0]["WORKBOOK_SHEET"] == "Sheet1"
    assert references[0]["WORKBOOK_CELL"] == "Sheet1!A17:L54"

    workbook = load_workbook(
        FDAS / "drilling_and_completion_days.xlsx", read_only=True, data_only=True
    )
    rows = list(
        workbook["Sheet1"].iter_rows(
            min_row=17, max_row=54, min_col=1, max_col=12, values_only=True
        )
    )
    workbook.close()

    assert len(rows) == 38
    assert {row[0] for row in rows} == {"Big Foot"}
    assert sum(row[7] or 0 for row in rows) == 1207
    assert sum(row[8] or 0 for row in rows) == 1826


def test_crosswalk_pins_native_units_and_currencies() -> None:
    located = {
        (row["WORKBOOK_FILE"], row["WORKBOOK_CELL"]): (
            row["VALUE_UNIT"],
            row["CURRENCY"],
        )
        for row in _rows()
        if row["WORKBOOK_CELL"]
    }

    assert located == EXPECTED_UNITS_AND_CURRENCIES


def test_workbook_values_are_assumptions_not_disclosures() -> None:
    value_rows = [row for row in _rows() if row["WORKBOOK_VALUE"]]

    assert value_rows
    assert {row["EVIDENCE_DERIVATION"] for row in value_rows} == {"assumed"}
    assert {row["SOURCE_PROVENANCE"] for row in value_rows} == {"workbook_assumption"}
    assert not any(
        row["EVIDENCE_DERIVATION"] in {"disclosed", "award_derived"}
        for row in value_rows
    )


def test_fdas_total_is_4517_3_mm_and_stale_5200_mm_config_is_not_disclosure() -> None:
    from openpyxl import load_workbook

    path = FDAS / "financial_project_summary.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Project_Summary"]
    bridge = sum(Decimal(str(sheet[cell].value)) for cell in ("J3", "K3", "L3"))
    facilities = sum(
        Decimal(str(sheet[cell].value)) for cell in ("E3", "F3", "G3", "H3", "I3")
    )
    counts_and_days = tuple(
        sheet[cell].value for cell in ("V3", "W3", "X3", "Y3", "Z3", "AA3", "AB3")
    )
    workbook.close()
    assumptions = load_workbook(
        FDAS / "lease_assumptions.xlsx", read_only=True, data_only=True
    )
    injection_facility = Decimal(str(assumptions["assumptions"]["D15"].value))
    assumptions.close()

    additive = [
        Decimal(row["WORKBOOK_VALUE"]) / Decimal("1000000")
        for row in _rows()
        if row["COUNTING_DISPOSITION"] == "included"
    ]
    assert bridge / Decimal("1000000") == Decimal("4517.3")
    assert facilities / Decimal("1000000") + injection_facility == Decimal("2730")
    assert counts_and_days == (7, 1, 0, 0, 38, 1207, 1826)
    assert counts_and_days[-2] + counts_and_days[-1] == 3033
    assert sum(additive) == Decimal("4517.3")
    stale_config = (
        ROOT / "config/analysis/lower_tertiary/fields/big_foot.yml"
    ).read_text(encoding="utf-8")
    assert "5200" in stale_config
    assert not any(
        row["WORKBOOK_VALUE"] == "5200" or row["EVIDENCE_DERIVATION"] == "disclosed"
        for row in _rows()
    )


def test_opex_never_enters_development_capex() -> None:
    opex = [row for row in _rows() if row["WORKBOOK_CATEGORY"] == "OPEX"]

    assert {row["WORKBOOK_CELL"] for row in opex} == {"P3", "Q3", "D16", "D17"}
    assert {row["CAPEX_BASIS"] for row in opex} == {"opex"}
    assert {row["MAPPING_STATUS"] for row in opex} == {"excluded"}
    assert {row["COUNTING_DISPOSITION"] for row in opex} == {"excluded"}
    assert {row["COMPARISON_ELIGIBILITY"] for row in opex} == {"ineligible"}


def test_installation_and_hookup_are_explicitly_unmapped_in_fdas() -> None:
    rows = [row for row in _rows() if row["WORKBOOK_CATEGORY"] == "installation/hookup"]

    assert len(rows) == 1
    assert rows[0]["REQUIREMENT_IDS"] == "req-000007"
    assert rows[0]["WORKBOOK_CELL"] == rows[0]["WORKBOOK_VALUE"] == ""
    assert rows[0]["MAPPING_STATUS"] == "unmapped"
    assert rows[0]["COUNTING_DISPOSITION"] == "excluded"


def test_fdas_crosswalk_preserves_workbook_vintage_and_assumption_status() -> None:
    rows = _rows()

    assert {row["ASSUMPTION_VINTAGE"] for row in rows} == {"FDAS_V30"}
    assert {row["EVIDENCE_DERIVATION"] for row in rows} == {"assumed"}
    assert {row["SOURCE_PROVENANCE"] for row in rows} == {"workbook_assumption"}
    assert {row["MAPPING_STATUS"] for row in rows} <= {
        "mapped",
        "unmapped",
        "excluded",
    }
    assert all(row["PRICE_BASIS"] == "nominal" for row in rows)
    assert all(row["BASIS_YEAR"] == "" for row in rows)


def test_source_workbooks_are_byte_unchanged_after_inspection() -> None:
    from openpyxl import load_workbook

    before = {name: _digest(FDAS / name) for name in FROZEN_HASHES}
    for name in FROZEN_HASHES:
        workbook = load_workbook(FDAS / name, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            next(sheet.iter_rows(values_only=True), ())
        workbook.close()
    after = {name: _digest(FDAS / name) for name in FROZEN_HASHES}

    assert before == after == FROZEN_HASHES


def test_workbook_core_metadata_is_not_published() -> None:
    with CROSSWALK.open(encoding="utf-8", newline="") as stream:
        reader = csv.DictReader(stream)
        rows = list(reader)

    assert tuple(reader.fieldnames or ()) == EXPECTED_COLUMNS
    assert not any(
        token in column.lower()
        for column in EXPECTED_COLUMNS
        for token in ("author", "creator", "modified_by", "lastprinted", "metadata")
    )
    assert not any(
        prefix in value
        for row in rows
        for value in row.values()
        for prefix in ABSOLUTE_PATH_PREFIXES
    )


def test_crosswalk_is_deterministically_ordered() -> None:
    category_order = {
        name: index
        for index, name in enumerate(
            (
                "host",
                "SURF",
                "booster pump",
                "water-injection pump/facility",
                "dry-well system",
                "facilities total",
                "drilling",
                "completion",
                "D&C total",
                "installation/hookup",
                "OPEX",
            )
        )
    }
    role_order = {
        name: index
        for index, name in enumerate(
            (
                "project_summary_component",
                "project_summary_total",
                "project_summary_subtotal",
                "project_summary_opex",
                "assumption_input",
                "fingerprint_reference",
                "coverage_gap",
            )
        )
    }
    rows = _rows()

    def key(row: dict[str, str]) -> tuple[int, int, str, str]:
        return (
            category_order[row["WORKBOOK_CATEGORY"]],
            role_order[row["WORKBOOK_ROLE"]],
            row["WORKBOOK_FILE"],
            row["WORKBOOK_CELL"],
        )

    assert rows == sorted(rows, key=key)
