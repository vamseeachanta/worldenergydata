# Build_Development_Financials_V20.py
# (Formatting-only updates added; economics untouched)

import os, sys, re, warnings, subprocess
import numpy as np
import pandas as pd

# Ensure openpyxl is available
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    print("⚙️ Installing openpyxl …")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=FutureWarning)

LEASES_FP = "leases.xlsx"
ASSUMP_FP = "leases_assumptions.xlsx"
PROD_FP   = "multi_year_lease_matrix_with_charts.xlsx"
DNC_FP    = "drilling_and_completion_days_by_api.xlsx"
WTI_FP    = "wti_full_monthly.xlsx"
OUT_XLSX  = "DEVNAME_Financials_V20.xlsx"

# -------------------- Utilities --------------------

def std_cols(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    return out

def pick_col(df: pd.DataFrame, keys, required: bool = False):
    colsU = [c.upper() for c in df.columns]
    for k in keys:
        if k.upper() in colsU:
            return df.columns[colsU.index(k.upper())]
    if required:
        raise KeyError(f"Missing required column; tried {keys}; have {list(df.columns)}")
    return None

def ym_index(start, end):
    start = pd.Timestamp(start).to_period("M").to_timestamp()
    end   = pd.Timestamp(end).to_period("M").to_timestamp()
    if end < start:
        end = start
    return pd.date_range(start, end, freq="MS")

def first_nonzero_month(series: pd.Series):
    if series is None or series.empty:
        return None
    nz = series[series.fillna(0) > 0]
    return None if nz.empty else pd.to_datetime(nz.index[0])

def normalize_lease_num(val):
    s = str(val).strip().upper()
    m = re.fullmatch(r"G?(\d+)", s)
    return "G" + m.group(1) if m else s

def _get_rate_usd(A: dict, usd_keys, k_keys):
    for k in usd_keys:
        if k in A and pd.notna(A[k]):
            try:
                return float(A[k])
            except Exception:
                pass
    for k in k_keys:
        if k in A and pd.notna(A[k]):
            try:
                return float(A[k]) * 1000.0
            except Exception:
                pass
    return 0.0

# -------------------- Production ingestion --------------------

def _is_matrix_style(df: pd.DataFrame) -> bool:
    cols = [str(c).strip() for c in df.columns]
    has_well = any(c.upper() == "WELL_NAME" for c in cols)
    month_like = [c for c in cols if re.fullmatch(r"\d{4}-\d{2}", str(c))]
    return has_well and len(month_like) >= 3

def _matrix_to_timeseries(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    id_cols = [c for c in df.columns if c.upper() in ("WELL_NAME", "API_WELL_NUMBER")]
    mcols = [c for c in df.columns if re.fullmatch(r"\d{4}-\d{2}", str(c))]
    well_key = "WELL_NAME" if "WELL_NAME" in df.columns else (id_cols[0] if id_cols else None)
    if not well_key:
        raise ValueError("Cannot identify well name column in matrix-style sheet.")
    long = df.melt(id_vars=[well_key], value_vars=mcols, var_name="YearMonth", value_name="BBLS_PER_DAY")
    long["YearMonth"] = pd.to_datetime(long["YearMonth"], format="%Y-%m", errors="coerce")
    long["BBLS_PER_DAY"] = pd.to_numeric(long["BBLS_PER_DAY"], errors="coerce").fillna(0.0)
    dim = long["YearMonth"].dt.days_in_month
    long["OIL_BBL"] = (long["BBLS_PER_DAY"] * dim).astype(float)
    ts = (long.pivot_table(index="YearMonth", columns=well_key, values="OIL_BBL", aggfunc="sum")
            .sort_index().reset_index())
    ts.columns = [c if c == "YearMonth" else str(c).strip() for c in ts.columns]
    return ts

def load_production_by_sheet(xlsx_path: str) -> dict:
    xls = pd.ExcelFile(xlsx_path)
    out = {}
    for sh in xls.sheet_names:
        df = std_cols(xls.parse(sh))
        mcol = pick_col(df, ["MONTH", "DATE"], required=False)
        if mcol:
            df[mcol] = pd.to_datetime(df[mcol], errors="coerce")
            df = df.rename(columns={mcol: "YearMonth"})
            df["YearMonth"] = df["YearMonth"].dt.to_period("M").dt.to_timestamp()
            num_cols = [c for c in df.columns if c != "YearMonth" and pd.api.types.is_numeric_dtype(df[c])]
            out[sh] = pd.concat([df[["YearMonth"]], df[num_cols]], axis=1).dropna(subset=["YearMonth"])
        elif _is_matrix_style(df):
            out[sh] = _matrix_to_timeseries(df)
    return out

def merge_prod_for_dev(prod_by_sheet: dict, leases_df: pd.DataFrame, dev_name: str) -> pd.DataFrame:
    L = leases_df.copy()
    L.columns = [c.upper() for c in L.columns]
    lease_names = set(L.loc[L["DEV_NAME"].astype(str).str.upper() == str(dev_name).strip().upper(),
                           "LEASE_NAME"].dropna().astype(str).str.strip())
    frames = []
    for sh, df in prod_by_sheet.items():
        if str(sh).strip().upper() in {s.upper() for s in lease_names}:
            frames.append(df.copy())
    if not frames:
        if dev_name in prod_by_sheet:
            frames.append(prod_by_sheet[dev_name])
        else:
            return pd.DataFrame(columns=["YearMonth"]).assign(YearMonth=pd.Series([], dtype="datetime64[ns]"))
    prod = frames[0]
    for f in frames[1:]:
        prod = prod.merge(f, on="YearMonth", how="outer")
    prod = prod.sort_values("YearMonth").reset_index(drop=True)
    for c in prod.columns:
        if c != "YearMonth":
            prod[c] = pd.to_numeric(prod[c], errors="coerce").fillna(0.0)
    return prod

# -------------------- D&C ingestion --------------------

def load_dc_both(dnc_path: str, leases_df: pd.DataFrame):
    xls = pd.ExcelFile(dnc_path)
    monthly_frames = []
    totals_frames = []
    for sh in xls.sheet_names:
        df = std_cols(xls.parse(sh))
        if df.empty:
            continue
        ren = {
            "DRILL DAYS":"DRILL_DAYS","DRILLING_DAYS":"DRILL_DAYS","DRILLING DAYS":"DRILL_DAYS",
            "COMP DAYS":"COMP_DAYS","COMPLETION_DAYS":"COMP_DAYS","COMPLETION DAYS":"COMP_DAYS",
            "TOTAL_DEPTH":"TOTAL_DEPTH_DATE","TD_DATE":"TOTAL_DEPTH_DATE"
        }
        for k,v in ren.items():
            if k in df.columns and v not in df.columns:
                df = df.rename(columns={k:v})
        well_col = None
        for cand in ["WELL_NAME","WELL","API_WELL_NUMBER","API"]:
            if cand in df.columns:
                well_col=cand; break
        mcol = pick_col(df, ["YEARMONTH","MONTH","DATE"], required=False)
        # Monthly rows
        have_drill = any(c in df.columns for c in ["DRILL_DAYS","DRILLING_DAYS","DRILL DAYS"])
        have_comp  = any(c in df.columns for c in ["COMP_DAYS","COMPLETION_DAYS","COMP DAYS"])
        if well_col and mcol and have_drill and have_comp:
            df = df.rename(columns={"DRILLING_DAYS":"DRILL_DAYS","DRILL DAYS":"DRILL_DAYS",
                                    "COMPLETION_DAYS":"COMP_DAYS","COMP DAYS":"COMP_DAYS"})
            df[mcol] = pd.to_datetime(df[mcol], errors="coerce")
            df = df.rename(columns={mcol:"YearMonth"})
            df["YearMonth"] = df["YearMonth"].dt.to_period("M").dt.to_timestamp()
            keep = [c for c in ["YearMonth", well_col, "DRILL_DAYS","COMP_DAYS","LEASE_NUM","LEASE_NAME",
                                "DEV_NAME","WELL_SPUD_DATE","TOTAL_DEPTH_DATE"] if c in df.columns]
            monthly_frames.append(df[keep].copy())
        # Totals rows
        if well_col and ("WELL_SPUD_DATE" in df.columns) and ("TOTAL_DEPTH_DATE" in df.columns) and have_drill and have_comp:
            if "DRILL_DAYS" not in df.columns and "DRILLING_DAYS" in df.columns:
                df = df.rename(columns={"DRILLING_DAYS":"DRILL_DAYS"})
            if "COMP_DAYS" not in df.columns and "COMPLETION_DAYS" in df.columns:
                df = df.rename(columns={"COMPLETION_DAYS":"COMP_DAYS"})
            keep = [c for c in [well_col,"WELL_SPUD_DATE","TOTAL_DEPTH_DATE","DRILL_DAYS","COMP_DAYS",
                                "LEASE_NUM","LEASE_NAME","DEV_NAME"] if c in df.columns]
            totals_frames.append(df[keep].copy())
    monthly = pd.concat(monthly_frames, ignore_index=True) if monthly_frames else pd.DataFrame()
    totals  = pd.concat(totals_frames,  ignore_index=True) if totals_frames  else pd.DataFrame()
    for df in (monthly, totals):
        if not df.empty and "LEASE_NUM" in df.columns:
            df["LEASE_NUM"] = df["LEASE_NUM"].map(normalize_lease_num)
    return monthly, totals

def build_day_maps_for_dev(dev_name: str, leases_df: pd.DataFrame, prod_df: pd.DataFrame,
                           monthly: pd.DataFrame, totals: pd.DataFrame):
    L = leases_df.copy(); L.columns = [c.upper() for c in L.columns]
    def _dev_filter(df):
        if df is None or df.empty: return df
        d = df.copy(); d.columns = [c.upper() for c in d.columns]
        if "DEV_NAME" not in d.columns:
            if "LEASE_NUM" in d.columns and "LEASE_NUM" in L.columns:
                d = d.merge(L[["LEASE_NUM","DEV_NAME"]], on="LEASE_NUM", how="left")
            elif "LEASE_NAME" in d.columns and "LEASE_NAME" in L.columns:
                d = d.merge(L[["LEASE_NAME","DEV_NAME"]], on="LEASE_NAME", how="left")
        d = d[d["DEV_NAME"].astype(str).str.strip().str.upper() == str(dev_name).strip().upper()]
        return d
    mon = _dev_filter(monthly); tot = _dev_filter(totals)

    wells = set(); wells_with_totals = set()
    if tot is not None and not tot.empty:
        wk=None
        for cand in ["WELL_NAME","WELL","API_WELL_NUMBER","API"]:
            if cand in tot.columns: wk=cand; break
        if wk:
            wells_with_totals |= set(tot[wk].astype(str).str.strip().tolist())
            wells |= wells_with_totals
    if mon is not None and not mon.empty:
        wk=None
        for cand in ["WELL_NAME","WELL","API_WELL_NUMBER","API"]:
            if cand in mon.columns: wk=cand; break
        if wk: wells |= set(mon[wk].astype(str).str.strip().tolist())
    if "YearMonth" in prod_df.columns:
        wells |= set([c for c in prod_df.columns if c != "YearMonth"])

    fo_by_well = {}
    if "YearMonth" in prod_df.columns:
        pdf = prod_df.set_index("YearMonth")
        for w in wells:
            if w in pdf.columns:
                fo_by_well[w] = first_nonzero_month(pdf[w])

    drill_map = {w: {} for w in wells}
    comp_map  = {w: {} for w in wells}

    def alloc_forward(start_ts, n_days):
        out={}
        if pd.isna(start_ts) or n_days<=0: return out
        start=pd.Timestamp(start_ts).normalize(); days_left=int(round(n_days)); cur=start
        while days_left>0:
            m0=pd.Timestamp(cur.year,cur.month,1)
            m1=m0+pd.offsets.MonthBegin(1)
            m_end=m1-pd.Timedelta(days=1)
            alloc=int(min(days_left,(m_end-cur).days+1))
            ym=m0.to_period("M").to_timestamp()
            out[ym]=out.get(ym,0)+alloc
            days_left-=alloc
            cur=m1
        return out

    def alloc_backward(end_month, n_days):
        out={}
        if end_month is None or n_days<=0: return out
        cur=pd.Timestamp(end_month).to_period("M").to_timestamp()
        cur=cur-pd.offsets.MonthBegin(1)
        days_left=int(round(n_days))
        while days_left>0:
            m0=pd.Timestamp(cur.year,cur.month,1)
            m1=m0+pd.offsets.MonthBegin(1)
            days_in_month=(m1-pd.Timedelta(days=1)).day
            ym=m0.to_period("M").to_timestamp()
            alloc=min(days_left, days_in_month)
            out[ym]=out.get(ym,0)+alloc
            days_left-=alloc
            cur=m0-pd.offsets.MonthBegin(1)
        return out

    # Prefer totals per well; else monthly
    if tot is not None and not tot.empty:
        well_col=None
        for cand in ["WELL_NAME","WELL","API_WELL_NUMBER","API"]:
            if cand in tot.columns: well_col=cand; break
        for _,r in tot.iterrows():
            w=str(r.get(well_col,"")).strip()
            if not w: continue
            spud=pd.to_datetime(r.get("WELL_SPUD_DATE", pd.NaT), errors="coerce")
            td  =pd.to_datetime(r.get("TOTAL_DEPTH_DATE", pd.NaT), errors="coerce")
            dd=float(pd.to_numeric(r.get("DRILL_DAYS",0), errors="coerce") or 0.0)
            cd=float(pd.to_numeric(r.get("COMP_DAYS",0),  errors="coerce") or 0.0)
            for ym,v in alloc_forward(spud, dd).items():
                drill_map[w][ym]=drill_map[w].get(ym,0)+v
            fo=fo_by_well.get(w, None)
            if fo is not None:
                for ym,v in alloc_backward(fo, cd).items():
                    comp_map[w][ym]=comp_map[w].get(ym,0)+v
            else:
                for ym,v in alloc_forward(td, cd).items():
                    comp_map[w][ym]=comp_map[w].get(ym,0)+v

    if mon is not None and not mon.empty:
        well_col=None
        for cand in ["WELL_NAME","WELL","API_WELL_NUMBER","API"]:
            if cand in mon.columns: well_col=cand; break
        for _,r in mon.iterrows():
            w=str(r.get(well_col,"")).strip()
            if pd.isna(r.get("YearMonth", pd.NaT)) or not w or w in wells_with_totals: continue
            m=pd.to_datetime(r.get("YearMonth", pd.NaT), errors="coerce").to_period("M").to_timestamp()
            dd=float(pd.to_numeric(r.get("DRILL_DAYS",0), errors="coerce") or 0.0)
            cd=float(pd.to_numeric(r.get("COMP_DAYS",0),  errors="coerce") or 0.0)
            if dd>0: drill_map[w][m]=drill_map[w].get(m,0)+dd
            if cd>0: comp_map[w][m]=comp_map[w].get(m,0)+cd

    return drill_map, comp_map, list(wells), fo_by_well

# -------------------- WTI / assumptions --------------------

def load_assumptions(xlsx_path: str) -> dict:
    df = std_cols(pd.read_excel(xlsx_path))
    pcol = pick_col(df, ["PARAMETER","PARAM","KEY","NAME"], required=False) or df.columns[0]
    vcol = pick_col(df, ["VALUE","VAL","DATA"], required=False) or df.columns[1]
    P={}
    for _,r in df.iterrows():
        P[str(r[pcol]).strip()] = r[vcol]
    return P

def load_wti_map(path: str, base: float) -> dict:
    if not os.path.exists(path): return {}
    df = std_cols(pd.read_excel(path))
    dcol = pick_col(df, ["DATE","MONTH","YEARMONTH"], required=False) or df.columns[0]
    pcol = pick_col(df, ["WTI","PRICE","WTI_PRICE","WTI_USD_BBL"], required=False) or df.columns[1]
    df[dcol]=pd.to_datetime(df[dcol], errors="coerce")
    df["YearMonth"]=df[dcol].dt.to_period("M").dt.to_timestamp()
    df[pcol]=pd.to_numeric(df[pcol], errors="coerce")
    df=df.dropna(subset=["YearMonth"])
    mp={}
    for _,r in df.iterrows():
        mp[pd.to_datetime(r["YearMonth"]).to_period("M").to_timestamp()] = float(r[pcol]) if pd.notna(r[pcol]) else base
    return mp

# -------------------- Finance helpers --------------------

def npv_from_monthly(cf: np.ndarray, r_month: float) -> float:
    if cf.size == 0:
        return 0.0
    disc = (1.0 + r_month) ** np.arange(cf.size)
    return float(np.sum(cf / disc))

def mirr_from_monthly(cf: np.ndarray, r_fin_month: float, r_reinv_month: float) -> float:
    n = cf.size
    if n <= 1:
        return np.nan
    pos = cf.copy(); pos[pos < 0] = 0.0
    neg = cf.copy(); neg[cf > 0] = 0.0
    fv_pos = 0.0
    for t in range(n):
        if pos[t] > 0:
            fv_pos += pos[t] * ((1.0 + r_reinv_month) ** (n - 1 - t))
    pv_neg = 0.0
    for t in range(n):
        if neg[t] < 0:
            pv_neg += neg[t] / ((1.0 + r_fin_month) ** t)
    if pv_neg >= 0 or fv_pos <= 0:
        return np.nan
    mirr_month = (fv_pos / -pv_neg) ** (1.0 / (n - 1)) - 1.0
    return mirr_month

# -------------------- Main --------------------

def main():
    print(f"Using production workbook: {PROD_FP}")
    leases = std_cols(pd.read_excel(LEASES_FP, sheet_name=0))
    for col in ["LEASE_NAME","DEV_NAME"]:
        if col not in leases.columns:
            raise SystemExit(f"❌ leases.xlsx missing '{col}' column.")
    if "DEV_SYSTEM" not in leases.columns and "DEV_TYPE" not in leases.columns:
        raise SystemExit("❌ leases.xlsx must have DEV_SYSTEM or DEV_TYPE column.")
    leases["DEV_TYPE_EFF"] = leases["DEV_TYPE"] if "DEV_TYPE" in leases.columns else leases["DEV_SYSTEM"]
    if "LEASE_NUM" in leases.columns:
        leases["LEASE_NUM"] = leases["LEASE_NUM"].map(normalize_lease_num)

    A = load_assumptions(ASSUMP_FP)

    MODU_usd=_get_rate_usd(A,
        usd_keys=["MODU_LOADED_DAYRATE_USD","MODU_Loaded_dayrate_USD","MODU_LOADED_DAYRATE"],
        k_keys=["MODU_Loaded_dayrate","MODU_LOADED_DAYRATE_MM"])
    DRY_usd =_get_rate_usd(A,
        usd_keys=["DRY_LOADED_DAYRATE_USD","Dry_Loaded_dayrate_USD","DRY_LOADED_DAYRATE"],
        k_keys=["Dry_Loaded_dayrate","DRY_LOADED_DAYRATE_MM"])
    print(f"Dayrates used — MODU: ${MODU_usd:,.0f}/day, DRY: ${DRY_usd:,.0f}/day")

    flat_wti=float(A.get("WTI_base_$/bbl",75.0))
    royalty=float(A.get("Royalty_Rate",0.0))
    severance=float(A.get("Severance_Tax_Rate",0.0))
    subsea_opex=float(A.get("Variable_OPEX_Subsea_$/bbl",16.0))
    dry_opex=float(A.get("Variable_OPEX_Dry_$/bbl",10.0))
    if "Variable_OPEX_$/bbl" in A:
        subsea_opex=float(A["Variable_OPEX_$/bbl"])
        dry_opex=float(A.get("Variable_OPEX_Dry_$/bbl",subsea_opex))
    fixed_opex_mm=float(A.get("Fixed_OPEX_MM_per_year",0.0))
    fixed_opex_usd=fixed_opex_mm*1_000_000.0 if fixed_opex_mm else 0.0
    corp_tax=float(A.get("Corporate_Tax_Rate",0.21))

    host_subsea_mm=float(A.get("Host_SUBSEA_MM",0.0))
    host_dry_mm=float(A.get("Host_DRY_MM",0.0))
    surf_per_well_mm=float(A.get("SURF_per_well_MM",0.0))
    host_prefo=int(A.get("Host_PreFO_Months",24))
    surf_prefo=int(A.get("SURF_PreFO_Months",12))
    subsea_pump_mm=float(A.get("Subsea_Pump_MM",A.get("Pump_pkg_per_5_wells_MM",0.0)))

    # Dry well systems per producer (USD)
    dry_sys_per_prod_usd = A.get("Dry_Well_System_Per_Producer_USD", None)
    if dry_sys_per_prod_usd is not None and pd.notna(dry_sys_per_prod_usd):
        try:
            dry_sys_per_prod_usd = float(dry_sys_per_prod_usd)
        except Exception:
            dry_sys_per_prod_usd = 0.0
    else:
        dry_sys_per_prod_mm = float(A.get("Dry_Well_System_Per_Producer_MM",
                                     A.get("Dry_Well_System_Per_Producer",
                                           A.get("DRY_WELL_SYSTEM_PER_PRODUCER_MM",0.0))))
        dry_sys_per_prod_usd = dry_sys_per_prod_mm * 1_000_000.0
    print(f"Dry Well System per producer (USD): {dry_sys_per_prod_usd:,.0f}")

    # Discount / finance / reinvestment used in CALCULATIONS (unchanged logic)
    disc_ann = float(A.get("Discount_Rate_Annual", A.get("NPV_Discount_Rate", 0.10)))

    # Prefer explicit MIRR_* keys; fall back to Reinvest/Finance, then to discount, for CALC
    reinv_ann = float(A.get("MIRR_Reinvest_Rate_annual", A.get("Reinvest_Rate_Annual", disc_ann)))
    fin_ann   = float(A.get("MIRR_Finance_Rate_annual",  A.get("Finance_Rate_Annual",  disc_ann)))

    r_d = (1.0 + disc_ann) ** (1.0/12.0) - 1.0
    r_f = (1.0 + fin_ann ) ** (1.0/12.0) - 1.0
    r_r = (1.0 + reinv_ann) ** (1.0/12.0) - 1.0

    prod_by_sheet=load_production_by_sheet(PROD_FP)
    dnc_monthly, dnc_totals = load_dc_both(DNC_FP, leases)

    if not dnc_monthly.empty:
        global_start = pd.to_datetime(dnc_monthly["YearMonth"]).min()
    elif not dnc_totals.empty:
        global_start = pd.to_datetime(dnc_totals["WELL_SPUD_DATE"]).min()
    else:
        global_start = pd.Timestamp("2000-01-01")

    wti_map=load_wti_map(WTI_FP, base=flat_wti)
    price_model_used = "Monthly WTI file" if len(wti_map)>0 else f"Flat WTI ${flat_wti:.2f}/bbl"

    per_dev_outputs=[]
    proj_rows=[]
    exec_rows=[]
    qc_rows=[]
    qc_dev_rows=[]

    def build_input_totals_map(df_totals, df_monthly):
        res={}
        if df_totals is not None and not df_totals.empty:
            well_col=None
            for cand in ["WELL_NAME","WELL","API_WELL_NUMBER","API"]:
                if cand in df_totals.columns: well_col=cand; break
            for _,r in df_totals.iterrows():
                w=str(r.get(well_col,"")).strip()
                if not w: continue
                dd=float(pd.to_numeric(r.get("DRILL_DAYS",0), errors="coerce") or 0.0)
                cd=float(pd.to_numeric(r.get("COMP_DAYS",0),  errors="coerce") or 0.0)
                spud=pd.to_datetime(r.get("WELL_SPUD_DATE", pd.NaT), errors="coerce")
                td  =pd.to_datetime(r.get("TOTAL_DEPTH_DATE", pd.NaT), errors="coerce")
                res[w]=(dd,cd,spud,td)
        if df_monthly is not None and not df_monthly.empty:
            well_col=None
            for cand in ["WELL_NAME","WELL","API_WELL_NUMBER","API"]:
                if cand in df_monthly.columns: well_col=cand; break
            grp=(df_monthly.groupby(well_col, dropna=True)
                 .agg(DRILL_DAYS=pd.NamedAgg(column="DRILL_DAYS", aggfunc="sum"),
                      COMP_DAYS=pd.NamedAgg(column="COMP_DAYS", aggfunc="sum"))
                 ).reset_index()
            for _,r in grp.iterrows():
                w=str(r[well_col]).strip()
                if not w: continue
                if w not in res:
                    res[w]=(float(r["DRILL_DAYS"] or 0.0), float(r["COMP_DAYS"] or 0.0), pd.NaT, pd.NaT)
        return res

    input_totals_all = build_input_totals_map(dnc_totals, dnc_monthly)

    for dev_name, g in leases.groupby("DEV_NAME"):
        system=str(g["DEV_TYPE_EFF"].iloc[0]).strip().lower()
        var_opex = subsea_opex if system=="subsea" else dry_opex

        prod_df = merge_prod_for_dev(prod_by_sheet, leases, dev_name)
        if prod_df.empty:
            prod_df = pd.DataFrame({"YearMonth": ym_index(global_start, global_start)})

        drill_map, comp_map, all_wells, fo_by_well = build_day_maps_for_dev(
            dev_name, leases, prod_df, dnc_monthly, dnc_totals
        )

        # FO dates
        fo_candidates=[]
        if "YearMonth" in prod_df.columns:
            for c in [col for col in prod_df.columns if col != "YearMonth"]:
                m=first_nonzero_month(prod_df.set_index("YearMonth")[c])
                if m is not None:
                    fo_candidates.append(m)
        fo_dev = min(fo_candidates) if fo_candidates else None

        latest_prod = prod_df["YearMonth"].max() if "YearMonth" in prod_df.columns and not prod_df.empty else global_start
        months_alloc=[]
        for m2d in list(drill_map.values())+list(comp_map.values()):
            months_alloc.extend(list(m2d.keys()))
        earliest_alloc = min(months_alloc) if months_alloc else global_start
        latest_dc = max(months_alloc) if months_alloc else None
        overall_end = max([x for x in [latest_prod, latest_dc] if x is not None]) if (latest_prod or latest_dc) else global_start
        overall_start = min(global_start, earliest_alloc)
        idx=ym_index(overall_start, overall_end)

        out = pd.DataFrame(index=idx)
        out.index.name="YearMonth"

        oil_cols=[]
        if "YearMonth" in prod_df.columns:
            oil_cols=[c for c in prod_df.columns if c!="YearMonth"]
            out = out.join(prod_df.set_index("YearMonth")[oil_cols], how="left")
            out[oil_cols]=out[oil_cols].fillna(0.0)

        drill_tot=pd.Series(0.0, index=idx)
        comp_tot =pd.Series(0.0, index=idx)
        for w in sorted(all_wells):
            dcol=f"DRILL_DAYS::{w}"
            ccol=f"COMP_DAYS::{w}"
            out[dcol]=0.0; out[ccol]=0.0
            for m,v in drill_map.get(w, {}).items():
                if m in out.index: out.loc[m,dcol]+=float(v)
            for m,v in comp_map.get(w, {}).items():
                if m in out.index: out.loc[m,ccol]+=float(v)
            drill_tot += out[dcol]; comp_tot += out[ccol]
        out["DRILL_DAYS_TOT"]=drill_tot.values
        out["COMP_DAYS_TOT"]=comp_tot.values

        out["WTI_Price"]=out.index.map(lambda m: wti_map.get(m, flat_wti))
        out["Gross_Oil_bbls"]=out[oil_cols].sum(axis=1) if oil_cols else 0.0
        rev_gross=out["Gross_Oil_bbls"]*out["WTI_Price"]
        royalty_amt=rev_gross*royalty
        severance_amt=rev_gross*severance
        out["Revenue_Adjustments"]=0.0
        out["Revenue_Net"]=rev_gross-royalty_amt-severance_amt+out["Revenue_Adjustments"]

        out["OPEX_Var"]=out["Gross_Oil_bbls"]*var_opex
        out["OPEX_Fixed"]=(fixed_opex_usd/12.0) if fixed_opex_usd>0 else 0.0
        out["OPEX"]=out["OPEX_Var"]+out["OPEX_Fixed"]

        out["CAPEX_Drill"]=0.0
        out["CAPEX_Comp"]=0.0
        for m in out.index:
            rate = MODU_usd if (system=="subsea" or (fo_dev is None or m<fo_dev)) else DRY_usd
            out.loc[m,"CAPEX_Drill"] = out.loc[m,"DRILL_DAYS_TOT"] * rate
            out.loc[m,"CAPEX_Comp"]  = out.loc[m,"COMP_DAYS_TOT"]  * rate

        host_mm = host_subsea_mm if system=="subsea" else host_dry_mm
        host_usd=host_mm*1_000_000.0
        out["CAPEX_Facilities"]=0.0

        producer_cols=[c for c in oil_cols if out[c].sum()>0]
        producers_count=len(producer_cols)
        injector_count=(producers_count//7) if (system=="subsea" and producers_count>=7) else 0

        # SURF totals (subsea)
        surf_usd_total=0.0
        if system=="subsea":
            surf_usd_total += surf_per_well_mm*1_000_000.0*producers_count
            if injector_count>0:
                surf_usd_total += injector_count*surf_per_well_mm*1_000_000.0

        # Host allocation
        if fo_dev is not None and host_usd>0:
            total_host_months=max(host_prefo+6,1)
            win=pd.date_range(fo_dev-pd.offsets.DateOffset(months=total_host_months),
                              fo_dev-pd.offsets.DateOffset(months=1), freq="MS")
            if len(win)>0:
                per_m=host_usd/len(win)
                for mm in win:
                    if mm in out.index: out.loc[mm,"CAPEX_Facilities"] += per_m

        # SURF allocation
        if system=="subsea" and surf_usd_total>0 and surf_prefo>0 and fo_dev is not None:
            win=pd.date_range(fo_dev-pd.offsets.DateOffset(months=surf_prefo),
                              fo_dev-pd.offsets.DateOffset(months=1), freq="MS")
            if len(win)>0:
                per_m=surf_usd_total/len(win)
                for mm in win:
                    if mm in out.index: out.loc[mm,"CAPEX_Facilities"] += per_m

        # Dry Well Systems: per producer, 12 months pre FO (USD)
        dry_sys_total_usd=0.0
        if system!="subsea" and dry_sys_per_prod_usd>0 and producers_count>0:
            for c in producer_cols:
                fm=first_nonzero_month(out[c])
                if fm is not None:
                    alloc=dry_sys_per_prod_usd
                    dry_sys_total_usd += alloc
                    win=pd.date_range(fm-pd.offsets.DateOffset(months=12),
                                      fm-pd.offsets.DateOffset(months=1), freq="MS")
                    if len(win)>0:
                        per_m=alloc/len(win)
                        for mm in win:
                            if mm in out.index: out.loc[mm,"CAPEX_Facilities"] += per_m

        # Subsea pumps: 1 per 8 producers
        pump_count=(producers_count//8) if (system=="subsea" and producers_count>=8) else 0
        pump_total_usd = pump_count * subsea_pump_mm * 1_000_000.0
        anchor_fo=None
        if pump_count>0:
            prod_fo_sorted = sorted([m for m in [first_nonzero_month(out[c]) for c in producer_cols] if m is not None])
            anchor_fo = prod_fo_sorted[7] if len(prod_fo_sorted)>=8 else fo_dev
        if pump_total_usd>0 and anchor_fo is not None and surf_prefo>0:
            win=pd.date_range(anchor_fo-pd.offsets.DateOffset(months=surf_prefo),
                              anchor_fo-pd.offsets.DateOffset(months=1), freq="MS")
            if len(win)>0:
                per_m=pump_total_usd/len(win)
                for mm in win:
                    if mm in out.index: out.loc[mm,"CAPEX_Facilities"] += per_m

        # CAPEX & cash flow
        out["CAPEX"]=out["CAPEX_Drill"]+out["CAPEX_Comp"]+out["CAPEX_Facilities"]
        out["CAPEX_Drill_Net"]=out["CAPEX_Drill"]*(1.0-corp_tax)
        out["CAPEX_Comp_Net"]=out["CAPEX_Comp"]*(1.0-corp_tax)
        out["CAPEX_Facilities_Net"]=out["CAPEX_Facilities"]*(1.0-corp_tax)
        out["Tax_Savings"]=out["CAPEX"]*corp_tax
        out["Net_Cash_Flow"]=out["Revenue_Net"]-out["OPEX"]-out["CAPEX"]+out["Tax_Savings"]
        out["Cum_Cash_Flow"]=out["Net_Cash_Flow"].cumsum()

        total_oil=float(out["Gross_Oil_bbls"].sum())
        total_fac=float(out["CAPEX_Facilities"].sum())
        total_drill=float(out["CAPEX_Drill"].sum())
        total_comp=float(out["CAPEX_Comp"].sum())
        total_dnc=total_drill+total_comp
        total_opex=float(out["OPEX"].sum())

        if fo_dev is not None:
            pre_mask=out.index<fo_dev; post_mask=out.index>=fo_dev
        else:
            pre_mask=out.index<out.index.min(); post_mask=~pre_mask
        dnc_pre_gross=float(out.loc[pre_mask,["CAPEX_Drill","CAPEX_Comp"]].sum().sum())
        dnc_post_gross=float(out.loc[post_mask,["CAPEX_Drill","CAPEX_Comp"]].sum().sum())
        dnc_pre_net=dnc_pre_gross*(1.0-corp_tax)
        dnc_post_net=dnc_post_gross*(1.0-corp_tax)

        # Finance metrics (monthly; economics unchanged)
        cf = out["Net_Cash_Flow"].values.astype(float)
        npv10 = npv_from_monthly(cf, r_d)
        mirr_m = mirr_from_monthly(cf, r_f, r_r)
        mirr_ann = (1.0 + mirr_m) ** 12 - 1.0 if np.isfinite(mirr_m) else np.nan

        # Save DEV sheet
        per_dev_outputs.append((str(dev_name)[:31], out.reset_index()))

        # QC (per-well)
        for w in sorted(set(list(drill_map.keys())+list(comp_map.keys()))):
            dd_alloc=float(sum(drill_map.get(w,{}).values()))
            cd_alloc=float(sum(comp_map.get(w,{}).values()))
            dd_in,cd_in,spud,td=input_totals_all.get(w,(np.nan,np.nan,pd.NaT,pd.NaT))
            fo=fo_by_well.get(w,None)
            qc_rows.append({
                "DEV_NAME":dev_name,"WELL":w,"SPUD":spud,"TD":td,"FO_MONTH":fo,
                "Drill Days — Input":dd_in,"Drill Days — Alloc":dd_alloc,
                "Drill Days — Diff":(dd_alloc-dd_in) if pd.notna(dd_in) else np.nan,
                "Comp Days — Input":cd_in,"Comp Days — Alloc":cd_alloc,
                "Comp Days — Diff":(cd_alloc-cd_in) if pd.notna(cd_in) else np.nan
            })

        # Project Summary row
        proj_rows.append({
            "Project Name":dev_name,
            "TOTAL OIL BBL":total_oil,
            "Facilities Host USD":float(host_usd),
            "Facilities SURF USD":float(surf_usd_total if system=="subsea" else 0.0),
            "Subsea Pumps USD":float(pump_total_usd if system=="subsea" else 0.0),
            "Dry Well Systems USD":float(dry_sys_total_usd if system!="subsea" else 0.0),
            "Facilities Cost USD":float(total_fac),
            "DnC Drill Total USD":total_drill,
            "DnC Comp Total USD":total_comp,
            "DnC Total USD":total_dnc,
            "OPEX Total USD":total_opex,
            "Producer Wells Used":int(len(producer_cols)),
            "Injector Wells":int((len(producer_cols)//7) if (system=='subsea' and len(producer_cols)>=7) else 0),
            "Pumps Count (8/prod)":int((len(producer_cols)//8) if (system=='subsea' and len(producer_cols)>=8) else 0),
            "DEV SYSTEM USED":system,
            "ECON PATH USED":('dry: MODU->DRY at FO' if system!='subsea' else 'subsea: MODU'),
            "FO Month":fo_dev
        })

        # Executive Summary row
        exec_rows.append({
            "Project Name":dev_name,
            "TOTAL OIL BBL":total_oil,
            "Facilities Cost USD":float(total_fac),
            "DnC Drill Total USD":total_drill,
            "DnC Comp Total USD":total_comp,
            "DnC Total USD":total_dnc,
            "DnC Total PreFO USD":dnc_pre_gross,
            "DnC Total PostFO USD":dnc_post_gross,
            "DnC Net PreFO USD":dnc_pre_net,
            "DnC Net PostFO USD":dnc_post_net,
            "NPV10 afterTax":npv10,
            "MIRR afterTax":mirr_ann,
            "Price Model Used":price_model_used
        })

        # QC (per-dev)
        wells_in_dev=[row["WELL"] for row in qc_rows if row["DEV_NAME"]==dev_name]
        dd_in_dev=float(np.nansum([input_totals_all.get(w,(np.nan,np.nan,np.nan,np.nan))[0] for w in wells_in_dev]))
        cd_in_dev=float(np.nansum([input_totals_all.get(w,(np.nan,np.nan,np.nan,np.nan))[1] for w in wells_in_dev]))
        dd_alloc_dev=float(sum([sum(drill_map.get(w,{}).values()) for w in wells_in_dev]))
        cd_alloc_dev=float(sum([sum(comp_map.get(w,{}).values()) for w in wells_in_dev]))
        qc_dev_rows.append({"DEV_NAME":dev_name,
                            "Drill Days — Input (Dev)":dd_in_dev,
                            "Drill Days — Alloc (Dev)":dd_alloc_dev,
                            "Drill Days — Diff (Dev)":dd_alloc_dev-dd_in_dev,
                            "Comp Days — Input (Dev)":cd_in_dev,
                            "Comp Days — Alloc (Dev)":cd_alloc_dev,
                            "Comp Days — Diff (Dev)":cd_alloc_dev-cd_in_dev})

    import datetime as _dt
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        # placeholder README (will be overwritten by formatter)
        readme = pd.DataFrame({
            "README — Financial Analysis V20 Output Workbook":[None,"File Name:","Generated By:","Version:","Last Updated:"],
            "":[None,OUT_XLSX,"Build_Development_Financials_V20.py","V20",_dt.datetime.now().strftime("%B %d, %Y")]
        })
        readme.to_excel(writer, sheet_name="README", index=False)

        exec_df=pd.DataFrame(exec_rows, columns=[
            "Project Name","TOTAL OIL BBL","Facilities Cost USD",
            "DnC Drill Total USD","DnC Comp Total USD","DnC Total USD",
            "DnC Total PreFO USD","DnC Total PostFO USD",
            "DnC Net PreFO USD","DnC Net PostFO USD",
            "NPV10 afterTax","MIRR afterTax","Price Model Used"
        ])
        exec_df.to_excel(writer, sheet_name="Executive Summary (V20)", index=False)

        proj_df=pd.DataFrame(proj_rows, columns=[
            "Project Name","TOTAL OIL BBL","Facilities Host USD","Facilities SURF USD","Subsea Pumps USD",
            "Dry Well Systems USD","Facilities Cost USD","DnC Drill Total USD","DnC Comp Total USD",
            "DnC Total USD","OPEX Total USD","Producer Wells Used","Injector Wells","Pumps Count (8/prod)",
            "DEV SYSTEM USED","ECON PATH USED","FO Month"
        ])
        proj_df.to_excel(writer, sheet_name="Project Summary (V20)", index=False)

        qc_df=pd.DataFrame(qc_rows, columns=[
            "DEV_NAME","WELL","SPUD","TD","FO_MONTH",
            "Drill Days — Input","Drill Days — Alloc","Drill Days — Diff",
            "Comp Days — Input","Comp Days — Alloc","Comp Days — Diff"
        ])
        qc_df.to_excel(writer, sheet_name="QC — D&C Allocation", index=False)

        qc_dev_df=pd.DataFrame(qc_dev_rows, columns=[
            "DEV_NAME","Drill Days — Input (Dev)","Drill Days — Alloc (Dev)","Drill Days — Diff (Dev)",
            "Comp Days — Input (Dev)","Comp Days — Alloc (Dev)","Comp Days — Diff (Dev)"
        ])
        qc_dev_df.to_excel(writer, sheet_name="QC — Dev Totals", index=False)

        for sh_name, df in per_dev_outputs:
            df.to_excel(writer, sheet_name=sh_name, index=False)

    print(f"✅ Wrote {OUT_XLSX}")

    # ---- Post-processing formatting (no economics changed) ----
    _apply_v20_formatting(out_name=OUT_XLSX, discount_rate_annual=disc_ann)

# -------------------- Formatting helpers (post-write only) --------------------

def _header_map(ws):
    h = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            h[str(v).strip()] = c
    return h

def _wrap_header(ws):
    for cell in ws[1]:
        cell.alignment = Alignment(wrap_text=True)

def _set_all_col_width(ws, width=15.0):
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = width

def _insert_columns_after(ws, after_header, headers_to_insert):
    hmap = _header_map(ws)
    if after_header in hmap:
        insert_pos = hmap[after_header] + 1
        ws.insert_cols(insert_pos, amount=len(headers_to_insert))
    else:
        insert_pos = ws.max_column + 1
    for i, name in enumerate(headers_to_insert):
        ws.cell(row=1, column=insert_pos + i, value=name)
    return insert_pos

def _collect_exec_values_by_project(ws, headers):
    hmap = _header_map(ws)
    if "Project Name" not in hmap:
        return {}
    proj_col = hmap["Project Name"]
    out = {}
    for r in range(2, ws.max_row + 1):
        pname = ws.cell(row=r, column=proj_col).value
        if pname is None:
            continue
        vals = {}
        for h in headers:
            cidx = hmap.get(h)
            vals[h] = ws.cell(row=r, column=cidx).value if cidx else None
        out[str(pname)] = vals
    return out

def _append_rate_columns(ws, rates):
    start_col = ws.max_column + 1
    cols = ["Discount Rate (Annual)", "MIRR Reinvest Rate (Annual)", "MIRR Finance Rate (Annual)"]
    for i, name in enumerate(cols):
        c = start_col + i
        ws.cell(row=1, column=c, value=name)
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c, value=rates.get(name))

def _populate_readme_sheet(wb, out_name):
    name = "README"
    if name in wb.sheetnames:
        ws = wb[name]
        for row in ws["A1:Z200"]:
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(name, 0)

    from datetime import datetime as _dt
    ws["A1"] = "README — Financial Analysis V20 Output Workbook"
    ws["A2"] = "Generated On"; ws["B2"] = _dt.now().strftime("%B %d, %Y %H:%M")
    ws["A3"] = "File Name";   ws["B3"] = out_name

    ws["A5"] = "Overview"
    ws["B5"] = (
        "This workbook contains per-development financial results produced by the V20 engine. "
        "It aggregates monthly oil production, drilling & completion (DnC) day allocations, facilities CAPEX, and OPEX "
        "to construct after-tax cash flows per development. Economics (NPV and MIRR) are computed from those monthly cash flows."
    )

    ws["A7"] = "Inputs (files)"
    ws["B7"] = "\n".join([
        "- leases.xlsx (DEV_NAME, LEASE_NAME/NUM, DEV_TYPE/DEV_SYSTEM)",
        "- leases_assumptions.xlsx (rates, costs, taxes)",
        "- multi_year_lease_matrix_with_charts.xlsx (monthly production by lease/well)",
        "- drilling_and_completion_days_by_api.xlsx (DnC per-well totals/monthlies)",
        "- wti_full_monthly.xlsx (optional monthly WTI; otherwise flat WTI base is used)",
    ])

    ws["A10"] = "Calculations (summary)"
    ws["B10"] = "\n".join([
        "• Revenue: monthly gross oil barrels × WTI, minus royalty & severance.",
        "• OPEX: variable per-barrel + fixed monthly.",
        "• CAPEX: DnC days × dayrate (MODU or DRY per system/FO timing), plus facilities allocations (host, SURF, pumps, dry well systems).",
        "• Taxes: corporate tax savings applied to CAPEX in-month.",
        "• Net Cash Flow: Revenue_Net − OPEX − CAPEX + Tax_Savings.",
        "• NPV10 & MIRR: computed from monthly cash flows; MIRR uses finance & reinvestment rates.",
    ])

    ws["A15"] = "Outputs (tabs)"
    ws["B15"] = "\n".join([
        "- Executive Summary (V20): high-level per-development metrics (with NPV/MIRR and rate columns).",
        "- Project Summary (V20): detailed CAPEX/OPEX breakdown, wells & equipment counts, FO month, and DnC Pre/Post-FO metrics.",
        "- QC — D&C Allocation / QC — Dev Totals: reconciliation tables of input vs allocated DnC days.",
        "- <DEV_NAME> tabs: month-by-month cash flow build, WTI, OPEX, CAPEX components, and cumulative cash flow.",
    ])

    for cell in ws[1]:
        cell.alignment = Alignment(wrap_text=True)
    for r in (5, 7, 10, 15):
        ws[f"B{r}"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 120

def _read_mirr_rates_from_assumptions(path="leases_assumptions.xlsx"):
    """Read Inputs_Assumptions!B6 (Reinvest) and B7 (Finance). If the sheet isn't there, use active."""
    try:
        wb = load_workbook(path, data_only=True)
        sh = wb["Inputs_Assumptions"] if "Inputs_Assumptions" in wb.sheetnames else wb.active
        reinvest = sh["B6"].value  # MIRR_Reinvest_Rate_annual
        finance  = sh["B7"].value  # MIRR_Finance_Rate_annual
        return reinvest, finance
    except Exception:
        return None, None

def _format_sheet_numbers(ws, percent_headers=None, skip_headers=None):
    """Format numeric cells: numbers -> '#,##0'; rates -> '0.00%'; dates (FO Month) -> 'mmm-yy'."""
    percent_headers = set(h.strip() for h in (percent_headers or []))
    skip_headers = set(h.strip() for h in (skip_headers or []))
    h = _header_map(ws)
    for header, col in h.items():
        if header in skip_headers:
            continue
        if header in percent_headers:
            numfmt = "0.00%"
        elif header == "FO Month":
            numfmt = "mmm-yy"
        else:
            numfmt = "#,##0"
        # apply only to numeric cells (date col handled by header name)
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            v = cell.value
            if header == "FO Month":
                cell.number_format = numfmt
            elif isinstance(v, (int, float)) and v is not None:
                cell.number_format = numfmt

def _format_numbers_all(wb):
    # Executive Summary
    if "Executive Summary (V20)" in wb.sheetnames:
        ws = wb["Executive Summary (V20)"]
        _format_sheet_numbers(
            ws,
            percent_headers={
                "Discount Rate (Annual)",
                "MIRR Reinvest Rate (Annual)",
                "MIRR Finance Rate (Annual)",
                "MIRR afterTax",
            },
            skip_headers={"Project Name", "Price Model Used"}  # text columns
        )
    # Project Summary
    if "Project Summary (V20)" in wb.sheetnames:
        ws = wb["Project Summary (V20)"]
        _format_sheet_numbers(
            ws,
            percent_headers=set(),
            skip_headers={"Project Name", "DEV SYSTEM USED", "ECON PATH USED"}  # FO Month formatted by name
        )
    # QC sheets
    for name in wb.sheetnames:
        if str(name).startswith("QC"):
            ws = wb[name]
            _format_sheet_numbers(ws, percent_headers=set(),
                                  skip_headers={"DEV_NAME","WELL","SPUD","TD","FO_MONTH"})

def _apply_v20_formatting(out_name, discount_rate_annual):
    wb = load_workbook(out_name)

    # 1) README page
    _populate_readme_sheet(wb, out_name)

    exec_sheet = "Executive Summary (V20)"
    proj_sheet = "Project Summary (V20)"
    move_headers = [
        "DnC Total PreFO USD",
        "DnC Total PostFO USD",
        "DnC Net PreFO USD",
        "DnC Net PostFO USD",
    ]

    # 2) Exec Summary: wrap + width 15 + add 3 rate columns
    reinvest_b6, finance_b7 = _read_mirr_rates_from_assumptions(ASSUMP_FP)
    rates = {
        "Discount Rate (Annual)": discount_rate_annual,                   # mirrors calc input
        "MIRR Reinvest Rate (Annual)": reinvest_b6 if reinvest_b6 is not None else discount_rate_annual,
        "MIRR Finance Rate (Annual)":  finance_b7 if finance_b7 is not None else discount_rate_annual,
    }

    if exec_sheet in wb.sheetnames:
        ws_exec = wb[exec_sheet]
        _wrap_header(ws_exec)
        _set_all_col_width(ws_exec, 15.0)
        _append_rate_columns(ws_exec, rates)

    # 3) Project Summary: wrap; set A..O width 14.0; move Exec G–J to Project (after 'DnC Total USD')
    if proj_sheet in wb.sheetnames and exec_sheet in wb.sheetnames:
        ws_proj = wb[proj_sheet]
        ws_exec = wb[exec_sheet]
        _wrap_header(ws_proj)

        # Set Project Summary columns A..O to width 14.0
        for idx in range(1, 16):  # A=1..O=15
            ws_proj.column_dimensions[get_column_letter(idx)].width = 14.0

        exec_map = _collect_exec_values_by_project(ws_exec, move_headers)
        start_col = _insert_columns_after(ws_proj, "DnC Total USD", move_headers)
        hmap_proj = _header_map(ws_proj)
        proj_name_col = hmap_proj.get("Project Name", 1)
        for r in range(2, ws_proj.max_row + 1):
            pname = ws_proj.cell(row=r, column=proj_name_col).value
            if pname is None:
                continue
            vals = exec_map.get(str(pname), {})
            for i, h in enumerate(move_headers):
                ws_proj.cell(row=r, column=start_col + i, value=vals.get(h))

        # delete the moved columns from Exec (right-to-left)
        hmap_exec = _header_map(ws_exec)
        to_delete = sorted([hmap_exec[h] for h in move_headers if h in hmap_exec], reverse=True)
        for col_idx in to_delete:
            ws_exec.delete_cols(col_idx, 1)

    # 4) Wrap headers on QC sheets
    for name in wb.sheetnames:
        if str(name).startswith("QC"):
            _wrap_header(wb[name])

    # 5) Number formats (commas, no decimals; % for rates; mmm-yy for FO Month)
    _format_numbers_all(wb)

    wb.save(out_name)
    print(f"✅ V20 formatting applied to {out_name}")

if __name__ == "__main__":
    main()
