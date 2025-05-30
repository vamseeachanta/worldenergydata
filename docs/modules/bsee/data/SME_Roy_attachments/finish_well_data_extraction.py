import pandas as pd
import csv
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def extract_remarks_by_api():
    mapping_file = "api_snwar_pairs.csv"
    remark_file = "mv_war_main_prop_remark.txt"
    output_file = "api_remarks_output.xlsx"

    try:
        mapping_df = pd.read_csv(mapping_file, dtype=str)
        mapping_df['SN_WAR'] = mapping_df['SN_WAR'].apply(lambda x: str(abs(int(str(x).strip()))) if str(x).strip().lstrip('-').isdigit() else '')
        mapping_df = mapping_df[mapping_df['SN_WAR'] != '']
    except Exception as e:
        print(f"❌ Failed to read API-SN_WAR mapping: {e}")
        return

    snwar_to_api = mapping_df.set_index('SN_WAR')['API_WELL_NUMBER'].to_dict()
    remarks_by_api = defaultdict(list)

    try:
        with open(remark_file, 'r', encoding='ISO-8859-1') as f:
            reader = csv.DictReader(f, delimiter=',', quotechar='"')
            for row in reader:
                raw_snwar = row.get('SN_WAR', '').strip()
                try:
                    snwar = str(abs(int(raw_snwar)))
                except:
                    continue
                if snwar in snwar_to_api:
                    api = snwar_to_api[snwar]
                    text = row.get('TEXT_REMARK', '').strip()
                    if not text:
                        continue

                    for line in text.splitlines():
                        match = re.match(r'^[A-Za-z]{2} (\d{1,2}/\d{1,2}/\d{2}) - (.+)', line)
                        if match:
                            date_str, remark = match.groups()
                            remarks_by_api[api].append((date_str, remark.strip()[:500]))

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "API Remarks"

        # Headers
        ws.append(["API Well Number", "Date", "Narrative Remark"])

        # Populate data by API
        for api, entries in remarks_by_api.items():
            # Sort by parsed date for each API
            try:
                sorted_entries = sorted(entries, key=lambda x: datetime.strptime(x[0], "%m/%d/%y"))
                start_date = datetime.strptime(sorted_entries[0][0], "%m/%d/%y")
                end_date = datetime.strptime(sorted_entries[-1][0], "%m/%d/%y")
                duration = (end_date - start_date).days
            except:
                sorted_entries = entries
                duration = ""

            for date_str, remark in sorted_entries:
                ws.append([api, date_str, remark])

            # Insert duration row
            ws.append(["", f"{duration} days", ""])
            # Insert 2 additional blank rows
            ws.append(["", "", ""])
            ws.append(["", "", ""])

        # Adjust column widths
        ws.column_dimensions[get_column_letter(1)].width = 16
        ws.column_dimensions[get_column_letter(2)].width = 10
        ws.column_dimensions[get_column_letter(3)].width = 100

        # Save file
        wb.save(output_file)
        print(f"✅ Extracted remarks to {output_file} with duration rows and spacing")

    except Exception as e:
        print(f"❌ Failed to process remark file: {e}")

if __name__ == "__main__":
    extract_remarks_by_api()








