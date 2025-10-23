#!/usr/bin/env python3
# generate_financial_summary_V30.py -- trimmed NPV/MIRR (Excel-like) + dry-tree 1e6 fix

import os, math, re, warnings
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore', category=FutureWarning)

def month_floor(ts):
    if pd.isna(ts): return ts
    return pd.Timestamp(ts).to_period('M').to_timestamp()

def ym_index(start, end):
    s = month_floor(start); e = month_floor(end)
    if pd.isna(s) or pd.isna(e): return pd.DatetimeIndex([])
    if e < s: e = s
    return pd.date_range(s, e, freq='MS')

def safe_float(x, default=0.0):
    try:
        v = float(x)
        if np.isnan(v): return default
        return v
    except Exception:
        return default

def norm_dev_system(s):
    if pd.isna(s): return 'unknown'
    return str(s).strip().lower().replace(' ','')

def norm_name(s):
    if pd.isna(s): return s
    s = str(s).strip()
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('St. Malo','St Malo')
    return s

def read_xlsx_any(fp):
    xl = pd.ExcelFile(fp)
    cands = []
    for sh in xl.sheet_names:
        df = xl.parse(sh); df.columns=[str(c).strip() for c in df.columns]
        cands.append((sh, df))
    return cands

def load_leases(fp: str) -> pd.DataFrame:
    sh, df = read_xlsx_any(fp)[0]
    cols = {c.upper(): c for c in df.columns}
    def pick(*names):
        for n in names:
            if n in cols: return cols[n]
        raise RuntimeError(f'{fp} missing one of: {names}')
    out = pd.DataFrame({
        'LEASE_NAME': df[pick('LEASE_NAME','LEASE','LEASE TITLE','LEASE_ID')].astype(str).map(norm_name),
        'DEV_NAME':   df[pick('DEV_NAME','DEV NAME','DEVELOPMENT','PROJECT NAME','PROJECT')].astype(str).map(norm_name),
        'DEV_SYSTEM': df[pick('DEV_SYSTEM','SYSTEM','DEV TYPE','DEVTYPE')].map(norm_dev_system),
    }).drop_duplicates(subset=['LEASE_NAME'])
    return out

def load_assumptions_fixed(fp: str) -> pd.DataFrame:
    df = pd.read_excel(fp, sheet_name=0)
    df.columns = [str(c).strip().upper() for c in df.columns]
    if 'DEV_SYSTEM' in df.columns and len(set(df.columns) - {'DEV_SYSTEM','NOTES'}) >= 2:
        df = df.rename(columns={'DEV_SYSTEM':'METRIC'})
        systems = [c for c in df.columns if c not in {'METRIC','NOTES'}]
        long = df.melt(id_vars=['METRIC'], var_name='DEV_SYSTEM', value_name='VAL', value_vars=systems)
        long['DEV_SYSTEM'] = long['DEV_SYSTEM'].str.lower().str.replace(' ','', regex=False)
        long['METRIC'] = long['METRIC'].str.upper()
        wide = long.pivot_table(index='DEV_SYSTEM', columns='METRIC', values='VAL', aggfunc='last').reset_index()
        for c in wide.columns:
            if c != 'DEV_SYSTEM':
                wide[c] = pd.to_numeric(wide[c], errors='coerce')
        return wide
    if 'DEV_SYSTEM' in df.columns:
        df['DEV_SYSTEM'] = df['DEV_SYSTEM'].str.lower().str.replace(' ','', regex=False)
        for c in df.columns:
            if c != 'DEV_SYSTEM':
                df[c] = pd.to_numeric(df[c], errors='coerce')
        return df
    return pd.DataFrame([{'DEV_SYSTEM':'default'}])

ASSUMP_ALIASES = {
    'HOST_CAPEX_MM':['HOST_CAPEX_MM'],
    'SURF_PER_WELL_MM':['SURF_PER_WELL_MM'],
    'PUMP_WELLS_PER_PKG':['PUMP_WELLS_PER_PKG'],
    'DRY_WELL_SYSTEM_PER_PRODUCER_USD':['DRY_WELL_SYSTEM_PER_PRODUCER_USD'],
    'MODU_LOADED_DAYRATE_MM':['MODU_LOADED_DAYRATE_MM'],
    'DRY_TREE_RIG_RATE_MM':['DRY_TREE_RIG_RATE_MM'],
    'ROYALTY_RATE':['ROYALTY_RATE'],
    'VARIABLE_OPEX_$/BBL':['VARIABLE_OPEX_$/BBL'],
    'FIXED_OPEX_MM_PER_YEAR':['FIXED_OPEX_MM_PER_YEAR'],
    'WTI_BASE_$/BBL':['WTI_BASE_$/BBL'],
    'INJECTORS_PER_PRODUCER':['INJECTORS_PER_PRODUCER'],
    'HOST_PREFO_MONTHS':['HOST_PREFO_MONTHS'],
    'DISCOUNT_RATE_ANNUAL':['DISCOUNT_RATE_ANNUAL'],
    'BOOSTER_PUMP_15K_MM':['BOOSTER_PUMP_15K_MM'],
    'BOOSTER_PUMP_20K_MM':['BOOSTER_PUMP_20K_MM'],
    'BOOSTER_PUMP_TRIGGER_PRODUCERS':['BOOSTER_PUMP_TRIGGER_PRODUCERS'],
    'WATER_INJECTION_PUMP_MM':['WATER_INJECTION_PUMP_MM'],
    'WATER_INJECTION_TRIGGER_PRODUCERS':['WATER_INJECTION_TRIGGER_PRODUCERS'],
    'WATER_INJECTION_FACILITY_COST_MM':['WATER_INJECTION_FACILITY_COST_MM'],
}

def Aget(A: pd.DataFrame, sys_name, key, default=0.0):
    sysn = norm_dev_system(sys_name); keyU = key.upper()
    if 'DEV_SYSTEM' not in A.columns: return default
    if keyU in A.columns:
        m = (A['DEV_SYSTEM']==sysn); 
        if not m.any(): m = (A['DEV_SYSTEM']=='default')
        if m.any():
            v = A.loc[m, keyU].iloc[0]
            try: return float(v) if pd.notna(v) else default
            except Exception: return default
    for alias in ASSUMP_ALIASES.get(keyU, []):
        if alias in A.columns and alias != keyU:
            m = (A['DEV_SYSTEM']==sysn); 
            if not m.any(): m = (A['DEV_SYSTEM']=='default')
            if m.any():
                v = A.loc[m, alias].iloc[0]
                try: return float(v) if pd.notna(v) else default
                except Exception: return default
    return default

def load_production(fp: str) -> pd.DataFrame:
    chosen=None
    for sh, df in read_xlsx_any(fp):
        up={c.upper():c for c in df.columns}
        if not all(n in up for n in ['DEV_NAME','API_WELL_NUMBER']): continue
        if ('MONTH' in up) and (('MONTHLY_OIL_VOLUME' in up) or ('TOTAL_OIL_BBL' in up)):
            chosen=(sh,df); break
    if not chosen: raise RuntimeError(f'{fp} missing required columns')
    _, df = chosen
    cols = {c.upper(): c for c in df.columns}
    df = df.rename(columns={cols['DEV_NAME']: 'DEV_NAME',
                            cols['API_WELL_NUMBER']: 'API_WELL_NUMBER',
                            cols['MONTH']: 'MONTH'})
    if 'MONTHLY_OIL_VOLUME' in cols:
        df = df.rename(columns={cols['MONTHLY_OIL_VOLUME']: 'MONTHLY_OIL_VOLUME'})
    elif 'TOTAL_OIL_BBL' in cols:
        df = df.rename(columns={cols['TOTAL_OIL_BBL']: 'MONTHLY_OIL_VOLUME'})
    df['DEV_NAME'] = df['DEV_NAME'].map(norm_name)
    df['API_WELL_NUMBER'] = pd.to_numeric(df['API_WELL_NUMBER'], errors='coerce')
    df['MONTH'] = pd.to_datetime(df['MONTH'], errors='coerce').dt.to_period('M').dt.to_timestamp()
    df['MONTHLY_OIL_VOLUME'] = pd.to_numeric(df['MONTHLY_OIL_VOLUME'], errors='coerce').fillna(0.0)
    return df

def load_wti(fp: str, A: pd.DataFrame) -> pd.DataFrame:
    if not os.path.exists(fp): return pd.DataFrame(columns=['MONTH','WTI_USD'])
    sh, df = read_xlsx_any(fp)[0]
    df.columns = [c.upper() for c in df.columns]
    price_col = next((c for c in ['WTI_USD','WTI_PRICE','PRICE'] if c in df.columns), None)
    out = df[['MONTH', price_col]].copy()
    out['MONTH'] = pd.to_datetime(out['MONTH'], errors='coerce').dt.to_period('M').dt.to_timestamp()
    out = out.rename(columns={price_col: 'WTI_USD'})
    out['WTI_USD'] = pd.to_numeric(out['WTI_USD'], errors='coerce').fillna(0.0)
    return out

def load_dnc(fp: str, leases: pd.DataFrame) -> pd.DataFrame:
    sh, df = read_xlsx_any(fp)[0]
    df.columns = [c.strip().upper() for c in df.columns]
    if 'API' in df.columns and 'API_WELL_NUMBER' not in df.columns: df = df.rename(columns={'API':'API_WELL_NUMBER'})
    if 'DRILL DAYS' in df.columns and 'DRILLING_DAYS' not in df.columns: df = df.rename(columns={'DRILL DAYS':'DRILLING_DAYS'})
    if 'COMP DAYS' in df.columns and 'COMPLETION_DAYS' not in df.columns: df = df.rename(columns={'COMP DAYS':'COMPLETION_DAYS'})
    if 'TD_DATE' in df.columns and 'TOTAL_DEPTH_DATE' not in df.columns: df = df.rename(columns={'TD_DATE':'TOTAL_DEPTH_DATE'})
    if 'TOTAL DEPTH DATE' in df.columns and 'TOTAL_DEPTH_DATE' not in df.columns: df = df.rename(columns={'TOTAL DEPTH DATE':'TOTAL_DEPTH_DATE'})
    need = ['LEASE_NAME','API_WELL_NUMBER','WELL_SPUD_DATE','DRILLING_DAYS','COMPLETION_DAYS']
    for n in need:
        if n not in df.columns: raise RuntimeError('dnc missing '+n)
    df['LEASE_NAME'] = df['LEASE_NAME'].astype(str).map(norm_name)
    df['API_WELL_NUMBER'] = pd.to_numeric(df['API_WELL_NUMBER'], errors='coerce')
    df['WELL_SPUD_DATE'] = pd.to_datetime(df['WELL_SPUD_DATE'], errors='coerce')
    if 'TOTAL_DEPTH_DATE' in df.columns:
        df['TOTAL_DEPTH_DATE'] = pd.to_datetime(df['TOTAL_DEPTH_DATE'], errors='coerce')
    else:
        df['TOTAL_DEPTH_DATE'] = pd.NaT
    j = df.merge(leases[['LEASE_NAME','DEV_NAME','DEV_SYSTEM']], on='LEASE_NAME', how='left')
    j['DEV_NAME'] = j['DEV_NAME'].map(norm_name); j['DEV_SYSTEM'] = j['DEV_SYSTEM'].map(norm_dev_system)
    return j

def excel_like_mirr(cf: np.ndarray, r_ann: float) -> float:
    nz = np.where(np.abs(cf) > 1e-6)[0]
    if nz.size == 0: return np.nan
    cf = cf[nz[0]:nz[-1]+1]
    if not (np.any(cf > 0) and np.any(cf < 0)): return np.nan
    n = cf.size - 1
    r = (1.0 + r_ann)**(1/12) - 1.0
    fv_pos = sum(cf[t] * ((1.0 + r) ** (n - t)) for t in range(cf.size) if cf[t] > 0)
    pv_neg = sum(cf[t] / ((1.0 + r) ** t) for t in range(cf.size) if cf[t] < 0)
    if pv_neg >= 0 or fv_pos <= 0: return np.nan
    return (fv_pos / -pv_neg) ** (1.0 / n) - 1.0

def generate(leases_xlsx, assump_xlsx, prod_xlsx, dnc_xlsx, wti_xlsx, out_xlsx):
    leases = load_leases(leases_xlsx)
    A      = load_assumptions_fixed(assump_xlsx)
    prod   = load_production(prod_xlsx)
    wti    = load_wti(wti_xlsx, A)
    dnc    = load_dnc(dnc_xlsx, leases)

    for df in (leases, A, prod, dnc):
        if 'DEV_NAME' in df.columns:   df['DEV_NAME'] = df['DEV_NAME'].map(norm_name)
        if 'DEV_SYSTEM' in df.columns: df['DEV_SYSTEM'] = df['DEV_SYSTEM'].map(norm_dev_system)

    fopw = (prod[prod['MONTHLY_OIL_VOLUME']>0]
            .sort_values(['API_WELL_NUMBER','MONTH'])
            .groupby('API_WELL_NUMBER')['MONTH'].min().to_dict())
    fo_by_dev = (prod[prod['MONTHLY_OIL_VOLUME']>0]
                 .groupby('DEV_NAME')['MONTH'].min().reset_index().rename(columns={'MONTH':'FO_MONTH'}))

    oil_sum = prod.groupby('DEV_NAME')['MONTHLY_OIL_VOLUME'].sum().reset_index(name='TOTAL OIL BBL')
    producer_wells = (prod[prod['MONTHLY_OIL_VOLUME']>0]
                      .groupby('DEV_NAME')['API_WELL_NUMBER'].nunique()
                      .reset_index(name='PRODUCER_WELLS'))
    total_wellbores = dnc.groupby('DEV_NAME')['API_WELL_NUMBER'].nunique().reset_index(name='TOTAL_WELLBORES')
    dev_sys = leases[['DEV_NAME','DEV_SYSTEM']].drop_duplicates()

    base = (oil_sum
            .merge(dev_sys, on='DEV_NAME', how='left')
            .merge(fo_by_dev, on='DEV_NAME', how='left')
            .merge(producer_wells, on='DEV_NAME', how='left')
            .merge(total_wellbores, on='DEV_NAME', how='left')
            .fillna({'PRODUCER_WELLS':0,'TOTAL_WELLBORES':0}))
    base['DEV_SYSTEM'] = base['DEV_SYSTEM'].fillna('unknown')

    g_start = min([x for x in [
        prod['MONTH'].min() if len(prod) else pd.NaT,
        dnc['WELL_SPUD_DATE'].min() if len(dnc) else pd.NaT,
        dnc['TOTAL_DEPTH_DATE'].min() if len(dnc) else pd.NaT,
    ] if pd.notna(x)])
    g_end   = max([x for x in [
        prod['MONTH'].max() if len(prod) else pd.NaT,
        dnc['WELL_SPUD_DATE'].max() if len(dnc) else pd.NaT,
        dnc['TOTAL_DEPTH_DATE'].max() if len(dnc) else pd.NaT,
    ] if pd.notna(x)])
    if pd.isna(g_start) or pd.isna(g_end):
        raise RuntimeError('Cannot establish a global timeline')

    proj_rows = []; dev_sheets = {}

    dnc_totals = (dnc.groupby('DEV_NAME')[['DRILLING_DAYS','COMPLETION_DAYS']]
                  .sum(min_count=1).reset_index())
    spud_td = (dnc.groupby('DEV_NAME')[['WELL_SPUD_DATE','TOTAL_DEPTH_DATE']]
               .agg({'WELL_SPUD_DATE':'min','TOTAL_DEPTH_DATE':'max'}).reset_index())

    WTI_base_default = Aget(A, 'default','WTI_BASE_$/BBL', 75.0)

    for dev in base['DEV_NAME'].dropna().unique():
        sys_name = base.loc[base['DEV_NAME']==dev,'DEV_SYSTEM'].iloc[0]
        fo       = base.loc[base['DEV_NAME']==dev,'FO_MONTH'].iloc[0] if 'FO_MONTH' in base.columns else pd.NaT

        pm = prod[prod['DEV_NAME']==dev][['MONTH','API_WELL_NUMBER','MONTHLY_OIL_VOLUME']].copy()
        sp_min = spud_td.loc[spud_td['DEV_NAME']==dev,'WELL_SPUD_DATE']
        dev_start = min([x for x in [sp_min.iloc[0] if len(sp_min) else pd.NaT, pm['MONTH'].min() if len(pm) else pd.NaT, g_start] if pd.notna(x)])
        idx = pd.DataFrame({'MONTH': ym_index(dev_start, g_end)})

        monthly = idx.merge(pm.groupby('MONTH')['MONTHLY_OIL_VOLUME'].sum().reset_index(), on='MONTH', how='left').fillna({'MONTHLY_OIL_VOLUME':0.0})
        if not wti.empty:
            monthly = monthly.merge(wti, on='MONTH', how='left')
            monthly['WTI_USD'] = monthly['WTI_USD'].fillna(WTI_base_default)
        else:
            monthly['WTI_USD'] = WTI_base_default

        dev_dnc = dnc[dnc['DEV_NAME']==dev].copy()
        drill_rows, comp_rows = [], []
        for _, r in dev_dnc.iterrows():
            dd, cd = safe_float(r['DRILLING_DAYS']), safe_float(r['COMPLETION_DAYS'])
            sp, td = r['WELL_SPUD_DATE'], r['TOTAL_DEPTH_DATE']
            api    = r['API_WELL_NUMBER']
            fow    = fopw.get(api, pd.NaT)
            m_d = {}
            if pd.notna(sp) and dd>0:
                cur = pd.Timestamp(sp).normalize(); left=int(round(dd))
                while left>0:
                    m0 = pd.Timestamp(cur.year, cur.month, 1); m1 = m0 + pd.offsets.MonthBegin(1)
                    alloc = min(left, (m1 - cur).days)
                    m_d[m0.to_period('M').to_timestamp()] = m_d.get(m0.to_period('M').to_timestamp(), 0) + alloc
                    left -= alloc; cur = m1
            for m,v in m_d.items(): drill_rows.append((m,v))
            # completion backwards to FO or TD
            m_c = {}
            end_month = fow if pd.notna(fow) else fo if pd.notna(fo) else pd.NaT
            if pd.notna(end_month) and cd>0:
                cur_m0 = pd.Timestamp(end_month).to_period('M').to_timestamp()
                cur = cur_m0 + pd.offsets.MonthEnd(0); left=int(round(cd))
                while left>0:
                    m0 = pd.Timestamp(cur.year, cur.month, 1)
                    days_in_month = (m0 + pd.offsets.MonthBegin(1) - pd.Timedelta(days=1)).day
                    alloc = min(left, days_in_month)
                    m_c[m0.to_period('M').to_timestamp()] = m_c.get(m0.to_period('M').to_timestamp(), 0) + alloc
                    left -= alloc; cur = m0 - pd.Timedelta(days=1)
            elif pd.notna(td) and cd>0:
                cur = pd.Timestamp(td).normalize(); left=int(round(cd))
                while left>0:
                    m0 = pd.Timestamp(cur.year, cur.month, 1); m1 = m0 + pd.offsets.MonthBegin(1)
                    alloc = min(left, (m1 - cur).days)
                    m_c[m0.to_period('M').to_timestamp()] = m_c.get(m0.to_period('M').to_timestamp(), 0) + alloc
                    left -= alloc; cur = m1
            for m,v in m_c.items(): comp_rows.append((m,v))

        dnc_m = pd.DataFrame(drill_rows, columns=['MONTH','DRILLING_DAYS']).groupby('MONTH', as_index=False)['DRILLING_DAYS'].sum() if drill_rows else pd.DataFrame(columns=['MONTH','DRILLING_DAYS'])
        cmp_m = pd.DataFrame(comp_rows,  columns=['MONTH','COMPLETION_DAYS']).groupby('MONTH', as_index=False)['COMPLETION_DAYS'].sum() if comp_rows else pd.DataFrame(columns=['MONTH','COMPLETION_DAYS'])
        monthly = monthly.merge(dnc_m, on='MONTH', how='left').merge(cmp_m, on='MONTH', how='left').fillna({'DRILLING_DAYS':0.0,'COMPLETION_DAYS':0.0})

        prod_by_api = (prod[(prod['DEV_NAME']==dev)].groupby('API_WELL_NUMBER')['MONTHLY_OIL_VOLUME'].sum())
        producer_cnt = int((prod_by_api > 0).sum())
        inj = int(math.floor(producer_cnt * Aget(A, sys_name,'INJECTORS_PER_PRODUCER', 0.2)))

        sysn = norm_dev_system(sys_name); sys_is_dry=(sysn=='dry'); sys_is_sub20=(sysn=='subsea20')

        r_modu_sub15 = Aget(A, 'subsea15','MODU_LOADED_DAYRATE_MM', 0.0)
        r_modu_sys   = Aget(A, sys_name,'MODU_LOADED_DAYRATE_MM', 0.0)
        r_dry_comp   = Aget(A, 'dry','DRY_TREE_RIG_RATE_MM', 0.0)

        def drill_rate_mm(m):
            if sys_is_dry: return r_modu_sub15
            if sys_is_sub20 and pd.notna(fo) and (m < month_floor(fo)): return r_modu_sub15
            return r_modu_sys
        def comp_rate_mm(m):
            if sys_is_dry: return r_dry_comp
            return r_modu_sys

        monthly['DRILL_RATE_MM'] = monthly['MONTH'].apply(drill_rate_mm)
        monthly['COMP_RATE_MM']  = monthly['MONTH'].apply(comp_rate_mm)
        monthly['DRILLING_COST_USD']   = monthly['DRILLING_DAYS']   * monthly['DRILL_RATE_MM'] * 1_000_000.0
        monthly['COMPLETION_COST_USD'] = monthly['COMPLETION_DAYS'] * monthly['COMP_RATE_MM']  * 1_000_000.0
        monthly['DnC_Total_USD']       = monthly['DRILLING_COST_USD'] + monthly['COMPLETION_COST_USD']

        Host_MM       = Aget(A, sys_name,'HOST_CAPEX_MM',0.0)
        Host_Pre_M    = int(safe_float(Aget(A, sys_name,'HOST_PREFO_MONTHS',0)))
        SURF_per_w_MM = Aget(A, sys_name,'SURF_PER_WELL_MM',0.0) if not sys_is_dry else 0.0

        booster_15k_MM = Aget(A, sys_name, 'BOOSTER_PUMP_15K_MM', 0.0)
        booster_20k_MM = Aget(A, sys_name, 'BOOSTER_PUMP_20K_MM', 0.0)
        booster_trig   = int(max(1, safe_float(Aget(A, sys_name,'BOOSTER_PUMP_TRIGGER_PRODUCERS', 99))))
        wi_pump_MM     = Aget(A, sys_name, 'WATER_INJECTION_PUMP_MM', 0.0)
        wi_trig        = int(max(1, safe_float(Aget(A, sys_name,'WATER_INJECTION_TRIGGER_PRODUCERS', 99))))
        wi_fac_MM      = Aget(A, sys_name, 'WATER_INJECTION_FACILITY_COST_MM', 0.0)

        booster_cost_MM = booster_20k_MM if sys_is_sub20 else booster_15k_MM
        booster_count = 0; wi_pump_count = 0
        if pd.notna(fo) and not sys_is_dry:
            booster_count = int(min(2, math.floor(producer_cnt / booster_trig))) if booster_trig>0 else 0
            wi_pump_count = int(min(2, math.floor(producer_cnt / wi_trig))) if wi_trig>0 and inj>0 else 0

        Pump_Package_USD = booster_count * booster_cost_MM * 1_000_000.0
        Water_Injection_Pump_USD = wi_pump_count * wi_pump_MM * 1_000_000.0

        Dry_per_prod_MM = Aget(A, sys_name,'DRY_WELL_SYSTEM_PER_PRODUCER_USD',0.0)
        Dry_Well_System_USD = (Dry_per_prod_MM * producer_cnt * 1_000_000.0) if (sys_is_dry and Dry_per_prod_MM>0 and pd.notna(fo)) else 0.0

        monthly['Host_CAPEX_USD'] = 0.0
        if pd.notna(fo) and Host_MM>0 and Host_Pre_M>0:
            start = month_floor(fo) - pd.DateOffset(months=Host_Pre_M)
            mask  = (monthly['MONTH'] >= start) & (monthly['MONTH'] < month_floor(fo))
            monthly.loc[mask,'Host_CAPEX_USD'] = (Host_MM * 1_000_000.0) / Host_Pre_M

        SURF_USD = 0.0
        if pd.notna(fo) and SURF_per_w_MM>0:
            SURF_USD = SURF_per_w_MM * (producer_cnt + inj) * 1_000_000.0

        monthly['Facilities_USD'] = monthly['Host_CAPEX_USD']
        if pd.notna(fo):
            fmask = (monthly['MONTH'] == month_floor(fo))
            monthly.loc[fmask, 'Facilities_USD'] += (SURF_USD + Pump_Package_USD + Water_Injection_Pump_USD + Dry_Well_System_USD + wi_fac_MM*1_000_000.0)

        Royalty_Rate     = Aget(A, sys_name,'ROYALTY_RATE',0.0)
        Var_Opex_per_bbl = Aget(A, sys_name,'VARIABLE_OPEX_$/BBL',0.0)
        Fixed_OPEX_Yr_MM = Aget(A, sys_name,'FIXED_OPEX_MM_PER_YEAR',0.0)

        monthly['Revenue_USD'] = monthly['MONTHLY_OIL_VOLUME'] * monthly['WTI_USD']
        monthly['Royalty_USD'] = monthly['Revenue_USD'] * Royalty_Rate
        monthly['Variable_OPEX_USD'] = monthly['MONTHLY_OIL_VOLUME'] * Var_Opex_per_bbl
        monthly['Fixed_OPEX_USD'] = 0.0
        if pd.notna(fo) and Fixed_OPEX_Yr_MM>0:
            monthly.loc[monthly['MONTHLY_OIL_VOLUME']>0, 'Fixed_OPEX_USD'] = (Fixed_OPEX_Yr_MM*1_000_000.0)/12.0

        monthly['Net_CashFlow_USD'] = (monthly['Revenue_USD'] - monthly['Royalty_USD']
                                       - monthly['Variable_OPEX_USD'] - monthly['Fixed_OPEX_USD']
                                       - monthly['Facilities_USD'] - monthly['DnC_Total_USD'])

        # Trim to first..last non-zero month for NPV/MIRR
        cf = monthly['Net_CashFlow_USD'].values.astype(float)
        nz = np.where(np.abs(cf) > 1e-6)[0]
        if nz.size > 0:
            monthly_eval = monthly.iloc[nz[0]:nz[-1]+1].copy()
        else:
            monthly_eval = monthly.iloc[:0].copy()

        Disc_ann = Aget(A, sys_name,'DISCOUNT_RATE_ANNUAL',0.10)
        r_m = (1.0 + Disc_ann) ** (1.0/12.0) - 1.0
        monthly_eval = monthly_eval.reset_index(drop=True)
        monthly_eval['t'] = monthly_eval.index.astype(float)
        monthly_eval['Disc_Factor'] = (1.0 + r_m) ** monthly_eval['t']
        monthly_eval['PV_CashFlow_USD'] = np.where(monthly_eval['Disc_Factor']!=0,
                                                   monthly_eval['Net_CashFlow_USD']/monthly_eval['Disc_Factor'], 0.0)
        cf_eval = monthly_eval['Net_CashFlow_USD'].values.astype(float)
        mirr_m = excel_like_mirr(cf_eval, Disc_ann)
        mirr_a = (1.0 + mirr_m) ** 12 - 1.0 if pd.notna(mirr_m) else np.nan

        sums = monthly_eval[['Net_CashFlow_USD','PV_CashFlow_USD']].sum(numeric_only=True)

        proj_rows.append({
            'Project Name': dev,
            'DEV_SYSTEM': sysn,
            'FO Month': month_floor(fo),
            'TOTAL OIL BBL': float(monthly['MONTHLY_OIL_VOLUME'].sum()),
            'Host_CAPEX_USD': float(monthly['Host_CAPEX_USD'].sum()),
            'SURF_USD': float(SURF_USD if pd.notna(fo) else 0.0),
            'Booster_Pump_USD': float(Pump_Package_USD if pd.notna(fo) else 0.0),
            'Water_Injection_Pump_USD': float(Water_Injection_Pump_USD if pd.notna(fo) else 0.0),
            'Dry_Well_System_USD': float(Dry_Well_System_USD if pd.notna(fo) else 0.0),
            'Facilities Cost USD': float(monthly['Facilities_USD'].sum()),
            'DRILLING_COST_USD': float(monthly['DRILLING_COST_USD'].sum()),
            'COMPLETION_COST_USD': float(monthly['COMPLETION_COST_USD'].sum()),
            'DnC_Total_USD': float(monthly['DnC_Total_USD'].sum()),
            'Revenue_USD': float(monthly['Revenue_USD'].sum()),
            'Royalty_USD': float(monthly['Royalty_USD'].sum()),
            'Variable_OPEX_USD': float(monthly['Variable_OPEX_USD'].sum()),
            'Fixed_OPEX_USD': float(monthly['Fixed_OPEX_USD'].sum()),
            'Net_CashFlow_USD': float(sums['Net_CashFlow_USD']),
            'NPV_USD': float(sums['PV_CashFlow_USD']),
            'MIRR_monthly': mirr_m,
            'MIRR_annual': mirr_a,
        })

        dev_sheets[dev] = monthly[['MONTH','WTI_USD','MONTHLY_OIL_VOLUME','Revenue_USD','Royalty_USD',
                                   'Variable_OPEX_USD','Fixed_OPEX_USD','DRILLING_DAYS','DRILLING_COST_USD',
                                   'COMPLETION_DAYS','COMPLETION_COST_USD','Facilities_USD','Net_CashFlow_USD']]

    proj_df = pd.DataFrame(proj_rows)

    with pd.ExcelWriter(out_xlsx, engine='openpyxl') as w:
        proj_df.to_excel(w, sheet_name='Project_Summary', index=False)
        for dev, df in dev_sheets.items():
            safe_name = (str(dev) or 'DEV').strip()[:31]
            df.to_excel(w, sheet_name=safe_name, index=False)

    wb = load_workbook(out_xlsx)
    ws = wb['Project_Summary']
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws)))
    for c in ws[1]: c.alignment = Alignment(wrap_text=True)
    headers = [c.value for c in ws[1]]
    cur_cols = {'Host_CAPEX_USD','SURF_USD','Booster_Pump_USD','Water_Injection_Pump_USD','Dry_Well_System_USD',
                'Facilities Cost USD','DRILLING_COST_USD','COMPLETION_COST_USD','DnC_Total_USD',
                'Revenue_USD','Royalty_USD','Variable_OPEX_USD','Fixed_OPEX_USD','Net_CashFlow_USD','NPV_USD'}
    pct_cols = {'MIRR_monthly','MIRR_annual'}
    date_cols = {'FO Month'}
    int_cols  = {'TOTAL OIL BBL'}
    for j, h in enumerate(headers, start=1):
        col = get_column_letter(j)
        if h in cur_cols: fmt = '"$"#,##0'
        elif h in pct_cols: fmt = '0.0%'
        elif h in date_cols: fmt = 'm/d/yy'
        elif h in int_cols: fmt = '#,##0'
        else: fmt = None
        if fmt:
            for r in range(2, ws.max_row+1):
                ws[f'{col}{r}'].number_format = fmt
        ws.column_dimensions[col].width = max(16, len(str(h))+2)
    wb.save(out_xlsx)

if __name__ == '__main__':
    base = Path('.')
    generate(str(base/'leases.xlsx'),
             str(base/'lease_assumptions.xlsx'),
             str(base/'chronological_lease_analysis.xlsx'),
             str(base/'drilling_and_completion_days.xlsx'),
             str(base/'wti_monthly.xlsx'),
             str(base/'financial_project_summary.xlsx'))
